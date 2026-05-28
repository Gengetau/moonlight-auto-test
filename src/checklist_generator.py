import argparse
import html
import json
import sys
from dataclasses import dataclass, replace
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

# Allows both `python -m src.checklist_generator` and direct `python src/checklist_generator.py`.
ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from src.coverage_policy import CHECKLIST_SECTIONS, CASE_DEPTH  # noqa: E402
try:  # noqa: E402
    from src.page_case_planner import PageCasePlanner
except ImportError:  # pragma: no cover - direct script execution fallback
    from page_case_planner import PageCasePlanner  # type: ignore


SEVERITY_ORDER = {"High": 0, "Medium": 1, "Low": 2}
CASE_GENERATING_KINDS = {"page", "form", "field", "select", "textarea", "hidden", "file", "button", "link", "scenario"}
FIELD_LIKE_KINDS = {"field", "select", "textarea"}
AUTO_ACTION_TYPES = {
    "snapshot",
    "wait",
    "click",
    "submit",
    "navigate",
    "download",
    "upload",
    "upload_submit",
    "fill",
    "select",
    "clear",
    "set_value",
    "press",
    "close_window",
}
RUNTIME_PROFILE_SCHEMA = "moonlight.runtime_page_profile.v1"
DEFAULT_RUNTIME_PROFILE_DIR = Path("generated/valid/runtime_profile")


@dataclass(frozen=True)
class TestCase:
    # Positional arguments are intentionally ordered as:
    # title, objective, steps, expected, severity.
    # Page/element metadata is supplied by **fields in the case builders.
    title: str
    objective: str
    steps: str
    expected: str
    severity: str
    page: str
    kind: str
    locator: str
    line: str
    evidence: str
    automation_mode: str = "manual/assist"
    case_type: str = ""
    action_type: str = ""
    test_data: str = ""
    submit_locator: str = ""
    expected_type: str = ""
    expected_value: str = ""
    pre_steps: str = ""
    main_step: str = ""
    case_id: str = ""
    parent_case_id: str = ""
    viewpoint_id: str = ""
    enabled: str = "true"
    generated_by: str = ""
    matched_capabilities: str = ""
    destructive: str = "false"
    priority: str = ""


def as_text(value: Any, default: str = "") -> str:
    if value is None:
        return default
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value)


def as_json(value: Any) -> str:
    if value in (None, "", [], {}):
        return ""
    try:
        return json.dumps(value, ensure_ascii=False)
    except Exception:
        return as_text(value)


def first_attr(attributes: Dict[str, Any], *names: str) -> str:
    lowered = {str(key).lower(): value for key, value in attributes.items()}
    for name in names:
        if name in attributes:
            return as_text(attributes[name])
        value = lowered.get(name.lower())
        if value is not None:
            return as_text(value)
    return ""


def page_entries(scan_data: Dict[str, Any]) -> Iterable[Dict[str, Any]]:
    if scan_data.get("schema") == RUNTIME_PROFILE_SCHEMA or isinstance(scan_data.get("controls"), list):
        yield scan_data
        return
    if isinstance(scan_data.get("runtime_profiles"), list):
        yield from scan_data["runtime_profiles"]
        return
    if "pages" in scan_data:
        yield from scan_data.get("pages", [])
        return
    if isinstance(scan_data.get("page_mappings"), list):
        yield from scan_data["page_mappings"]
        return
    yielded_mapping_subset = False
    for key in ["high_risk_pages", "medium_risk_pages"]:
        if isinstance(scan_data.get(key), list):
            yield from scan_data[key]
            yielded_mapping_subset = True
    if yielded_mapping_subset:
        return
    yield {
        "source": scan_data.get("source", scan_data.get("root", "<unknown>")),
        "counts": scan_data.get("counts", {}),
        "elements": scan_data.get("elements", []),
    }


def page_name_of(page: Dict[str, Any]) -> str:
    return as_text(
        page.get("page_id")
        or page.get("target_page_name")
        or page.get("target_page")
        or page.get("source")
        or page.get("legacy_sources", [None])[0],
        "<unknown>",
    )


def page_match_keys(value: Any) -> set:
    text = as_text(value).replace("\\", "/").strip().strip("'\"").lower()
    if not text:
        return set()
    leaf = text.rsplit("/", 1)[-1]
    stem = leaf.rsplit(".", 1)[0] if "." in leaf else leaf
    return {text, leaf, stem}


def _page_values_for_match(page: Dict[str, Any]) -> List[Any]:
    values: List[Any] = [
        page.get("page_id"),
        page.get("target_page_name"),
        page.get("target_page"),
        page.get("source"),
        page.get("legacy_page"),
        page.get("new_page"),
        page.get("view_page"),
    ]
    for key in ("legacy_sources", "new_sources"):
        source = page.get(key)
        if isinstance(source, list):
            values.extend(source)
    return values


def runtime_profile_paths(search_dir: Path = DEFAULT_RUNTIME_PROFILE_DIR) -> List[Path]:
    if not search_dir.exists():
        return []
    return sorted(
        (path for path in search_dir.rglob("*.json") if path.is_file()),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )


def load_runtime_profiles(search_dir: Path = DEFAULT_RUNTIME_PROFILE_DIR) -> List[Dict[str, Any]]:
    profiles: List[Dict[str, Any]] = []
    for path in runtime_profile_paths(search_dir):
        try:
            profile = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            continue
        if not isinstance(profile, dict):
            continue
        if profile.get("schema") != RUNTIME_PROFILE_SCHEMA and not isinstance(profile.get("controls"), list):
            continue
        profile.setdefault("runtime_profile_path", str(path))
        profile["_runtime_profile_mtime"] = path.stat().st_mtime
        profiles.append(profile)
    return profiles


def select_runtime_profile_for_page(page: Dict[str, Any], profiles: Sequence[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    target_keys = set()
    for value in _page_values_for_match(page):
        target_keys.update(page_match_keys(value))
    if not target_keys:
        return None

    best: Optional[Dict[str, Any]] = None
    best_score = -1
    for profile in profiles:
        profile_keys = set()
        for value in _page_values_for_match(profile):
            profile_keys.update(page_match_keys(value))
        if not target_keys.intersection(profile_keys):
            continue
        controls = profile.get("controls") if isinstance(profile.get("controls"), list) else []
        score = len(controls) + int(float(profile.get("_runtime_profile_mtime") or 0) % 100000)
        if page_match_keys(page_name_of(page)).intersection(profile_keys):
            score += 1000000
        if score > best_score:
            best = profile
            best_score = score
    if best is None:
        return None
    selected = dict(best)
    selected.setdefault("static_page_id", page_name_of(page))
    selected["profile_source"] = "runtime_profile"
    return selected


def runtime_profile_identity(profile: Dict[str, Any]) -> str:
    path = as_text(profile.get("runtime_profile_path"))
    if path:
        try:
            return str(Path(path).resolve()).lower()
        except OSError:
            return path.replace("\\", "/").lower()
    return "|".join(
        as_text(profile.get(key)).lower()
        for key in ("side", "login_entry", "route_id", "target_page", "target_page_name", "page_id")
    )


def runtime_profile_page_input(profile: Dict[str, Any]) -> Dict[str, Any]:
    page = dict(profile)
    page.setdefault("profile_source", "runtime_profile")
    page.setdefault("page_id", page.get("target_page") or page.get("target_page_name") or page.get("source"))
    return page


def page_risk(page: Dict[str, Any]) -> str:
    risk = as_text(page.get("risk"), "Medium")
    return risk if risk in CASE_DEPTH else "Medium"


def case_depth(page: Dict[str, Any]) -> str:
    return CASE_DEPTH.get(page_risk(page), "standard")


def take_by_depth(cases: List[TestCase], depth: str) -> List[TestCase]:
    if depth == "full":
        return cases
    if depth == "standard":
        return [case for case in cases if case.severity in {"High", "Medium"}]
    # smoke
    high = [case for case in cases if case.severity == "High"]
    return high[: max(1, min(3, len(high)))]


def element_attributes(element: Dict[str, Any]) -> Dict[str, Any]:
    attrs = element.get("attributes")
    if isinstance(attrs, dict):
        return attrs
    return {}


def element_field_name(element: Dict[str, Any]) -> str:
    if element.get("field_name"):
        return as_text(element.get("field_name"))
    attrs = element_attributes(element)
    kind = as_text(element.get("kind")).lower()
    tag = as_text(element.get("tag")).lower()

    if kind == "file" or tag == "html:file":
        return first_attr(attrs, "property", "path", "name", "id", "styleId")
    if tag.startswith("html:"):
        return first_attr(attrs, "property", "name", "id", "styleId")
    if tag.startswith("form:"):
        return first_attr(attrs, "path", "name", "id", "modelAttribute", "commandName")
    return first_attr(attrs, "id", "styleId", "name", "property", "path", "value", "title", "href", "action", "modelAttribute", "commandName")


def element_label(element: Dict[str, Any]) -> str:
    return (
        as_text(element.get("label"))
        or element_field_name(element)
        or as_text(element.get("semantic_key"))
        or as_text(element.get("key"))
        or as_text(element.get("legacy_locator") or element.get("new_locator") or element.get("locator"))
        or as_text(element.get("tag"), "unknown")
    )


def normalized_locator(element: Dict[str, Any]) -> str:
    kind = as_text(element.get("kind")).lower()
    tag = as_text(element.get("tag")).lower()
    attrs = element_attributes(element)
    if kind == "file" or tag == "html:file":
        field = first_attr(attrs, "property", "path", "name", "id") or element_field_name(element)
        if field:
            return f"input[name='{field}']"
    main_step = element.get("main_step") or {}
    return as_text(
        element.get("locator")
        or element.get("legacy_locator")
        or element.get("new_locator")
        or main_step.get("legacy_locator")
        or main_step.get("locator"),
        "(locator missing)",
    )


def element_evidence(element: Dict[str, Any]) -> str:
    raw = as_text(element.get("raw"))
    attrs = element_attributes(element)
    parts = []
    if raw:
        parts.append(raw)
    elif attrs:
        parts.append(", ".join(f"{key}={as_text(value)}" for key, value in sorted(attrs.items())))
    for key in ("key", "semantic_key", "legacy_locator", "new_locator", "action", "action_type"):
        if element.get(key):
            parts.append(f"{key}={as_text(element.get(key))}")
    return "\n".join(parts) if parts else as_text(element.get("tag"), "")


def field_summary(fields: Sequence[Dict[str, Any]]) -> str:
    summary = []
    for field in fields[:20]:
        label = element_label(field)
        line = as_text(field.get("line"), "-")
        locator = normalized_locator(field)
        summary.append(f"{as_text(field.get('tag'), 'field')} `{label}` line={line} locator={locator}")
    if len(fields) > 20:
        summary.append(f"... and {len(fields) - 20} more field(s)")
    return "; ".join(summary)


def common_element_fields(page: str, element: Dict[str, Any]) -> Dict[str, str]:
    evidence = element_evidence(element)
    related_fields = element.get("related_fields", [])
    if related_fields:
        evidence = f"{evidence}\n字段完整性校验：{field_summary(related_fields)}"
    return {
        "page": page,
        "kind": as_text(element.get("kind"), "unknown"),
        "locator": normalized_locator(element),
        "line": as_text(element.get("line"), "-"),
        "evidence": evidence,
    }


def attach_related_fields(elements: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Attach fields to nearest preceding form, but keep standalone field cases as well."""
    output: List[Dict[str, Any]] = []
    current_form: Optional[Dict[str, Any]] = None
    unassigned_fields: List[Dict[str, Any]] = []

    for element in elements:
        kind = as_text(element.get("kind")).lower()
        copied = dict(element)
        if kind in FIELD_LIKE_KINDS or kind == "hidden":
            field = dict(element)
            if current_form is None:
                unassigned_fields.append(field)
            else:
                current_form.setdefault("related_fields", []).append(field)
            output.append(copied)
            continue

        if kind == "form":
            copied["related_fields"] = []
            if unassigned_fields:
                copied["related_fields"].extend(unassigned_fields)
                unassigned_fields = []
            current_form = copied
        output.append(copied)

    return output


def button_semantic(element: Dict[str, Any]) -> str:
    attrs = element_attributes(element)
    onclick = first_attr(attrs, "onclick", "onClick").lower()
    action_type = as_text(element.get("action_type") or element.get("action_hint")).lower()
    if "window.close" in onclick or action_type == "close_window":
        return "close_window"
    if "download" in onclick or action_type == "download":
        return "download"
    if "fnsubmit" in onclick or ".submit" in onclick or action_type == "submit":
        return "submit"
    return action_type or "click"


def page_cases(page: Dict[str, Any]) -> List[TestCase]:
    page_name = page_name_of(page)
    counts = page.get("counts") or page.get("legacy_counts") or {}
    risk = page_risk(page)
    fields = {
        "page": page_name,
        "kind": "page",
        "locator": "-",
        "line": "-",
        "evidence": f"risk={risk}, counts={counts}, legacy_sources={page.get('legacy_sources')}, new_sources={page.get('new_sources')}",
    }
    cases = [
        TestCase("画面初期表示確認", "Legacy/New の同一業務入口から画面が正常に表示されることを確認する。", "対象 .do にアクセスし、白画面、HTTP エラー、権限エラー、タイトル/見出し/主要文言を確認する。", "両環境で対象画面が表示され、重大な表示欠落がない。", "High", **fields),
        TestCase("画面レイアウト差分確認", "移行前後の画面レイアウト差分を確認する。", "同一条件でスクリーンショットを取得し、差分率と目視差分を確認する。", "業務影響のある差分がない。差分がある場合は許容可否を記録する。", "High", **fields),
        TestCase("DOM/文言差分確認", "主要文言、項目名、ボタン名、説明文が移行前後で一致することを確認する。", "DOM テキストを取得し、Legacy/New の差分を確認する。", "業務文言やメッセージに意図しない差分がない。", "Medium", **fields),
        TestCase("文字化け確認", "Windows-31J/UTF-8 変換により日本語・記号が文字化けしないことを確認する。", "日本語、英語、全角半角、記号、改行を含む表示文言を確認する。", "豆腐文字、記号欠落、改行崩れ、エンコード崩れがない。", "High", **fields),
        TestCase("ブラウザ戻る/再読込確認", "戻る・再読込時に不正な再送信やエラー画面にならないことを確認する。", "画面表示後に戻る、進む、再読込を実行し、画面状態とメッセージを確認する。", "安全に状態復元され、二重登録・二重送信・セッション破壊が発生しない。", "Medium", **fields),
        TestCase("権限別初期表示確認", "権限により表示/非表示になる項目が移行後も同等であることを確認する。", "一般ユーザ、管理者、参照権限などで同一画面を表示する。", "権限外ボタンやリンクが表示されず、表示範囲が Legacy と一致する。", "High", **fields),
        TestCase("マルチブラウザ表示確認", "対象ブラウザで画面が同等に動作することを確認する。", "Chrome/Edge/Firefox 等で初期表示、主要ボタン、スクリーンショットを確認する。", "ブラウザ差によるレイアウト崩れや JS エラーがない。", "Low", **fields),
    ]
    return take_by_depth(cases, case_depth(page))


def form_cases(page: str, element: Dict[str, Any]) -> List[TestCase]:
    fields = common_element_fields(page, element)
    attrs = element_attributes(element)
    action = first_attr(attrs, "action") or "(current page/default action)"
    method = first_attr(attrs, "method") or "unspecified"
    enctype = first_attr(attrs, "enctype") or "unspecified"
    target = first_attr(attrs, "target") or "unspecified"
    label = element_label(element)
    return [
        TestCase(f"form action/method 確認：{label}", f"action={action}, method={method}, enctype={enctype}, target={target} が移行前後で同等であることを確認する。", "画面初期表示時の form 属性を Legacy/New で比較する。", "submit 先、method、multipart、target/window/frame が同等である。", "High", **fields),
        TestCase(f"表单提交主路径：{label}", "正常データで submit した時に想定業務処理が完了することを確認する。", "会社環境で利用可能な正常値を設定し、submit ボタンを押下する。", "新旧系统の遷移、提示文案、关键字段、后端状态一致。", "High", **fields),
        TestCase(f"必填与空值校验：{label}", "必須・空値チェックが移行後も維持されていることを確認する。", "可见输入项逐个置空；hidden 可篡改项改空后提交。", "阻止提交或返回明确业务错误；不得出现 500、空指针或静默成功。", "High", **fields),
        TestCase(f"字段完整性確認：{label}", "form 内字段数量、name、初期値、readonly/disabled 状态保持一致。", "初期表示 DOM 中的 input/select/textarea/hidden を比較する。", "Legacy/New 字段集合一致，差异均有迁移理由。", "Medium", **fields),
        TestCase(f"二重送信防止確認：{label}", "重复点击/网络延迟时不会重复登録/更新。", "网络慢速条件下快速双击 submit，并刷新/重放请求。", "只产生一次有效业务处理，重复请求被拦截或幂等处理。", "High", **fields),
        TestCase(f"边界长度与特殊字符：{label}", "覆盖字段长度、编码和转义差异。", "输入最大长度、超长、多字节、换行、单双引号、反斜杠后提交。", "长度限制稳定，多字节不乱码，特殊字符不破坏页面或 SQL/API。", "Medium", **fields),
        TestCase(f"XSS 探针注入：{label}", "提交值在当前页/确认页/错误页被正确转义。", "输入 <script>alert(1)</script> 等探针并提交。", "浏览器不执行脚本；响应中无未转义用户输入。", "High", **fields),
    ]


def field_cases(page: str, element: Dict[str, Any]) -> List[TestCase]:
    fields = common_element_fields(page, element)
    attrs = element_attributes(element)
    label = element_label(element)
    maxlength = first_attr(attrs, "maxlength", "maxLength") or "(not specified)"
    readonly = first_attr(attrs, "readonly", "readOnly") or "false"
    disabled = first_attr(attrs, "disabled") or "false"
    return [
        TestCase(f"入力項目表示確認：{label}", "入力項目が移行前後で同じ名称・位置・初期値で表示されることを確認する。", "初期表示で対象項目の表示、活性/非活性、初期値、readonly/disabled を確認する。", f"Legacy/New で表示状態、初期値、readonly={readonly}, disabled={disabled} が一致する。", "High", **fields),
        TestCase(f"入力可能確認：{label}", "対象項目に通常文字列を入力できることを確認する。", "半角英数、全角日本語、数字、記号を入力し、フォーカスアウト/submit を実行する。", "入力値が保持され、文字化けや JS エラーが発生しない。", "Medium", **fields),
        TestCase(f"最大桁数確認：{label}", f"maxlength={maxlength} の制御が移行後も維持されていることを確認する。", "最大桁数ちょうど、最大桁数超過、全角/半角混在値を入力する。", "桁数内は正常、超過時は入力制限または業務エラーとなる。", "Medium", **fields),
        TestCase(f"空値/必須チェック：{label}", "必須または業務必須項目の空値制御を確認する。", "対象項目を空欄にして登録/検索/更新を実行する。", "必要に応じてエラーメッセージが表示され、不正データが登録されない。", "High", **fields),
        TestCase(f"前後空白・改行確認：{label}", "trim、改行、タブの扱いが移行前後で一致することを確認する。", "前後空白、タブ、改行を含む値を入力して submit する。", "保存/検索/表示時の空白処理が Legacy と一致する。", "Medium", **fields),
        TestCase(f"特殊文字入力確認：{label}", "特殊文字、記号、SQL/XSS 探針に対する入力制御を確認する。", "単引号、ダブルクォート、HTML タグ、円記号、全角記号を入力する。", "画面崩れ、SQL エラー、スクリプト実行、文字化けが発生しない。", "High", **fields),
        TestCase(f"IME/多バイト文字確認：{label}", "日本語 IME 入力やサロゲート文字の扱いを確認する。", "ひらがな、カタカナ、漢字、旧字体、機種依存文字を入力する。", "文字化け・欠落・桁数誤判定がない。", "Low", **fields),
    ]


def hidden_cases(page: str, element: Dict[str, Any]) -> List[TestCase]:
    fields = common_element_fields(page, element)
    label = element_label(element)
    return [
        TestCase(f"hidden 引継ぎ確認：{label}", "画面遷移・submit 時に hidden 値が移行前後で同等に引き継がれることを確認する。", "初期表示時と submit 直前の hidden 値を Legacy/New で比較する。", "userId、projectId、権限、mode などの hidden 値が意図通り保持される。", "High", **fields),
        TestCase(f"hidden 改ざん耐性確認：{label}", "hidden 値改ざんで権限回避や他データ操作ができないことを確認する。", "hidden 値を空値、他ユーザ値、不正値、超長値に変更して submit する。", "不正値は拒否され、権限外データ参照・更新が発生しない。", "High", **fields),
        TestCase(f"hidden 欠落時エラー制御：{label}", "hidden 欠落時に安全なエラー制御となることを確認する。", "対象 hidden を DOM から削除して submit する。", "NullPointer/500 ではなく業務エラーまたは安全な再表示となる。", "Medium", **fields),
    ]


def file_cases(page: str, element: Dict[str, Any]) -> List[TestCase]:
    fields = common_element_fields(page, element)
    attrs = element_attributes(element)
    accept = first_attr(attrs, "accept") or "(not specified)"
    label = element_label(element)
    return [
        TestCase(f"ファイル選択欄表示確認：{label}", "ファイル選択欄が移行前後で表示され、input[type=file] として操作可能であることを確認する。", "初期表示でファイル選択欄の表示、name、活性状態を確認する。", "Legacy/New で同じ name の file input が存在し、操作可能である。", "High", **fields),
        TestCase(f"未選択アップロード：{label}", "ファイル未選択でアップロードした場合のエラー制御を確認する。", "ファイルを選択せずアップロードボタンを押下する。", "業務エラーが表示され、500 エラーや空登録が発生しない。", "High", **fields),
        TestCase(f"正常ファイルアップロード：{label}", f"業務で許可されたファイルを正常にアップロードできることを確認する。accept={accept}。", "有効なテンプレート/サンプルファイルを選択して submit する。", "上传成功；メッセージ、遷移、登録結果が Legacy/New で一致する。", "High", **fields),
        TestCase(f"空ファイルアップロード：{label}", "0 byte または空内容ファイルの扱いを確認する。", "0 byte ファイル、空行のみのファイルをアップロードする。", "不正ファイルとして拒否、または仕様通り処理される。", "High", **fields),
        TestCase(f"拡張子不正：{label}", "許可外拡張子が拒否されることを確認する。", "txt、exe、zip 等の非許可拡張子ファイルをアップロードする。", "非法类型被拒绝；错误消息清晰；服务端不保存危险文件。", "High", **fields),
        TestCase(f"ファイルタイプ偽装：{label}", "扩展名と MIME/内容が不一致のファイルを拒否できることを確認する。", "テキストを .xls/.xlsx に改名、または MIME が異なるファイルをアップロードする。", "真实内容校验或业务校验生效，不发生异常处理。", "High", **fields),
        TestCase(f"サイズ上限境界：{label}", "multipart 上限、业务上限、反向代理限制が一致することを確認する。", "上限直下、上限ちょうど、上限超過ファイルをアップロードする。", "境界内は正常、超過は安定拒否。500/504 や临时文件残留がない。", "High", **fields),
        TestCase(f"日本語ファイル名：{label}", "日本語ファイル名が文字化けせず処理されることを確認する。", "漢字、かな、全角スペース、長音、括弧を含むファイル名でアップロードする。", "メッセージ、ログ、登録結果、ダウンロード時のファイル名が文字化けしない。", "Medium", **fields),
        TestCase(f"記号付きファイル名：{label}", "記号・空白を含むファイル名の安全性を確認する。", "../、..\\、单双引号、空白、括号、换行、超长文件名でアップロードする。", "路径穿越、日志污染、页面崩れがなく、安全に拒否または正規化される。", "Medium", **fields),
        TestCase(f"テンプレート形式不正：{label}", "Excel/CSV 等の列不足、型不正、必須列欠落を検出できることを確認する。", "必須列缺失、列名错误、重复行、型不正、件数超過のファイルをアップロードする。", "該当行/列の業務エラーが表示され、不正データは登録されない。", "High", **fields),
        TestCase(f"二重アップロード：{label}", "同一ファイルを連続アップロードした場合の重複処理を確認する。", "同じファイルを連続でアップロードし、二重クリックも実行する。", "重複登録が防止される、または仕様通り上書き/エラーとなる。", "Medium", **fields),
        TestCase(f"アップロード後メッセージ/画面状態：{label}", "アップロード後のメッセージ、画面遷移、入力欄状態が移行前後で一致することを確認する。", "正常/异常アップロード後の画面、メッセージ、戻る操作、再アップロード可否を確認する。", "Legacy/New でメッセージ、遷移、再操作状態が一致する。", "High", **fields),
    ]


def button_cases(page: str, element: Dict[str, Any]) -> List[TestCase]:
    fields = common_element_fields(page, element)
    attrs = element_attributes(element)
    onclick = first_attr(attrs, "onclick", "onClick") or "(not specified)"
    label = element_label(element)
    semantic = button_semantic(element)
    cases = [
        TestCase(f"ボタン表示/活性確認：{label}", "ボタンが移行前後で同じ表示名・位置・活性状態で表示されることを確認する。", "初期表示でボタンの文言、表示、disabled、権限別表示を確認する。", "表示/非表示、活性/非活性、文言が Legacy/New で一致する。", "High", **fields),
        TestCase(f"ボタンクリック主路径：{label}", "ボタン押下時に移行前後で同じ業務動作が実行されることを確認する。", "正常页面状态でクリックし、请求、跳转、弹窗、页面刷新、后端状态变化を記録する。", "新旧系统行为一致；按钮不会无响应、重复触发或触发错误 action。", "High", **fields),
        TestCase(f"前端脚本依赖检查：{label}", f"onclick 或关联脚本迁移后没有丢失。onclick={onclick}。", "打开浏览器控制台点击按钮；观察 JS 错误、缺失函数、未定义变量和被拦截请求。", "控制台无脚本错误；动态校验、确认框、参数拼接、页面状态更新正常。", "Medium", **fields),
        TestCase(f"重复点击与防重提交：{label}", "重复点击时不会重复登记/更新/发送。", "快速双击/多击按钮；网络慢速条件下重复点击。", "只产生一次有效业务处理；重复请求被拦截或安全幂等。", "High", **fields),
    ]
    if semantic == "submit":
        cases.extend([
            TestCase(f"submit 先確認：{label}", "按钮押下时提交到预期 action，target/frame/window 保持一致。", "点击按钮并记录 request URL、method、target、遷移先。", "Legacy/New で submit 先、遷移、メッセージが一致する。", "High", **fields),
            TestCase(f"submit エラー時再表示確認：{label}", "submit 后业务错误时页面可安全再显示。", "输入异常数据后点击按钮，确认错误消息和输入保持。", "不发生 500/白画面；错误消息、输入保持、焦点位置与 Legacy 一致。", "High", **fields),
        ])
    elif semantic == "close_window":
        cases.extend([
            TestCase(f"閉じる/キャンセル動作確認：{label}", "キャンセル/閉じるボタン押下時に対象ウィンドウが仕様通り閉じることを確認する。", "クリック後の window close、親画面状態、セッション状態を確認する。", "対象ウィンドウが閉じる、または仕様通り前画面へ戻る。親画面に異常がない。", "High", **fields),
            TestCase(f"閉じる後の再表示確認：{label}", "閉じる操作後に再度同じ画面を開けることを確認する。", "キャンセル後、メニューまたは業務入口から同じ画面を再表示する。", "再表示でき、session/context が壊れていない。", "Medium", **fields),
        ])
    elif semantic == "download":
        cases.extend([
            TestCase(f"ダウンロード起動確認：{label}", "ボタン押下で想定ファイル出力が開始されることを確認する。", "クリック後の download event、ファイル名、拡張子、サイズを確認する。", "Legacy/New でファイル名、形式、内容概要が一致する。", "High", **fields),
            TestCase(f"ダウンロード権限確認：{label}", "権限外ユーザがファイル出力できないことを確認する。", "権限の異なるユーザで同ボタンを表示/クリックする。", "権限外では非表示またはエラーとなり、ファイルが出力されない。", "High", **fields),
        ])
    return cases


def link_cases(page: str, element: Dict[str, Any]) -> List[TestCase]:
    fields = common_element_fields(page, element)
    attrs = element_attributes(element)
    href = first_attr(attrs, "href", "action", "page") or first_attr(attrs, "onclick", "onClick") or "(not specified)"
    target = first_attr(attrs, "target") or "(not specified)"
    label = element_label(element)
    return [
        TestCase(f"リンク表示確認：{label}", "リンクが移行前後で同じ文言・位置・表示条件で表示されることを確認する。", "初期表示でリンク文言、href/onclick、target、権限別表示を確認する。", "Legacy/New で表示状態、文言、href/target が一致する。", "High", **fields),
        TestCase(f"リンク导航主路径：{label}", f"迁移后的链接仍进入正确页面或业务动作，target={href}。", "点击链接，记录目标 URL、请求参数、页面标题和关键内容。", "新旧系统目标一致；参数未丢失；无 404/500；登录态和权限状态保持正确。", "High", **fields),
        TestCase(f"リンク target/popup 確認：{label}", f"target={target} の挙動が移行前後で同等であることを確認する。", "通常クリック、新タブ/別窗口、popup の有無を確認する。", "popup/window/frame の開き方が Legacy/New で一致する。", "Medium", **fields),
        TestCase(f"パラメータ引継ぎ確認：{label}", "リンク押下時の id、mode、returnUrl 等のパラメータが正しく引き継がれることを確認する。", "クリック前後の URL、hidden、request parameter を確認する。", "必要パラメータが欠落せず、余計な機密情報が露出しない。", "High", **fields),
        TestCase(f"参数篡改与权限校验：{label}", "リンクパラメータ改ざんで権限外データを参照できないことを確認する。", "URL 参数を不存在、越权、空值、特殊字符、超长值に変更して访问。", "非法参数被拒绝或回到安全页面；不得泄漏数据、堆栈或内部路径。", "High", **fields),
        TestCase(f"返回与打开方式：{label}", "浏览器返回、刷新、新标签页打开时状态一致。", "点击后执行返回、刷新、新标签页打开；对弹窗/下载额外确认。", "页面状态可恢复；不会重复提交危险操作；下载/弹窗行为与旧系统一致。", "Medium", **fields),
        TestCase(f"リンク切れ確認：{label}", "ヘルプ・外部リンク・静的ファイルリンクが切れていないことを確認する。", "リンク先にアクセスし、HTTP status、文字化け、404/500 を確認する。", "リンク先が正常表示される。環境差分がある場合は許容理由を記録する。", "Medium", **fields),
        TestCase(f"別言語リンク確認：{label}", "言語切替によりヘルプ/文言リンクが正しく切り替わることを確認する。", "日本語/英語等の言語設定でリンク先と表示文言を確認する。", "言語に応じた正しいリンク先・文言が表示される。", "Low", **fields),
    ]




def scenario_cases(page: str, element: Dict[str, Any]) -> List[TestCase]:
    fields = common_element_fields(page, element)
    label = element_label(element)
    case_type = as_text(element.get("case_type") or element.get("action_hint"), "scenario")
    pre_steps = element.get("pre_steps") or []
    main_step = element.get("main_step") or {}
    if case_type == "upload_submit":
        meta = executable_metadata(element)
        return [
            TestCase(
                title=f"自动化実行シナリオ：ファイル選択→submit：{label}",
                objective="ファイルアップロード画面で、ファイル選択後に form submit する一連の主経路を自動実行対象として確認する。",
                steps="pre_steps で uploadFile にテストファイルを設定し、main_step で form submit を実行する。",
                expected="Legacy/New とも submit が完了し、遷移先・メッセージ・画面状態に重大差分がない。",
                severity="High",
                **fields,
                automation_mode="auto",
                case_type="upload_submit",
                action_type="upload_submit",
                test_data=meta.get("test_data", ""),
                submit_locator=meta.get("submit_locator", ""),
                expected_type=meta.get("expected_type", "visual_or_message"),
                expected_value=meta.get("expected_value", ""),
                pre_steps=meta.get("pre_steps", ""),
                main_step=meta.get("main_step", ""),
            )
        ]
    meta = executable_metadata(element)
    return [
        TestCase(
            title=f"自动化実行シナリオ：{case_type}：{label}",
            objective="page_mapping の executable_cases として生成された実行対象シナリオを確認する。",
            steps=f"case_type={case_type}, pre_steps={len(pre_steps)}, main_step={as_text(main_step.get('action_type'), '-')} を実行する。",
            expected="Legacy/New で同等の実行結果となり、BLOCKED または重大 DIFF が発生しない。",
            severity="Medium",
            **fields,
            automation_mode="auto",
            case_type=meta.get("case_type", case_type),
            action_type=meta.get("action_type", as_text(main_step.get("action_type"), "")),
            test_data=meta.get("test_data", ""),
            submit_locator=meta.get("submit_locator", ""),
            expected_type=meta.get("expected_type", "visual_or_message"),
            expected_value=meta.get("expected_value", ""),
            pre_steps=meta.get("pre_steps", ""),
            main_step=meta.get("main_step", ""),
        )
    ]

CASE_BUILDERS = {
    "form": form_cases,
    "field": field_cases,
    "select": field_cases,
    "textarea": field_cases,
    "hidden": hidden_cases,
    "file": file_cases,
    "button": button_cases,
    "link": link_cases,
    "scenario": scenario_cases,
}


def normalize_mapping_element(element: Dict[str, Any]) -> Dict[str, Any]:
    copied = dict(element)
    if not copied.get("locator"):
        copied["locator"] = copied.get("legacy_locator") or copied.get("new_locator")
    if not copied.get("action_hint") and copied.get("action_type"):
        copied["action_hint"] = copied.get("action_type")
    if not copied.get("label"):
        copied["label"] = element_label(copied)
    return copied


def collect_elements(page: Dict[str, Any]) -> List[Dict[str, Any]]:
    elements: List[Dict[str, Any]] = []

    def add_items(items: Any, source: str) -> None:
        for item in items or []:
            if not isinstance(item, dict):
                continue
            copied = dict(item)
            copied.setdefault("_source", source)
            elements.append(copied)

    # Scanner output.
    add_items(page.get("elements", []), "elements")

    # Mapping output. Keep matched elements so successful pairs also generate checklist rows.
    for key in ("matched_elements", "locator_changes", "missing_legacy_elements", "missing_new_elements", "full_action_steps"):
        add_items(page.get(key, []), key)

    # Executable scenarios are the bridge between the human checklist and the
    # automated regression runner. Include them so the Excel clearly shows
    # which generated items are actually executable.
    for case in page.get("executable_cases", []) or []:
        if isinstance(case, dict):
            item = dict(case)
            item.setdefault("kind", "scenario")
            item.setdefault("locator", item.get("legacy_locator") or (item.get("main_step") or {}).get("legacy_locator"))
            item.setdefault("action_hint", item.get("case_type"))
            item.setdefault("_source", "executable_cases")
            elements.append(item)

    # Some mapping formats only have legacy/new side elements embedded under comparable pairs.
    for key in ("actions", "action_items", "test_actions"):
        add_items(page.get(key, []), key)

    normalized: List[Dict[str, Any]] = []
    seen: set[Tuple[str, str, str, str, str]] = set()
    for element in elements:
        if not isinstance(element, dict):
            continue
        normalized_element = normalize_mapping_element(element)
        kind = as_text(normalized_element.get("kind")).lower()
        locator = normalized_locator(normalized_element)
        label = element_label(normalized_element).strip()
        
        # 过滤低价值元素，减少 Checklist 噪音
        if kind in {"link", "button"} and not label and locator == "(locator missing)":
            continue
        if kind == "hidden":
            continue # 隐藏域通常不需要出现在人工 Checklist 中
            
        line = as_text(normalized_element.get("line"), "-")
        source = as_text(normalized_element.get("_source"), "")
        key = (kind, locator, label, line, source)
        if key in seen:
            continue
        seen.add(key)
        normalized.append(normalized_element)
    return normalized


def infer_action_type(element: Dict[str, Any]) -> str:
    kind = as_text(element.get("kind")).lower()
    if kind == "button":
        return button_semantic(element)
    action = as_text(element.get("action_type") or element.get("action_hint") or element.get("case_type")).lower()
    if action:
        return action
    if kind == "file":
        return "upload"
    if kind == "link":
        attrs = element_attributes(element)
        onclick = first_attr(attrs, "onclick", "onClick").lower()
        href = first_attr(attrs, "href").lower()
        if "download" in onclick:
            return "download"
        if "fnsubmit" in onclick:
            return "submit"
        if href:
            return "navigate"
    if kind == "form":
        return "submit"
    return ""


def executable_metadata(element: Dict[str, Any]) -> Dict[str, str]:
    kind = as_text(element.get("kind")).lower()
    case_type = as_text(element.get("case_type") or element.get("action_hint") or element.get("action_type"))
    action_type = infer_action_type(element)

    pre_steps = element.get("pre_steps") or []
    main_step = element.get("main_step") or {}

    locator = normalized_locator(element)
    submit_locator = as_text(
        element.get("submit_locator")
        or element.get("submit_legacy_locator")
        or main_step.get("legacy_locator")
        or main_step.get("locator")
    )

    test_data = as_text(element.get("test_data") or element.get("value"))
    if not test_data and isinstance(pre_steps, list):
        for step in pre_steps:
            if isinstance(step, dict) and as_text(step.get("action_type")).lower() in {"upload", "set_input_files"}:
                test_data = as_text(step.get("value") or step.get("test_data"))
                locator = as_text(step.get("legacy_locator") or step.get("locator") or locator)
                break

    if not test_data and action_type in {"upload", "upload_submit"}:
        test_data = "test_data/upload/default_valid.xlsx"

    if not case_type:
        case_type = action_type or kind

    if case_type == "upload_submit":
        action_type = "upload_submit"
        if not submit_locator:
            submit_locator = as_text(main_step.get("legacy_locator") or main_step.get("locator"))

    expected_type = as_text(element.get("expected_type"))
    if not expected_type:
        if action_type in {"download"}:
            expected_type = "download"
        elif action_type in {"close_window"}:
            expected_type = "window_closed"
        elif action_type in {"upload", "upload_submit", "submit", "click", "navigate"}:
            expected_type = "visual_or_message"
        else:
            expected_type = "manual_review"

    return {
        "automation_mode": "auto",
        "case_type": case_type,
        "action_type": action_type,
        "test_data": test_data,
        "submit_locator": submit_locator,
        "expected_type": expected_type,
        "expected_value": as_text(element.get("expected_value")),
        "pre_steps": as_json(pre_steps),
        "main_step": as_json(main_step),
        "locator": locator,
    }


def infer_case_action_type(case: TestCase, element: Dict[str, Any]) -> str:
    kind = as_text(element.get("kind") or case.kind).lower()
    title_blob = " ".join([case.title, case.objective, case.steps]).lower()

    if kind == "page":
        return "snapshot"
    if kind == "form":
        return "submit"
    if kind == "file":
        return "upload"
    if kind == "button":
        return button_semantic(element)
    if kind == "link":
        return infer_action_type(element) or "navigate"
    if kind == "hidden":
        return "set_value"
    if kind in {"select"} or "select" in as_text(element.get("tag")).lower():
        return "select"
    if kind in FIELD_LIKE_KINDS:
        if any(token in title_blob for token in ("空", "empty", "required", "必須")):
            return "clear"
        return "fill"
    return infer_action_type(element)


def default_test_data_for_case(action_type: str, case: TestCase, element: Dict[str, Any]) -> str:
    existing = as_text(case.test_data or element.get("test_data") or element.get("value"))
    if existing:
        return existing

    title_blob = " ".join([case.title, case.objective, case.steps]).lower()
    attrs = element_attributes(element)
    maxlength = first_attr(attrs, "maxlength", "maxLength")

    if action_type in {"clear"}:
        return ""
    if action_type == "set_value":
        return "moonlight-hidden-auto"
    if action_type == "press":
        return "Enter"
    if action_type == "upload":
        if any(token in title_blob for token in ("0 byte", "空ファイル", "empty")):
            return "test_data/upload/empty.txt"
        if any(token in title_blob for token in ("拡張子", "invalid", "不正", ".exe")):
            return "test_data/upload/invalid_type.exe"
        if any(token in title_blob for token in ("サイズ", "large", "上限")):
            return "test_data/upload/large_sample.tsv"
        if any(token in title_blob for token in ("日本語", "文字化け")):
            return "test_data/upload/日本語ファイル名.tsv"
        return "test_data/upload/プロジェクトリストアップロード.tsv"
    if action_type == "fill":
        if "xss" in title_blob or "script" in title_blob:
            return "<script>alert(1)</script>"
        if "sql" in title_blob:
            return "' OR '1'='1"
        if any(token in title_blob for token in ("数字", "number", "件数")):
            return "1234567890"
        if maxlength and maxlength.isdigit():
            size = max(1, min(int(maxlength), 128))
            return "M" * size
        if any(token in title_blob for token in ("最大", "maxlength", "桁")):
            return "M" * 64
        if any(token in title_blob for token in ("日本語", "ime")):
            return "自動化テスト"
        return "moonlight-auto"
    return ""


def expected_type_for_action(action_type: str) -> str:
    if action_type == "download":
        return "download"
    if action_type == "close_window":
        return "window_closed"
    if action_type in {"snapshot", "wait"}:
        return "visual"
    if action_type in {"fill", "clear", "select", "set_value", "press"}:
        return "control_state"
    if action_type in {"upload", "upload_submit", "submit", "click", "navigate"}:
        return "visual_or_message"
    return "manual_review"


def auto_enrich_case(case: TestCase, element: Dict[str, Any]) -> TestCase:
    if case.automation_mode == "auto" and case.action_type:
        return case

    action_type = infer_case_action_type(case, element)
    locator = normalized_locator(element)
    can_auto = action_type in AUTO_ACTION_TYPES and (
        locator not in {"", "(locator missing)"} or action_type == "snapshot"
    )
    if action_type == "snapshot":
        locator = "__page__"

    return replace(
        case,
        automation_mode="auto" if can_auto else "manual/assist",
        case_type=case.case_type or action_type,
        action_type=case.action_type or action_type,
        test_data=case.test_data or default_test_data_for_case(action_type, case, element),
        submit_locator=case.submit_locator or as_text(element.get("submit_locator")),
        expected_type=case.expected_type or expected_type_for_action(action_type),
        expected_value=case.expected_value,
        locator=locator if can_auto else case.locator,
    )


def auto_enrich_cases(cases: List[TestCase], element: Dict[str, Any]) -> List[TestCase]:
    return [auto_enrich_case(case, element) for case in cases]


def is_executable_element(element: Dict[str, Any]) -> bool:
    source = as_text(element.get("_source"))
    kind = as_text(element.get("kind")).lower()
    if source == "executable_cases":
        return True
    if source == "full_action_steps" and kind in {"link", "button", "file"}:
        return bool(infer_action_type(element) and normalized_locator(element))
    return False


def automation_case(page: str, element: Dict[str, Any]) -> Optional[TestCase]:
    if not is_executable_element(element):
        return None

    meta = executable_metadata(element)
    label = element_label(element)
    action_type = meta["action_type"] or "-"
    case_type = meta["case_type"] or action_type
    fields = common_element_fields(page, element)
    fields["locator"] = meta["locator"] or fields["locator"]

    if case_type == "upload_submit":
        title = f"AUTO 実行：ファイル選択→submit：{label}"
        objective = "Excel checklist から自動実行できるファイルアップロード主経路を定義する。"
        steps = "locator にテストファイルを設定し、submit_locator の form/button/link で submit する。"
        expected = "Legacy/New とも実行が完了し、BLOCKED または重大 DIFF が発生しない。"
    else:
        title = f"AUTO 実行：{action_type}：{label}"
        objective = "Excel checklist から自動実行できる画面操作を定義する。"
        steps = "対象 locator に対して action_type の操作を実行し、Legacy/New の結果を比較する。"
        expected = "Legacy/New とも同等の結果となり、BLOCKED または重大 DIFF が発生しない。"

    return TestCase(
        title,
        objective,
        steps,
        expected,
        "High",
        **fields,
        automation_mode=meta["automation_mode"],
        case_type=case_type,
        action_type=action_type,
        test_data=meta["test_data"],
        submit_locator=meta["submit_locator"],
        expected_type=meta["expected_type"],
        expected_value=meta["expected_value"],
        pre_steps=meta["pre_steps"],
        main_step=meta["main_step"],
    )


def uses_page_specific_planner(scan_data: Dict[str, Any]) -> bool:
    return (
        isinstance(scan_data.get("page_mappings"), list)
        or isinstance(scan_data.get("runtime_profiles"), list)
        or scan_data.get("schema") == RUNTIME_PROFILE_SCHEMA
        or isinstance(scan_data.get("controls"), list)
    )


def planned_case_to_test_case(case: Dict[str, Any]) -> TestCase:
    capabilities = as_text(case.get("matched_capabilities"))
    evidence_parts = [
        f"template_id={as_text(case.get('template_id'))}",
        f"generated_by={as_text(case.get('generated_by'), 'PageCasePlanner')}",
    ]
    if case.get("parent_case_id"):
        evidence_parts.append(f"parent_case_id={as_text(case.get('parent_case_id'))}")
    if case.get("viewpoint_id"):
        evidence_parts.append(f"viewpoint_id={as_text(case.get('viewpoint_id'))}")
    if capabilities:
        evidence_parts.append(f"matched_capabilities={capabilities}")
    if case.get("profile_source"):
        evidence_parts.append(f"profile_source={as_text(case.get('profile_source'))}")
    if case.get("runtime_profile_path"):
        evidence_parts.append(f"runtime_profile_path={as_text(case.get('runtime_profile_path'))}")
    return TestCase(
        title=as_text(case.get("title"), as_text(case.get("case_type"), "planned case")),
        objective=as_text(case.get("objective")),
        steps=as_text(case.get("steps")),
        expected=as_text(case.get("expected")),
        severity=as_text(case.get("severity") or case.get("risk"), "High"),
        page=as_text(case.get("page_id"), "<unknown>"),
        kind=as_text(case.get("case_type"), "page"),
        locator=as_text(case.get("locator"), "__page__"),
        line="-",
        evidence="\n".join(part for part in evidence_parts if part),
        automation_mode=as_text(case.get("automation_mode"), "auto"),
        case_type=as_text(case.get("case_type")),
        action_type=as_text(case.get("action_type")),
        test_data=as_text(case.get("test_data")),
        submit_locator=as_text(case.get("submit_locator")),
        expected_type=as_text(case.get("expected_type")),
        expected_value=as_text(case.get("expected_value")),
        pre_steps=as_text(case.get("pre_steps")),
        main_step=as_text(case.get("main_step")),
        case_id=as_text(case.get("case_id")),
        parent_case_id=as_text(case.get("parent_case_id")),
        viewpoint_id=as_text(case.get("viewpoint_id")),
        enabled=as_text(case.get("enabled"), "true"),
        generated_by=as_text(case.get("generated_by"), "PageCasePlanner"),
        matched_capabilities=capabilities,
        destructive=as_text(case.get("destructive"), "false"),
        priority=as_text(case.get("priority")),
    )


def case_priority_order(case: TestCase) -> Tuple[int, str]:
    try:
        return int(case.priority), case.priority
    except (TypeError, ValueError):
        return 1000 + SEVERITY_ORDER.get(case.severity, 99), case.severity


def plan_page_specific_cases(
    scan_data: Dict[str, Any],
    *,
    runtime_profile_dir: Path = DEFAULT_RUNTIME_PROFILE_DIR,
    include_runtime_profile_dir: bool = False,
) -> Tuple[List[TestCase], List[Dict[str, Any]], List[Dict[str, Any]]]:
    planner = PageCasePlanner()
    cases: List[TestCase] = []
    skipped_templates: List[Dict[str, Any]] = []
    profiles: List[Dict[str, Any]] = []
    runtime_profiles = [] if scan_data.get("schema") == RUNTIME_PROFILE_SCHEMA else load_runtime_profiles(runtime_profile_dir)
    consumed_runtime_profiles = set()

    for page in page_entries(scan_data):
        selected_profile = select_runtime_profile_for_page(page, runtime_profiles)
        profile_input = selected_profile or page
        if selected_profile:
            consumed_runtime_profiles.add(runtime_profile_identity(selected_profile))
        planned_cases, skipped, profile = planner.plan(profile_input)
        cases.extend(planned_case_to_test_case(case) for case in planned_cases)
        skipped_templates.extend(skipped)
        profiles.append(profile)

    if include_runtime_profile_dir:
        for runtime_profile in runtime_profiles:
            identity = runtime_profile_identity(runtime_profile)
            if identity in consumed_runtime_profiles:
                continue
            profile_input = runtime_profile_page_input(runtime_profile)
            planned_cases, skipped, profile = planner.plan(profile_input)
            cases.extend(planned_case_to_test_case(case) for case in planned_cases)
            skipped_templates.extend(skipped)
            profiles.append(profile)
            consumed_runtime_profiles.add(identity)

    cases.sort(
        key=lambda item: (
            item.page,
            item.destructive == "true",
            case_priority_order(item),
            SEVERITY_ORDER.get(item.severity, 99),
            item.case_type,
            item.case_id,
            item.viewpoint_id,
        )
    )
    return cases, skipped_templates, profiles


def generate_cases(
    scan_data: Dict[str, Any],
    *,
    runtime_profile_dir: Path = DEFAULT_RUNTIME_PROFILE_DIR,
    include_runtime_profile_dir: bool = False,
) -> List[TestCase]:
    if uses_page_specific_planner(scan_data):
        cases, _, _ = plan_page_specific_cases(
            scan_data,
            runtime_profile_dir=runtime_profile_dir,
            include_runtime_profile_dir=include_runtime_profile_dir,
        )
        return cases

    cases: List[TestCase] = []
    for page in page_entries(scan_data):
        page_name = page_name_of(page)
        depth = case_depth(page)
        page_element = {
            "kind": "page",
            "label": page_name,
            "locator": "__page__",
            "line": "-",
        }
        cases.extend(auto_enrich_cases(page_cases(page), page_element))

        for element in attach_related_fields(collect_elements(page)):
            auto_case = automation_case(page_name, element)
            if auto_case is not None:
                cases.append(auto_case)

            kind = as_text(element.get("kind")).lower()
            if kind not in CASE_GENERATING_KINDS:
                continue
            builder = CASE_BUILDERS.get(kind)
            if builder is None:
                continue
            built = auto_enrich_cases(builder(page_name, element), element)
            cases.extend(take_by_depth(built, depth))

    return sorted(cases, key=lambda item: (item.page, SEVERITY_ORDER.get(item.severity, 99), item.kind, item.title, item.locator))


def summarize_counts(scan_data: Dict[str, Any], cases: Sequence[TestCase]) -> Dict[str, int]:
    totals = {str(key): int(value) for key, value in scan_data.get("totals", {}).items()}
    if not totals:
        for page in page_entries(scan_data):
            for kind, count in (page.get("counts") or page.get("legacy_counts") or {}).items():
                try:
                    totals[str(kind)] = totals.get(str(kind), 0) + int(count)
                except Exception:
                    pass
    if not totals:
        for case in cases:
            totals[case.kind] = totals.get(case.kind, 0) + 1
    totals["test_cases"] = len(cases)
    return totals


def markdown_escape_cell(value: str) -> str:
    return html.escape(value).replace("\n", "<br>").replace("|", "\\|")


def unique_evidence_cases(cases: Sequence[TestCase]) -> List[TestCase]:
    unique: List[TestCase] = []
    seen = set()
    for case in cases:
        key = (case.page, case.line, case.kind, case.locator, case.evidence)
        if key in seen:
            continue
        seen.add(key)
        unique.append(case)
    return unique


def universal_checklist_markdown_lines() -> List[str]:
    lines = ["", "## Universal Migration Checklist", "", "| # | Checklist item | Automation mode | Evidence policy |", "|---|---|---|---|"]
    for index, section in enumerate(CHECKLIST_SECTIONS, start=1):
        lines.append("| " + " | ".join(markdown_escape_cell(value) for value in [str(index), section["title"], section["mode"], section["expected"]]) + " |")
    return lines


def render_markdown(scan_data: Dict[str, Any], cases: Sequence[TestCase]) -> str:
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    pages = list(page_entries(scan_data))
    totals = summarize_counts(scan_data, cases)
    lines = [
        "# 自动化测试建议报告",
        "",
        f"- 生成时间：{generated_at}",
        f"- 扫描根路径：{as_text(scan_data.get('root'), as_text(scan_data.get('source'), '<unknown>'))}",
        f"- 页面数量：{len(pages)}",
        f"- 元素统计：form={totals.get('form', 0)}，file={totals.get('file', 0)}，button={totals.get('button', 0)}，link={totals.get('link', 0)}",
        f"- 建议用例数：{totals.get('test_cases', 0)}",
        "",
        "## 主优先执行清单",
        "",
    ]
    high_cases = [case for case in cases if case.severity == "High"]
    for index, case in enumerate(high_cases[:20], start=1):
        lines.append(f"{index}. [{case.page}:{case.line}] {case.title} - {case.objective}")
    if not high_cases:
        lines.append("未发现可生成的高优先级用例。请确认输入 JSON 中包含页面/元素或 page_mapping 信息。")
    lines.extend(universal_checklist_markdown_lines())
    lines.extend(["", "## 用例明细", "", "| # | 优先级 | 页面 | 行 | 类型 | 定位器 | 用例 | 操作建议 | 期望结果 | 自动化模式 | action_type |", "|---|---|---|---|---|---|---|---|---|---|---|"])
    for index, case in enumerate(cases, start=1):
        lines.append("| " + " | ".join(markdown_escape_cell(value) for value in [str(index), case.severity, case.page, case.line, case.kind, case.locator, f"{case.title}\n{case.objective}", case.steps, case.expected, case.automation_mode, case.action_type]) + " |")
    lines.extend(["", "## 元素证据", "", "| 页面 | 行 | 类型 | 定位器 | JSP 线索 |", "|---|---|---|---|---|"])
    for case in unique_evidence_cases(cases):
        lines.append("| " + " | ".join(markdown_escape_cell(value) for value in [case.page, case.line, case.kind, case.locator, case.evidence]) + " |")
    return "\n".join(lines) + "\n"


def write_excel(
    path: Path,
    scan_data: Dict[str, Any],
    cases: Sequence[TestCase],
    *,
    runtime_profile_dir: Path = DEFAULT_RUNTIME_PROFILE_DIR,
    include_runtime_profile_dir: bool = False,
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise SystemExit("Excel output requires openpyxl. Install dependencies with: pip install -r requirements.txt") from exc

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    pages = list(page_entries(scan_data))
    totals = summarize_counts(scan_data, cases)
    skipped_templates: List[Dict[str, Any]] = []
    page_profiles: List[Dict[str, Any]] = []
    if uses_page_specific_planner(scan_data):
        _, skipped_templates, page_profiles = plan_page_specific_cases(
            scan_data,
            runtime_profile_dir=runtime_profile_dir,
            include_runtime_profile_dir=include_runtime_profile_dir,
        )
    summary_rows = [
        ("生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("扫描根路径", as_text(scan_data.get("root"), as_text(scan_data.get("source"), "<unknown>"))),
        ("页面数量", len(page_profiles) if page_profiles else len(pages)),
        ("form", totals.get("form", 0)),
        ("file", totals.get("file", 0)),
        ("button", totals.get("button", 0)),
        ("link", totals.get("link", 0)),
        ("建议用例数", totals.get("test_cases", 0)),
    ]
    for row in summary_rows:
        summary.append(row)
    summary.column_dimensions["A"].width = 18
    summary.column_dimensions["B"].width = 90

    universal = workbook.create_sheet("UniversalChecklist")
    universal.append(["#", "Checklist item", "Automation mode", "Evidence policy"])
    for index, section in enumerate(CHECKLIST_SECTIONS, start=1):
        universal.append([index, section["title"], section["mode"], section["expected"]])
    for cell in universal[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row in universal.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for index, width in enumerate([6, 36, 18, 72], start=1):
        universal.column_dimensions[get_column_letter(index)].width = width
    universal.freeze_panes = "A2"
    universal.auto_filter.ref = universal.dimensions

    if page_profiles:
        profile_sheet = workbook.create_sheet("PageProfile")
        profile_headers = [
            "page_id",
            "profile_source",
            "runtime_profile_path",
            "entry_url",
            "view_page",
            "ready_selector",
            "capabilities",
            "form_count",
            "file_count",
            "button_count",
            "link_count",
            "input_count",
            "select_count",
            "textarea_count",
            "table_count",
            "submit_action_count",
            "download_action_count",
            "close_action_count",
        ]
        profile_sheet.append(profile_headers)
        for profile in page_profiles:
            counts = profile.get("counts") or {}
            capabilities = profile.get("capabilities") or {}
            enabled_caps = ",".join(name for name, enabled in capabilities.items() if enabled)
            profile_sheet.append(
                [
                    profile.get("page_id"),
                    profile.get("profile_source"),
                    profile.get("runtime_profile_path"),
                    profile.get("entry_url"),
                    profile.get("view_page"),
                    profile.get("ready_selector"),
                    enabled_caps,
                    counts.get("form", 0),
                    counts.get("file", 0),
                    counts.get("button", 0),
                    counts.get("link", 0),
                    counts.get("input", 0),
                    counts.get("select", 0),
                    counts.get("textarea", 0),
                    counts.get("table", 0),
                    len(profile.get("submit_actions") or []),
                    len(profile.get("download_actions") or []),
                    len(profile.get("close_actions") or []),
                ]
            )
        for cell in profile_sheet[1]:
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        for row in profile_sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for index, width in enumerate([34, 22, 54, 34, 34, 34, 72, 12, 12, 12, 12, 12, 12, 12, 12, 18, 20, 18], start=1):
            profile_sheet.column_dimensions[get_column_letter(index)].width = width
        profile_sheet.freeze_panes = "A2"
        profile_sheet.auto_filter.ref = profile_sheet.dimensions

    if skipped_templates:
        skipped_sheet = workbook.create_sheet("SkippedTemplates")
        skipped_headers = [
            "page_id",
            "template_id",
            "case_type",
            "status",
            "reason",
            "missing_capabilities",
            "matched_capabilities",
        ]
        skipped_sheet.append(skipped_headers)
        for skipped in skipped_templates:
            skipped_sheet.append(
                [
                    skipped.get("page_id"),
                    skipped.get("template_id"),
                    skipped.get("case_type"),
                    skipped.get("status"),
                    skipped.get("reason"),
                    skipped.get("missing_capabilities"),
                    skipped.get("matched_capabilities"),
                ]
            )
        for cell in skipped_sheet[1]:
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        for row in skipped_sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for index, width in enumerate([34, 24, 24, 16, 36, 34, 72], start=1):
            skipped_sheet.column_dimensions[get_column_letter(index)].width = width
        skipped_sheet.freeze_panes = "A2"
        skipped_sheet.auto_filter.ref = skipped_sheet.dimensions

    detail = workbook.create_sheet("Checklist")
    headers = [
        "case_id",
        "parent_case_id",
        "viewpoint_id",
        "priority",
        "page_id",
        "line",
        "section",
        "element_kind",
        "locator",
        "test_title",
        "objective",
        "operation",
        "expected_result",
        "automation_mode",
        "case_type",
        "action_type",
        "test_data",
        "submit_locator",
        "expected_type",
        "expected_value",
        "pre_steps",
        "main_step",
        "evidence",
        "enabled",
        "generated_by",
        "matched_capabilities",
        "destructive",
    ]
    detail.append(headers)
    for index, case in enumerate(cases, start=1):
        case_id = case.case_id or f"{Path(case.page).stem or 'PAGE'}-{index:05d}"
        detail.append([
            case_id,
            case.parent_case_id,
            case.viewpoint_id,
            case.priority or case.severity,
            case.page,
            case.line,
            case.kind,
            case.kind,
            case.locator,
            case.title,
            case.objective,
            case.steps,
            case.expected,
            case.automation_mode,
            case.case_type,
            case.action_type,
            case.test_data,
            case.submit_locator,
            case.expected_type,
            case.expected_value,
            case.pre_steps,
            case.main_step,
            case.evidence,
            case.enabled,
            case.generated_by,
            case.matched_capabilities,
            case.destructive,
        ])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in detail[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row in detail.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    widths = [46, 46, 24, 10, 44, 8, 18, 18, 34, 42, 56, 64, 64, 18, 18, 18, 34, 34, 18, 28, 42, 42, 58, 10, 20, 72, 12]
    for index, width in enumerate(widths, start=1):
        detail.column_dimensions[get_column_letter(index)].width = width
    detail.freeze_panes = "A2"
    detail.auto_filter.ref = detail.dimensions

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def load_scan(path: Path) -> Dict[str, Any]:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise SystemExit(f"Invalid JSON input: {path} ({exc})") from exc


def write_report(
    scan_data: Dict[str, Any],
    output: Optional[Path],
    *,
    runtime_profile_dir: Path = DEFAULT_RUNTIME_PROFILE_DIR,
    include_runtime_profile_dir: bool = True,
) -> None:
    cases = generate_cases(
        scan_data,
        runtime_profile_dir=runtime_profile_dir,
        include_runtime_profile_dir=include_runtime_profile_dir,
    )
    if output and output.suffix.lower() == ".xlsx":
        write_excel(
            output,
            scan_data,
            cases,
            runtime_profile_dir=runtime_profile_dir,
            include_runtime_profile_dir=include_runtime_profile_dir,
        )
        return
    report = render_markdown(scan_data, cases)
    if output:
        output.parent.mkdir(parents=True, exist_ok=True)
        output.write_text(report, encoding="utf-8")
        return
    print(report)


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate a JSP migration test checklist from jsp_scanner/page_mapping JSON.")
    parser.add_argument("input", type=Path, help="elements.json or page_mapping.json generated by moonlight tools")
    parser.add_argument("-o", "--output", type=Path, help="Report output path. Use .md or .xlsx. Defaults to stdout Markdown.")
    parser.add_argument(
        "--runtime-profile-dir",
        type=Path,
        default=DEFAULT_RUNTIME_PROFILE_DIR,
        help="Runtime page profile JSON directory. Defaults to generated/valid/runtime_profile.",
    )
    parser.add_argument(
        "--include-runtime-profiles",
        action="store_true",
        help="Append all existing runtime_profile/*.json pages to the checklist.",
    )
    args = parser.parse_args()
    if not args.input.exists():
        raise SystemExit(f"Input JSON does not exist: {args.input}")
    write_report(
        load_scan(args.input),
        args.output,
        runtime_profile_dir=args.runtime_profile_dir,
        include_runtime_profile_dir=args.include_runtime_profiles,
    )


if __name__ == "__main__":
    main()
