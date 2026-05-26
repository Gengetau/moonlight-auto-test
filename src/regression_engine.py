import html
import json
import os
import re
import time
import glob
import fnmatch
from collections import Counter
from pathlib import Path, PureWindowsPath
from typing import Any, Dict, Iterable, List, Optional, Tuple
from urllib.parse import parse_qsl, urlencode, urlsplit, urljoin

from playwright.sync_api import Page, Error as PlaywrightError, TimeoutError as PlaywrightTimeoutError

from src.action_executor import _capture_state, execute_action, infer_semantic_action
from src.assert_engine import compare_visual_screenshot
from src.config_parser import Config
from src.route_navigator import RouteMapCatalog, RouteNavigator


NEGATIVE_CASE_TYPES = {
    "negative_js_error",
    "negative_network_abort",
    "negative_http_500",
    "negative_invalid_input",
    "negative_file_upload",
}


def _judge_compare_status(
    *,
    action_type: str,
    visual_status: Optional[str],
    url_match: bool,
    legacy_action_status: Optional[str] = None,
    new_action_status: Optional[str] = None,
    legacy_nav_status: Optional[str] = None,
    new_nav_status: Optional[str] = None,
) -> str:
    """Return PASS/WARN/DIFF/BLOCKED using action-aware comparison policy.

    DOM comparison is intentionally disabled. The regression result is judged
    by action status, normalized URL policy, and visual comparison only.
    """
    if "BLOCKED" in {
        legacy_action_status,
        new_action_status,
        legacy_nav_status,
        new_nav_status,
    }:
        return "BLOCKED"

    normalized_action = (action_type or "page_snapshot").lower()
    visual_required = normalized_action not in {"download", "close_window"}
    url_required = normalized_action in {"navigate", "page_snapshot"}

    if visual_status == "BLOCKED":
        return "BLOCKED"

    if visual_required and visual_status == "DIFF":
        return "DIFF"

    if url_required and not url_match:
        return "DIFF"

    if not url_match:
        return "WARN"

    return "PASS"


class RegressionEngine:
    """
    Risk-first Legacy/New regression runner backed by page_mapping.json.

    The engine opens equivalent pages in both environments, replays mapped
    locator-change actions, compares URL/visual state, and writes a
    side-by-side HTML report.
    """

    def __init__(
        self,
        mapping_path: str = "generated/valid/page_mapping.json",
        *,
        legacy_base_url: Optional[str] = None,
        new_base_url: Optional[str] = None,
        output_dir: str = "./output/regression",
        visual_threshold_percent: float = 0.1,
        timeout: int = 15000,
        checklist_path: Optional[str] = None,
        route_map_path: Optional[str] = None,
        force_route_map: bool = False,
        upload_file: Optional[str] = None,
        upload_profile_config: Optional[str] = None,
        include_semi_auto: bool = False,
        include_destructive: bool = False,
        include_negative: bool = False,
        negative_profile: Optional[str] = None,
    ) -> None:
        self.mapping_path = Path(mapping_path)
        self.checklist_path = self._resolve_checklist_path(checklist_path)
        self._last_checklist_debug: Dict[str, Any] = {}
        self._checklist_case_rows: Dict[str, List[Dict[str, Any]]] = {}
        self.legacy_base_url = legacy_base_url or Config.LEGACY_URL
        self.new_base_url = new_base_url or Config.NEW_URL
        self.output_dir = Path(output_dir)
        self.visual_threshold_percent = visual_threshold_percent
        self.timeout = timeout
        self.force_route_map = force_route_map
        self.upload_file = str(upload_file) if upload_file else None
        self.upload_profiles = self._load_upload_profiles(upload_profile_config)
        self.include_semi_auto = include_semi_auto
        self.include_destructive = include_destructive
        self.include_negative = include_negative
        self.negative_profiles = {
            item.strip().lower()
            for item in re.split(r"[\r\n,;]+", str(negative_profile or ""))
            if item.strip()
        }
        self.current_browser_name = ""
        self.mapping = self.load_mapping()
        self.route_map_catalog = RouteMapCatalog(self._route_map_paths(route_map_path))

    @staticmethod
    def _resolve_checklist_path(checklist_path: Optional[str]) -> Optional[Path]:
        if checklist_path:
            return Path(checklist_path)

        default_path = Path("generated/valid/migration_checklist.xlsx")
        return default_path if default_path.exists() else None

    @staticmethod
    def _load_upload_profiles(path: Optional[str]) -> List[Dict[str, Any]]:
        if not path:
            return []
        profile_path = Path(path)
        if not profile_path.exists():
            return []
        try:
            payload = json.loads(profile_path.read_text(encoding="utf-8"))
        except Exception:
            return []
        if isinstance(payload, list):
            profiles = payload
        else:
            profiles = payload.get("upload_profiles") if isinstance(payload, dict) else []
        return [item for item in profiles or [] if isinstance(item, dict)]

    def load_mapping(self) -> Dict[str, Any]:
        if not self.mapping_path.exists():
            raise FileNotFoundError(f"page_mapping.json not found: {self.mapping_path}")
        return json.loads(self.mapping_path.read_text(encoding="utf-8"))

    def select_pages(
        self,
        *,
        risk_only: bool = True,
        risk_levels: Optional[Iterable[str]] = None,
        limit: Optional[int] = None,
        target_page: Optional[str] = None,
    ) -> List[Dict[str, Any]]:
        """
        根据风险等级或 target_page 选择需要执行回归测试的页面。

        优先级：
        1. 指定 target_page 时，仅返回对应页面
        2. 否则按 risk_level 过滤并排序
        """

        # ------------------------------------------------------------------
        # 指定单页面执行
        # ------------------------------------------------------------------
        if target_page:
            normalized_target = self._target_page_name(target_page)

            pages = [
                page
                for page in self.mapping.get("page_mappings", [])
                if self._target_page_name(page.get("page_id"))
                == normalized_target
            ]

            if not pages:
                available = sorted(
                    self._target_page_name(page.get("page_id"))
                    for page in self.mapping.get("page_mappings", [])
                    if page.get("page_id")
                )

                sample = ", ".join(available[:10])

                suffix = (
                    f" Available pages include: {sample}"
                    if sample
                    else " Mapping contains no pages."
                )

                raise ValueError(
                    f"Target JSP page not found in mapping file "
                    f"{self.mapping_path}: {target_page}.{suffix}"
                )

            return pages[:1]

        # ------------------------------------------------------------------
        # 风险等级过滤
        # ------------------------------------------------------------------
        levels = list(
            risk_levels
            or (
                ["High", "Medium"]
                if risk_only
                else ["High", "Medium", "Low"]
            )
        )

        rank = {
            level.lower(): index
            for index, level in enumerate(levels)
        }

        pages = [
            page
            for page in self.mapping.get("page_mappings", [])
            if str(page.get("risk", "")).lower() in rank
        ]

        # ------------------------------------------------------------------
        # 按风险等级 + page_id 排序
        # ------------------------------------------------------------------
        pages.sort(
            key=lambda page: (
                rank[str(page.get("risk", "")).lower()],
                page.get("page_id", ""),
            )
        )

        # ------------------------------------------------------------------
        # limit 截断
        # ------------------------------------------------------------------
        return pages[:limit] if limit else pages


    def run(
        self,
        legacy_page: Page,
        new_page: Page,
        *,
        risk_only: bool = True,
        risk_levels: Optional[Iterable[str]] = None,
        limit: Optional[int] = None,
        target_page: Optional[str] = None,
        browser_name: str = "chrome",
        manual: bool = False,
    ) -> Dict[str, Any]:
        self.current_browser_name = browser_name
        self.output_dir.mkdir(parents=True, exist_ok=True)
        results: List[Dict[str, Any]] = []
        for page_index, mapping in enumerate(
            self.select_pages(
                risk_only=risk_only,
                risk_levels=risk_levels,
                limit=limit,
                target_page=target_page,
            ),
            start=1,
        ):
            results.extend(
                self._run_page_pair(
                    legacy_page,
                    new_page,
                    mapping,
                    page_index=page_index,
                    browser_name=browser_name,
                    manual=manual,
                )
            )

        report_path = self.render_report(results)
        overall_status = "PASS"
        if any(item["status"] == "BLOCKED" for item in results):
            overall_status = "BLOCKED"
        elif any(item["status"] == "DIFF" for item in results):
            overall_status = "DIFF"

        return {
            "status": overall_status,
            "results": results,
            "report_path": str(report_path),
            "summary": dict(Counter(item["status"] for item in results)),
        }

    def _run_page_pair(
        self,
        legacy_page: Page,
        new_page: Page,
        mapping: Dict[str, Any],
        *,
        page_index: int,
        browser_name: str,
        manual: bool = False,
    ) -> List[Dict[str, Any]]:
        page_id = mapping.get("page_id") or f"page_{page_index}"

        page_dir = (
            self.output_dir
            / f"{page_index:04d}_{self._safe_name(page_id)}"
        )
        page_dir.mkdir(parents=True, exist_ok=True)

        entry_url = (
            mapping.get("entry_url")
            or mapping.get("resolved_entry_url")
            or page_id
        )

        legacy_url = self._page_url(
            self.legacy_base_url,
            entry_url,
        )

        new_url = self._page_url(
            self.new_base_url,
            entry_url,
        )

        # ============================================================
        # MANUAL MODE
        # ============================================================
        if manual:
            print("\n[MANUAL MODE] 请手动操作浏览器并导航至目标页面:")
            print(f" - 期待的目标画面: {page_id}")
            print(f" - Legacy 入口: {legacy_url}")
            print(f" - New 入口:    {new_url}")

            def _interactive_takeover(
                page: Page,
                label: str,
            ) -> Page:
                print(
                    f"\n >>> [{label}] 准备就绪后，在此处按回车 [ENTER] 进行接管..."
                )

                print(
                    "     (注: 如果浏览器点击无响应，请按回车激活同步锁)"
                )

                while True:
                    # ------------------------------------------------
                    # 让 Playwright 持续处理 protocol 消息
                    # ------------------------------------------------
                    try:
                        page.wait_for_timeout(200)
                    except Exception:
                        pass

                    input_signal = input(
                        f" [{label} READY?] "
                        f"按回车扫描页面 (或输入 'q' 放弃): "
                    ).strip().lower()

                    if input_signal == "q":
                        raise InterruptedError(
                            "User cancelled manual takeover"
                        )

                    all_pages = page.context.pages

                    print(
                        f" [{label}] 探测到 {len(all_pages)} 个页面对象:"
                    )

                    match_idx = -1

                    # ------------------------------------------------
                    # 页面匹配优先级
                    # 1. 包含 JSP 名
                    # 2. 包含对应 Struts Action
                    # ------------------------------------------------
                    target_action = str(
                        mapping.get("entry_url")
                        or mapping.get("resolved_entry_url")
                        or page_id
                    ).lower()

                    for idx, page_item in enumerate(all_pages):
                        url_lower = page_item.url.lower()

                        print(
                            f"    [{idx}] "
                            f"{page_item.url[:120]}"
                        )

                        if (
                            page_id.lower() in url_lower
                            or target_action in url_lower
                        ):
                            match_idx = idx

                    # ------------------------------------------------
                    # 自动发现匹配页面
                    # ------------------------------------------------
                    if match_idx != -1:
                        print(
                            f" [SUCCESS] 发现潜在匹配页面: "
                            f"[{match_idx}]"
                        )

                        choice_idx = match_idx

                    # ------------------------------------------------
                    # 未匹配时允许人工指定
                    # ------------------------------------------------
                    else:
                        print(
                            f" [WARN] 未发现包含 "
                            f"'{page_id}' 或 Action "
                            f"'{target_action}' 的页面。"
                        )

                        raw_choice = input(
                            f" >>> 请输入页面索引 "
                            f"[0-{len(all_pages)-1}] "
                            f"手动指定，或直接回车重试: "
                        ).strip()

                        if (
                            raw_choice.isdigit()
                            and 0 <= int(raw_choice) < len(all_pages)
                        ):
                            choice_idx = int(raw_choice)
                        else:
                            continue

                    # ------------------------------------------------
                    # 接管目标页面
                    # ------------------------------------------------
                    target = all_pages[choice_idx]

                    try:
                        target.bring_to_front()

                        target.wait_for_load_state(
                            "domcontentloaded",
                            timeout=3000,
                        )

                        return target

                    except Exception as exc:
                        print(
                            f" [ERROR] 挂载失败 ({exc})，请重试。"
                        )

            # ========================================================
            # Legacy/New 手动接管
            # ========================================================
            legacy_page = _interactive_takeover(
                legacy_page,
                "Legacy",
            )

            new_page = _interactive_takeover(
                new_page,
                "New",
            )

            print(
                f" [INFO] 已重定向接管目标: "
                f"Legacy({legacy_page.url}) | "
                f"New({new_page.url})"
            )

            legacy_nav = {
                "status": "PASS",
                "url": legacy_page.url,
                "manual": True,
            }

            new_nav = {
                "status": "PASS",
                "url": new_page.url,
                "manual": True,
            }

            # ========================================================
            # 手动模式直接执行已接管页面
            # ========================================================
            return self._run_captured_page_pair(
                legacy_page,
                new_page,
                mapping,
                page_dir,
                browser_name,
                legacy_nav,
                new_nav,
                manual=True,
            )

        # ============================================================
        # AUTO MODE
        # ============================================================
        route = self.route_map_catalog.find_for_target(page_id)
        if self.force_route_map and route:
            navigator = RouteNavigator(
                self.route_map_catalog,
                timeout=self.timeout,
                browser_name=browser_name,
                upload_file=self.upload_file,
            )
            print(
                f"[{page_id}] FORCE ROUTE MAP navigation: "
                + json.dumps(
                    {
                        "route_id": route.get("route_id"),
                        "route_map_path": route.get("route_map_path"),
                    },
                    ensure_ascii=False,
                )
            )
            legacy_page, legacy_nav = navigator.navigate(
                legacy_page,
                entry_url=self.legacy_base_url,
                target_page=page_id,
                capture_dir=page_dir,
                side="legacy",
            )
            new_page, new_nav = navigator.navigate(
                new_page,
                entry_url=self.new_base_url,
                target_page=page_id,
                capture_dir=page_dir,
                side="new",
            )
            return self._run_captured_page_pair(
                legacy_page,
                new_page,
                mapping,
                page_dir,
                browser_name,
                legacy_nav,
                new_nav,
                manual=False,
            )

        legacy_nav = self._goto(
            legacy_page,
            legacy_url,
        )

        new_nav = self._goto(
            new_page,
            new_url,
        )

        direct_legacy_reached = legacy_nav.get("status") == "PASS" and self._page_matches_mapping(legacy_page, mapping)
        direct_new_reached = new_nav.get("status") == "PASS" and self._page_matches_mapping(new_page, mapping)

        if not direct_legacy_reached or not direct_new_reached:
            if route:
                navigator = RouteNavigator(
                    self.route_map_catalog,
                    timeout=self.timeout,
                    browser_name=browser_name,
                    upload_file=self.upload_file,
                )
                print(
                    f"[{page_id}] DIRECT NAVIGATION fallback to route map: "
                    + json.dumps(
                        {
                            "legacy_direct_reached": direct_legacy_reached,
                            "new_direct_reached": direct_new_reached,
                            "route_id": route.get("route_id"),
                            "route_map_path": route.get("route_map_path"),
                        },
                        ensure_ascii=False,
                    )
                )

                if not direct_legacy_reached:
                    legacy_page, legacy_nav = navigator.navigate(
                        legacy_page,
                        entry_url=self.legacy_base_url,
                        target_page=page_id,
                        capture_dir=page_dir,
                        side="legacy",
                    )
                else:
                    legacy_nav.update({"strategy": "direct_url", "target_reached": True})

                if not direct_new_reached:
                    new_page, new_nav = navigator.navigate(
                        new_page,
                        entry_url=self.new_base_url,
                        target_page=page_id,
                        capture_dir=page_dir,
                        side="new",
                    )
                else:
                    new_nav.update({"strategy": "direct_url", "target_reached": True})
            else:
                if not direct_legacy_reached:
                    legacy_nav.update(
                        {
                            "status": "BLOCKED",
                            "strategy": "direct_url",
                            "target_reached": False,
                            "reason": legacy_nav.get("reason") or f"Direct navigation did not reach target page and no route map exists: {page_id}",
                        }
                    )
                if not direct_new_reached:
                    new_nav.update(
                        {
                            "status": "BLOCKED",
                            "strategy": "direct_url",
                            "target_reached": False,
                            "reason": new_nav.get("reason") or f"Direct navigation did not reach target page and no route map exists: {page_id}",
                        }
                    )
        else:
            legacy_nav.update({"strategy": "direct_url", "target_reached": True})
            new_nav.update({"strategy": "direct_url", "target_reached": True})

        return self._run_captured_page_pair(
            legacy_page,
            new_page,
            mapping,
            page_dir,
            browser_name,
            legacy_nav,
            new_nav,
            manual=False,
        )


    @staticmethod
    def _skip_mapping_fallback_action(action: Dict[str, Any]) -> bool:
        kind = str(action.get("kind") or "").lower()
        action_hint = str(action.get("action_hint") or action.get("action_type") or "").lower()
        automation_mode = str(action.get("automation_mode") or "").lower()
        semantic_action = infer_semantic_action(action_hint or kind, action)

        # Source-only fallback does not know the business scenario. Bare forms
        # and terminal close-window buttons are only safe when provided by an
        # explicit checklist/executable case.
        if kind == "form" or automation_mode == "scenario_only" or semantic_action == "submit":
            return True
        if action_hint in {"close_window", "window_close"} or kind == "close_window":
            return True
        return False

    def _build_action_plan(self, page_id: str, mapping: Dict[str, Any]) -> Tuple[List[Dict[str, Any]], str]:
        """
        Build executable runtime action plan.

        Priority:
        1. Checklist Excel cases with automation_mode=auto
        2. page_mapping.json executable_cases
        3. locator_changes + full_action_steps fallback

        Form-only submit steps are intentionally skipped in fallback mode because
        they must be combined with related inputs/files into a scenario such as
        upload_submit.
        """
        checklist_cases = self._load_checklist_cases(page_id, mapping)
        if checklist_cases:
            return checklist_cases, "checklist"

        executable_cases = mapping.get("executable_cases") or []
        if executable_cases:
            return [self._normalize_action_case(case) for case in executable_cases], "executable_cases"

        actions: List[Dict[str, Any]] = []
        for change in mapping.get("locator_changes", []) or []:
            if self._skip_mapping_fallback_action(change):
                continue
            actions.append(self._normalize_action_case(change))

        for step in mapping.get("full_action_steps", []) or []:
            action_hint = str(step.get("action_hint") or "").lower()

            if self._skip_mapping_fallback_action(step):
                continue

            legacy_loc = step.get("legacy_locator") or step.get("locator")
            new_loc = step.get("new_locator") or legacy_loc
            merged = {
                "legacy_locator": legacy_loc,
                "new_locator": new_loc,
                "label": step.get("label"),
                "semantic_key": step.get("semantic_key"),
                "kind": step.get("kind"),
                "action_hint": step.get("action_hint"),
                "action_type": step.get("action_type") or action_hint,
                "raw": step.get("raw"),
                "attributes": step.get("attributes", {}),
                "line": step.get("line"),
                "value": step.get("value") or step.get("test_data"),
                "source": "full_action_steps",
            }
            actions.append(self._normalize_action_case(merged))

        return self._dedupe_runtime_actions(actions), "mapping_fallback"

    def _load_checklist_cases(self, page_id: str, mapping: Optional[Dict[str, Any]] = None) -> List[Dict[str, Any]]:
        """
        Load executable cases from generated migration checklist.

        This is intentionally tolerant. If the Excel does not yet contain
        machine-readable columns, it returns [] and the engine falls back to
        mapping executable_cases/full_action_steps.
        """
        target_name = self._target_page_name(page_id)
        debug: Dict[str, Any] = {
            "target": target_name,
            "path": str(self.checklist_path) if self.checklist_path else None,
            "status": "not_configured",
        }
        self._last_checklist_debug = debug
        coverage_rows: List[Dict[str, Any]] = []
        self._checklist_case_rows[target_name] = coverage_rows

        if not self.checklist_path:
            return []

        if not self.checklist_path.exists():
            debug["status"] = "missing_file"
            return []

        try:
            from openpyxl import load_workbook
        except Exception as exc:
            debug["status"] = "openpyxl_unavailable"
            debug["error"] = str(exc)
            return []

        try:
            wb = load_workbook(self.checklist_path, data_only=True, read_only=True)
            sheet = wb["Checklist"] if "Checklist" in wb.sheetnames else wb[wb.sheetnames[0]]
            rows = list(sheet.iter_rows(values_only=True))
        except Exception as exc:
            debug["status"] = "read_error"
            debug["error"] = str(exc)
            return []

        if not rows:
            debug["status"] = "empty_sheet"
            return []

        headers = [str(value or "").strip() for value in rows[0]]
        header_map = {header.lower(): idx for idx, header in enumerate(headers) if header}

        def col(*names: str) -> Optional[int]:
            for name in names:
                key = name.strip().lower()
                if key in header_map:
                    return header_map[key]
            return None

        page_col = col("page_id", "page", "页面", "画面", "ページ", "対象画面")
        mode_col = col("automation_mode", "自动化模式", "自動化モード", "自動化", "automation")
        case_type_col = col("case_type", "ケース種別", "用例类型", "case")
        action_type_col = col("action_type", "アクション", "动作类型", "操作类型")
        locator_col = col("locator", "selector", "定位器", "セレクタ")
        legacy_locator_col = col("legacy_locator", "legacy selector", "移行前locator")
        new_locator_col = col("new_locator", "new selector", "移行後locator")
        submit_locator_col = col("submit_locator", "submit selector", "提交locator")
        test_data_col = col("test_data", "value", "测试数据", "テストデータ")
        operation_col = col("operation", "操作", "操作内容")
        main_step_col = col("main_step", "main step")
        expected_type_col = col("expected_type", "期待種別")
        destructive_col = col("destructive", "破壊的", "destructive?")
        generated_by_col = col("generated_by")
        enabled_col = col("enabled", "enable", "有效", "有効")
        case_id_col = col("case_id", "id", "no", "項目no")
        title_col = col("title", "test_title", "test_viewpoint", "テスト観点", "测试项目", "用例")

        debug["sheet"] = sheet.title
        debug["columns"] = {
            "page_col": page_col,
            "mode_col": mode_col,
            "case_type_col": case_type_col,
            "action_type_col": action_type_col,
            "locator_col": locator_col,
            "submit_locator_col": submit_locator_col,
            "operation_col": operation_col,
            "main_step_col": main_step_col,
            "expected_type_col": expected_type_col,
            "destructive_col": destructive_col,
            "enabled_col": enabled_col,
        }

        if page_col is None or mode_col is None:
            debug["status"] = "missing_required_columns"
            return []

        cases: List[Dict[str, Any]] = []
        stats: Counter[str] = Counter()
        samples: List[Dict[str, str]] = []
        allowed_modes = {"auto", "automated", "true", "yes", "y", "1", "自動", "自动"}
        semi_auto_modes = {"semi-auto", "semiauto", "semi auto", "半自动", "半自動"}

        def add_sample(kind: str, **values: str) -> None:
            if len(samples) < 5:
                samples.append({"kind": kind, **values})

        def coverage_row(row_page: str, case_id: str, test_title: str, automation_mode: str, destructive_value: str) -> Dict[str, Any]:
            row_info = {
                "page_id": row_page,
                "case_id": case_id,
                "test_title": test_title,
                "automation_mode": automation_mode,
                "destructive": destructive_value or "false",
                "excluded_reason": "",
            }
            coverage_rows.append(row_info)
            return row_info

        for row in rows[1:]:
            def cell(idx: Optional[int]) -> str:
                if idx is None or idx >= len(row):
                    return ""
                value = row[idx]
                return "" if value is None else str(value).strip()

            row_page = cell(page_col)
            if self._target_page_name(row_page) != target_name:
                stats["page_not_matched"] += 1
                continue

            stats["page_matched"] += 1
            case_id = cell(case_id_col)
            test_title = cell(title_col)
            mode = cell(mode_col)
            destructive_value = cell(destructive_col) or "false"
            checklist_coverage = coverage_row(row_page, case_id, test_title, mode, destructive_value)

            enabled = cell(enabled_col).lower()
            if enabled in {"false", "0", "no", "n", "disabled", "off", "無効", "否"}:
                stats["disabled"] += 1
                add_sample("disabled", page=row_page, enabled=enabled)
                checklist_coverage["excluded_reason"] = f"enabled={enabled or '<blank>'}"
                continue

            mode = mode.lower()
            mode_allowed = mode in allowed_modes or mode.startswith("auto")
            if not mode_allowed and self.include_semi_auto:
                mode_allowed = mode in semi_auto_modes or mode.startswith("semi")
            if not mode_allowed:
                stats[f"mode_rejected:{mode or '<blank>'}"] += 1
                add_sample("mode_rejected", page=row_page, mode=mode, case_id=case_id)
                if mode in semi_auto_modes or mode.startswith("semi"):
                    checklist_coverage["excluded_reason"] = f"automation_mode={mode or '<blank>'}; requires --include-semi-auto"
                else:
                    checklist_coverage["excluded_reason"] = f"automation_mode={mode or '<blank>'} is not executable"
                continue

            stats["auto_matched" if mode.startswith("auto") or mode in allowed_modes else "semi_auto_matched"] += 1

            case_type = cell(case_type_col) or cell(action_type_col) or "click"
            action_type = cell(action_type_col) or case_type
            action_lower = str(action_type or case_type).strip().lower()
            case_lower = str(case_type or action_type).strip().lower()
            locator = cell(locator_col)
            legacy_locator = cell(legacy_locator_col) or locator
            new_locator = cell(new_locator_col) or locator or legacy_locator
            submit_locator = cell(submit_locator_col)
            test_data = cell(test_data_col)
            label = test_title or case_id or case_type
            operation = cell(operation_col)
            main_step_text = cell(main_step_col)
            expected_type = cell(expected_type_col)
            destructive = self._truthy(destructive_value)
            negative_case = self._is_negative_case(case_type, action_type)

            if destructive and not self.include_destructive:
                stats["destructive_rejected"] += 1
                add_sample("destructive_rejected", page=row_page, case_id=case_id, case_type=case_type)
                checklist_coverage["excluded_reason"] = "destructive=true; requires --include-destructive"
                continue

            if negative_case and not self.include_negative:
                stats["negative_rejected"] += 1
                add_sample("negative_rejected", page=row_page, case_id=case_id, case_type=case_type)
                checklist_coverage["excluded_reason"] = "negative case; requires --include-negative"
                continue
            if negative_case and self.include_negative and not self._negative_profile_enabled(case_type, action_type, cell(generated_by_col)):
                stats["negative_profile_rejected"] += 1
                add_sample("negative_profile_rejected", page=row_page, case_id=case_id, case_type=case_type)
                checklist_coverage["excluded_reason"] = "negative case; --negative-profile did not match"
                continue

            is_upload_submit = self._is_upload_submit_case(
                case_type=case_type,
                action_type=action_type,
                label=label,
                operation=operation,
                submit_locator=submit_locator,
                main_step_text=main_step_text,
            )

            if is_upload_submit:
                if not locator and not legacy_locator:
                    stats["locator_missing_upload"] += 1
                    add_sample("locator_missing_upload", page=row_page, case_id=case_id, mode=mode)
                    checklist_coverage["excluded_reason"] = "upload_submit case has no upload locator"
                    continue
                main_step = self._parse_step_json(main_step_text)
                main_locator = submit_locator or main_step.get("submit_locator") or main_step.get("locator") or self._submit_locator_from_mapping(mapping, upload_locator=locator)
                main_action_type = main_step.get("action_type") or ("click" if main_locator and not str(main_locator).strip().lower().startswith("form") else "submit")
                case: Dict[str, Any] = {
                    "case_type": "upload_submit",
                    "action_type": "upload_submit",
                    "label": label,
                    "page_id": row_page,
                    "source": "checklist",
                    "expected_type": expected_type,
                    "destructive": str(destructive).lower(),
                    "pre_steps": [
                        {
                            "action_type": "upload",
                            "legacy_locator": legacy_locator or locator,
                            "new_locator": new_locator or locator,
                            "locator": locator or legacy_locator,
                            "value": test_data,
                        }
                    ],
                    "main_step": {
                        "action_type": main_action_type,
                        "legacy_locator": main_locator,
                        "new_locator": main_locator,
                        "locator": main_locator,
                        "submit_locator": submit_locator,
                    },
                }
                cases.append(self._normalize_action_case(case))
                stats["loadable_upload_submit"] += 1
            else:
                if action_lower in {"snapshot", "page_snapshot", "visual_check", "wait", "initial_display"}:
                    locator = locator or "__page__"
                    legacy_locator = legacy_locator or locator
                    new_locator = new_locator or locator
                elif not legacy_locator and not new_locator:
                    stats["locator_missing"] += 1
                    add_sample("locator_missing", page=row_page, case_id=case_id, mode=mode)
                    checklist_coverage["excluded_reason"] = "locator is missing"
                    continue
                cases.append(
                    self._normalize_action_case(
                        {
                            "case_type": case_type,
                            "action_type": action_type,
                            "label": label,
                            "page_id": row_page,
                            "legacy_locator": legacy_locator,
                            "new_locator": new_locator or legacy_locator,
                            "locator": locator,
                            "value": test_data,
                            "expected_type": expected_type,
                            "destructive": str(destructive).lower(),
                            "source": "checklist",
                        }
                    )
                )
                stats["loadable_action"] += 1

        debug["status"] = "loaded" if cases else "no_cases"
        debug["loaded"] = len(cases)
        debug["stats"] = dict(stats)
        debug["samples"] = samples
        return cases

    @staticmethod
    def _truthy(value: Any) -> bool:
        return str(value or "").strip().lower() in {"true", "1", "yes", "y", "on", "破壊", "対象", "是"}

    @staticmethod
    def _is_negative_case(case_type: Any, action_type: Any) -> bool:
        text = " ".join(str(value or "").lower() for value in (case_type, action_type))
        return "negative" in text or text.startswith("error_") or any(case in text for case in NEGATIVE_CASE_TYPES)

    def _negative_profile_enabled(self, case_type: Any, action_type: Any, generated_by: Any = "") -> bool:
        if not self.negative_profiles:
            return False
        haystack = " ".join(str(value or "").lower() for value in (case_type, action_type, generated_by))
        return any(profile in haystack for profile in self.negative_profiles)

    @staticmethod
    def _is_upload_submit_case(
        *,
        case_type: Any,
        action_type: Any,
        label: Any,
        operation: Any,
        submit_locator: Any,
        main_step_text: Any,
    ) -> bool:
        case_lower = str(case_type or "").lower()
        action_lower = str(action_type or "").lower()
        label_text = str(label or "")
        operation_lower = str(operation or "").lower()
        main_step_lower = str(main_step_text or "").lower()
        if case_lower == "upload_submit" or action_lower == "upload_submit":
            return True
        if "アップロード確認" in label_text:
            return True
        if submit_locator:
            return True
        upload_tokens = ("upload", "アップロード", "file", "ファイル")
        submit_tokens = ("submit", "click", "確認", "送信", "押下")
        if any(token in operation_lower for token in upload_tokens) and any(token in operation_lower for token in submit_tokens):
            return True
        if any(token in main_step_lower for token in ("submit", "click", "submit_locator")):
            return True
        return False

    @staticmethod
    def _parse_step_json(value: Any) -> Dict[str, Any]:
        if not value:
            return {}
        try:
            parsed = json.loads(str(value))
        except Exception:
            return {}
        return parsed if isinstance(parsed, dict) else {}

    @staticmethod
    def _submit_locator_from_mapping(mapping: Optional[Dict[str, Any]], upload_locator: Any = "") -> str:
        if not mapping:
            return ""
        candidates = list(mapping.get("executable_cases") or []) + list(mapping.get("locator_changes") or []) + list(mapping.get("full_action_steps") or [])
        for item in candidates:
            locator = item.get("locator") or item.get("legacy_locator") or item.get("new_locator")
            evidence = " ".join(str(item.get(key) or "").lower() for key in ("action_hint", "action_type", "kind", "label", "raw", "semantic_key", "locator"))
            if locator and any(token in evidence for token in ("submit", "upload", "confirm", "確認", "アップロード")):
                return str(locator)
        return ""

    def _normalize_action_case(self, action: Dict[str, Any]) -> Dict[str, Any]:
        normalized = dict(action)
        if "main_step" in normalized or "pre_steps" in normalized:
            normalized.setdefault("case_type", normalized.get("action_type") or "scenario")
            normalized.setdefault("action_type", normalized.get("case_type"))
            normalized.setdefault("kind", normalized.get("case_type"))
            normalized.setdefault("label", normalized.get("label") or normalized.get("case_type"))
            return normalized

        normalized.setdefault("legacy_locator", normalized.get("locator"))
        normalized.setdefault("new_locator", normalized.get("legacy_locator") or normalized.get("locator"))
        normalized.setdefault("action_type", normalized.get("action_hint") or normalized.get("kind") or "click")
        normalized.setdefault("case_type", normalized.get("action_type"))
        normalized.setdefault("label", normalized.get("label") or normalized.get("semantic_key") or normalized.get("legacy_locator") or normalized.get("case_type"))
        return normalized

    def _dedupe_runtime_actions(self, actions: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        seen = set()
        unique: List[Dict[str, Any]] = []
        for action in actions:
            if action.get("pre_steps") or action.get("main_step"):
                key = (
                    action.get("case_type"),
                    action.get("label"),
                    json.dumps(action.get("pre_steps", []), sort_keys=True, ensure_ascii=False, default=str),
                    json.dumps(action.get("main_step", {}), sort_keys=True, ensure_ascii=False, default=str),
                )
            else:
                key = (
                    action.get("case_type"),
                    action.get("legacy_locator"),
                    action.get("new_locator"),
                    action.get("label"),
                )
            if key in seen:
                continue
            seen.add(key)
            unique.append(action)
        return unique

    @staticmethod
    def _action_side_locator(action: Dict[str, Any], side: str) -> Optional[str]:
        side_key = f"{side}_locator"
        return action.get(side_key) or action.get("locator") or action.get("selector")

    @staticmethod
    def _upload_value_is_placeholder(value: Any) -> bool:
        text = str(value or "").strip()
        if not text:
            return True
        lowered = text.lower()
        return (
            lowered.startswith("${")
            or lowered in {"$upload_file", "upload_file"}
            or "fakepath" in lowered
            or lowered in {"${upload_invalid_file}", "${upload_empty_file}", "${upload_large_file}"}
        )

    @staticmethod
    def _existing_upload_value(value: Any) -> Optional[Any]:
        if isinstance(value, list):
            paths = [str(item) for item in value if Path(str(item)).exists()]
            return paths if paths and len(paths) == len(value) else None
        text = str(value or "").strip()
        if text and Path(text).exists():
            return text
        return None

    def _resolve_upload_value(self, value: Any, action_case: Dict[str, Any], step: Dict[str, Any], locator: Any) -> Any:
        existing = self._existing_upload_value(value)
        if existing is not None:
            return existing

        profile_value = self._upload_value_from_profiles(action_case, step, locator)
        if profile_value:
            return profile_value

        if self.upload_file:
            return self.upload_file

        return "" if self._upload_value_is_placeholder(value) else value

    def _upload_value_from_profiles(self, action_case: Dict[str, Any], step: Dict[str, Any], locator: Any) -> Optional[Any]:
        page_id = str(action_case.get("page_id") or "").lower()
        case_type = str(action_case.get("case_type") or step.get("case_type") or step.get("action_type") or "").lower()
        locator_text = str(locator or step.get("locator") or "").strip()
        negative_case = self._is_negative_case(action_case.get("case_type"), step.get("action_type") or action_case.get("action_type"))

        def matches(profile: Dict[str, Any]) -> bool:
            profile_negative = self._truthy(profile.get("negative"))
            if profile_negative and not negative_case:
                return False
            page_patterns = profile.get("page_patterns") or []
            if page_patterns and not any(fnmatch.fnmatch(page_id, str(pattern).lower()) for pattern in page_patterns):
                return False
            case_types = [str(item).lower() for item in profile.get("case_types") or [] if str(item).strip()]
            if case_types and not any(fnmatch.fnmatch(case_type, pattern) for pattern in case_types):
                return False
            profile_locator = str(profile.get("locator") or "").strip()
            if profile_locator and profile_locator != locator_text and profile_locator not in locator_text and locator_text not in profile_locator:
                return False
            return True

        for profile in self.upload_profiles:
            if not matches(profile):
                continue
            files = profile.get("files")
            if isinstance(files, list):
                existing = self._existing_upload_value(files)
                if existing:
                    return existing
            file_value = profile.get("file")
            existing = self._existing_upload_value(file_value)
            if existing:
                return existing
        return None

    def _execute_action_case(
        self,
        page: Page,
        action_case: Dict[str, Any],
        *,
        side: str,
        browser_name: str,
        capture_dir: Path,
        test_id: str,
    ) -> Dict[str, Any]:
        """
        Execute a single action or a scenario composed of pre_steps + main_step.
        Example: upload_submit = set_input_files(...) then submit form.
        """
        steps: List[Dict[str, Any]] = []
        if action_case.get("pre_steps") or action_case.get("main_step"):
            steps.extend(action_case.get("pre_steps") or [])
            main_step = action_case.get("main_step")
            if main_step:
                steps.append(main_step)
        else:
            steps.append(action_case)

        last_result: Optional[Dict[str, Any]] = None
        executed = []

        for index, step in enumerate(steps, start=1):
            step = dict(step)
            action_type = step.get("action_type") or step.get("action_hint") or step.get("kind") or action_case.get("action_type") or "click"
            locator = self._action_side_locator(step, side)
            value = step.get("value") or step.get("test_data") or action_case.get("value")
            if (
                str(action_case.get("case_type") or "").lower() == "upload_submit"
                and str(action_type or "").lower() == "submit"
                and locator
                and not str(locator).strip().lower().startswith("form")
            ):
                action_type = "click"
            semantic_action = infer_semantic_action(action_type, step)
            action_lower = str(action_type or semantic_action).strip().lower()
            if semantic_action == "upload":
                value = self._resolve_upload_value(value, action_case, step, locator)

            if action_lower in {"snapshot", "page_snapshot", "visual_check", "wait"}:
                result = {
                    "status": "PASS",
                    "action_type": action_type,
                    "semantic_action": "wait",
                    "state": _capture_state(page, capture_dir, f"{test_id}_step{index}"),
                }
                executed.append(
                    {
                        "index": index,
                        "action_type": action_type,
                        "locator": locator or "__page__",
                        "status": result.get("status"),
                    }
                )
                last_result = result
                continue

            # If the main submit locator is missing, try to submit nearest form from the first upload field.
            if not locator and semantic_action == "submit":
                locator = action_case.get("submit_locator") or step.get("submit_locator")
                if not locator:
                    locator = "form"

            if not locator and semantic_action not in {"goto"}:
                return {
                    "status": "BLOCKED",
                    "reason": f"Missing locator for scenario step {index}: {action_type}",
                    "action_type": action_type,
                    "state": _capture_state(page, capture_dir, f"{test_id}_step{index}_blocked"),
                    "executed_steps": executed,
                }

            result = execute_action(
                page,
                action_type,
                locator,
                value,
                browser_name=browser_name,
                capture_dir=capture_dir,
                test_id=f"{test_id}_step{index}",
                timeout=self.timeout,
                action_context={**action_case, **step, "locator": locator},
            )
            executed.append(
                {
                    "index": index,
                    "action_type": action_type,
                    "locator": locator,
                    "status": result.get("status"),
                    "reason": result.get("reason"),
                    "upload_file": result.get("upload_file") if semantic_action == "upload" else None,
                    "submit_locator": locator if str(action_case.get("case_type") or "").lower() == "upload_submit" and index == len(steps) else None,
                }
            )
            last_result = result

            if result.get("status") != "PASS":
                break

        if last_result is None:
            return {
                "status": "BLOCKED",
                "reason": "No executable steps in action case",
                "state": _capture_state(page, capture_dir, f"{test_id}_empty"),
                "executed_steps": executed,
            }

        last_result = dict(last_result)
        last_result["case_type"] = action_case.get("case_type")
        last_result["executed_steps"] = executed
        return last_result

    @staticmethod
    def _executed_step_field(action_result: Dict[str, Any], field: str) -> Optional[Any]:
        for step in action_result.get("executed_steps") or []:
            value = step.get(field)
            if value:
                return value
        return None


    def _run_captured_page_pair(
        self,
        legacy_page: Page,
        new_page: Page,
        mapping: Dict[str, Any],
        page_dir: Path,
        browser_name: str,
        legacy_nav: Dict[str, Any],
        new_nav: Dict[str, Any],
        manual: bool,
    ) -> List[Dict[str, Any]]:
        page_id = mapping.get("page_id") or "unknown"
        results: List[Dict[str, Any]] = []

        legacy_state = _capture_state(legacy_page, page_dir, "00_legacy_initial")
        new_state = _capture_state(new_page, page_dir, "00_new_initial")
        results.append(
            self._compare_state(
                page_id,
                mapping.get("risk"),
                "page_snapshot",
                legacy_state,
                new_state,
                page_dir / "00_diff.png",
                legacy_nav=legacy_nav,
                new_nav=new_nav,
            )
        )

        if legacy_nav.get("status") != "PASS" or new_nav.get("status") != "PASS":
            print(
                f"[{page_id}] ROUTE NAVIGATION blocked; skip action plan: "
                + json.dumps(
                    {
                        "legacy_status": legacy_nav.get("status"),
                        "legacy_reason": legacy_nav.get("reason"),
                        "new_status": new_nav.get("status"),
                        "new_reason": new_nav.get("reason"),
                    },
                    ensure_ascii=False,
                    default=str,
                )
            )
            return results

        for blocked in mapping.get("missing_legacy_elements", []):
            if self._is_dynamic_jsp_row_control(blocked):
                print(
                    f"[{page_id}] SKIP static dynamic row control: "
                    + json.dumps(
                        {
                            "label": blocked.get("label") or blocked.get("key"),
                            "locator": blocked.get("locator"),
                            "reason": "covered by runtime/checklist CRUD action",
                        },
                        ensure_ascii=False,
                    )
                )
                continue
            missing_status, missing_reason = self._missing_legacy_element_status(blocked)
            results.append(
                {
                    "page_id": page_id,
                    "risk": mapping.get("risk"),
                    "action": blocked.get("label") or blocked.get("key") or "missing_legacy_element",
                    "action_type": infer_semantic_action(blocked.get("action_hint") or blocked.get("kind"), blocked),
                    "status": missing_status,
                    "reason": missing_reason,
                    "legacy_locator": blocked.get("locator"),
                    "new_locator": None,
                    "legacy_screenshot": legacy_state.get("screenshot"),
                    "new_screenshot": new_state.get("screenshot"),
                    "legacy_frame": legacy_state.get("target_frame"),
                    "new_frame": new_state.get("target_frame"),
                    "frame_candidates": {
                        "legacy": legacy_state.get("frame_candidates", []),
                        "new": new_state.get("frame_candidates", []),
                    },
                }
            )

        target_actions, plan_source = self._build_action_plan(page_id, mapping)
        print(
            f"[{page_id}] ACTION PLAN SOURCE: "
            + json.dumps(
                {
                    "source": plan_source,
                    "planned": len(target_actions),
                    "locator_changes": len(mapping.get("locator_changes", []) or []),
                    "executable_cases": len(mapping.get("executable_cases", []) or []),
                    "full_action_steps": len(mapping.get("full_action_steps", []) or []),
                    "checklist_path": str(self.checklist_path) if self.checklist_path else None,
                    "checklist_debug": self._last_checklist_debug,
                    "kinds": dict(Counter(action.get("kind") or action.get("case_type") or "unknown" for action in target_actions)),
                },
                ensure_ascii=False,
            )
        )

        for action_index, action_case in enumerate(target_actions, start=1):
            action_name = action_case.get("label") or action_case.get("semantic_key") or action_case.get("case_type") or f"action_{action_index}"
            action_type = action_case.get("case_type") or action_case.get("action_type") or action_case.get("action_hint") or action_case.get("kind") or "click"
            semantic_action = infer_semantic_action(action_type, action_case)
            legacy_locator = self._action_side_locator(action_case, "legacy")
            new_locator = self._action_side_locator(action_case, "new")
            action_file_id = self._safe_name(f"{action_index:02d}_{action_name}")

            if action_case.get("pre_steps") or action_case.get("main_step"):
                # Use first concrete step locator for report display.
                steps_for_display = list(action_case.get("pre_steps") or [])
                if action_case.get("main_step"):
                    steps_for_display.append(action_case.get("main_step"))
                for step in steps_for_display:
                    legacy_locator = legacy_locator or self._action_side_locator(step, "legacy")
                    new_locator = new_locator or self._action_side_locator(step, "new")

            print(
                f"[{page_id}] START action {action_index}/{len(target_actions)}: "
                + json.dumps(
                    {
                        "action": action_name,
                        "action_type": action_type,
                        "semantic_action": semantic_action,
                        "legacy_locator": legacy_locator,
                        "new_locator": new_locator,
                        "source": action_case.get("source") or plan_source,
                    },
                    ensure_ascii=False,
                )
            )

            if not legacy_locator and not action_case.get("pre_steps") and not action_case.get("main_step"):
                results.append(
                    {
                        "page_id": page_id,
                        "risk": mapping.get("risk"),
                        "action": action_name,
                        "action_type": semantic_action,
                        "status": "BLOCKED",
                        "reason": "legacy_locator is missing",
                        "legacy_locator": legacy_locator,
                        "new_locator": new_locator,
                    }
                )
                continue

            database_operation = self._database_operation_kind(action_case, action_type, semantic_action)
            legacy_before_state: Optional[Dict[str, Any]] = None
            new_before_state: Optional[Dict[str, Any]] = None
            if database_operation:
                legacy_before_state = _capture_state(legacy_page, page_dir, f"{action_file_id}_legacy_before")
                new_before_state = _capture_state(new_page, page_dir, f"{action_file_id}_new_before")

            legacy_action = self._execute_action_case(
                legacy_page,
                action_case,
                side="legacy",
                browser_name=browser_name,
                capture_dir=page_dir,
                test_id=f"{action_file_id}_legacy",
            )
            new_action = self._execute_action_case(
                new_page,
                action_case,
                side="new",
                browser_name=browser_name,
                capture_dir=page_dir,
                test_id=f"{action_file_id}_new",
            )

            print(
                f"[{page_id}] END action: "
                + json.dumps(
                    {
                        "action": action_name,
                        "legacy_status": legacy_action.get("status"),
                        "new_status": new_action.get("status"),
                        "legacy_reason": legacy_action.get("reason"),
                        "new_reason": new_action.get("reason"),
                        "legacy_steps": legacy_action.get("executed_steps"),
                        "new_steps": new_action.get("executed_steps"),
                    },
                    ensure_ascii=False,
                    default=str,
                )
            )

            if database_operation:
                compared = self._compare_database_operation(
                    page_id,
                    mapping.get("risk"),
                    action_name,
                    action_type,
                    database_operation,
                    legacy_before_state or {},
                    new_before_state or {},
                    legacy_action,
                    new_action,
                    page_dir,
                    action_file_id,
                )
            else:
                compared = self._compare_state(
                    page_id,
                    mapping.get("risk"),
                    action_name,
                    legacy_action.get("state") or {},
                    new_action.get("state") or {},
                    page_dir / f"{action_index:02d}_diff.png",
                    legacy_action=legacy_action,
                    new_action=new_action,
                    action_type=action_type,
                )
            compared.update(
                {
                    "action_type": action_type,
                    "legacy_locator": legacy_locator,
                    "new_locator": new_locator,
                    "upload_file": self._executed_step_field(legacy_action, "upload_file")
                    or self._executed_step_field(new_action, "upload_file"),
                    "submit_locator": self._executed_step_field(legacy_action, "submit_locator")
                    or self._executed_step_field(new_action, "submit_locator"),
                    "legacy_after_url": legacy_action.get("after_url"),
                    "new_after_url": new_action.get("after_url"),
                    "navigation_detected": bool(legacy_action.get("navigation_detected") or new_action.get("navigation_detected")),
                    "popup_detected": bool(legacy_action.get("popup_detected") or new_action.get("popup_detected")),
                    "frame_changed": bool(legacy_action.get("frame_changed") or new_action.get("frame_changed")),
                    "validation_only": bool(legacy_action.get("validation_only") and new_action.get("validation_only")),
                    "legacy_action": legacy_action,
                    "new_action": new_action,
                    "plan_source": plan_source,
                }
            )
            should_reopen_target = self._requires_target_reopen_after_action(
                action_case,
                action_type,
                semantic_action,
                legacy_action,
                new_action,
            )
            if should_reopen_target:
                legacy_page, new_page, reopen_result = self._reopen_target_pair(
                    legacy_page,
                    new_page,
                    mapping,
                    page_dir,
                    browser_name,
                    reason=f"after action {action_index}: {action_name}",
                )
                compared["post_action_reopen"] = reopen_result
                if reopen_result.get("status") != "PASS":
                    compared["status"] = "BLOCKED"
                    compared["reason"] = reopen_result.get("reason") or "Failed to reopen target page after leaving/closing action."
            print(
                f"[{page_id}] COMPARE result: "
                + json.dumps(
                    {
                        "action": action_name,
                        "status": compared.get("status"),
                        "url_match": compared.get("url_match"),
                        "visual_status": (compared.get("visual") or {}).get("status"),
                        "visual_diff_percent": (compared.get("visual") or {}).get("diff_percent"),
                        "comparison_mode": compared.get("comparison_mode"),
                        "database_operation": compared.get("database_operation"),
                        "post_action_reopen_status": (compared.get("post_action_reopen") or {}).get("status"),
                    },
                    ensure_ascii=False,
                    default=str,
                )
            )
            results.append(compared)
            if should_reopen_target and compared.get("status") == "BLOCKED":
                break

        return results

    @staticmethod
    def _is_dynamic_jsp_row_control(blocked: Dict[str, Any]) -> bool:
        evidence = " ".join(
            str(blocked.get(key) or "")
            for key in ("kind", "key", "label", "locator", "raw", "action", "semantic_key")
        ).lower()
        return (
            ("<bean:" in evidence or "<logic:" in evidence or "bean:write" in evidence)
            and any(marker in evidence for marker in ("delete", "削除", "button", "onclick"))
        )

    @staticmethod
    def _missing_legacy_element_status(blocked: Dict[str, Any]) -> Tuple[str, str]:
        return "BLOCKED", "Legacy element has no mapped New equivalent"

    @staticmethod
    def _database_operation_kind(action_case: Dict[str, Any], action_type: Any, semantic_action: Any) -> Optional[str]:
        normalized_type = str(action_type or "").strip().lower()
        if normalized_type in {
            "snapshot",
            "page_snapshot",
            "visual_check",
            "wait",
            "initial_display",
            "result_table_verify",
            "download_template",
            "download",
            "close_window",
            "back_action",
            "link_navigation",
            "upload_select",
            "upload_without_file",
        }:
            return None

        evidence = " ".join(
            str(value or "")
            for value in (
                action_case.get("case_type"),
                action_case.get("action_type"),
                action_case.get("label"),
                action_case.get("semantic_key"),
                action_case.get("locator"),
                action_case.get("legacy_locator"),
                action_case.get("new_locator"),
                action_case.get("submit_locator"),
                action_case.get("expected_type"),
                action_type,
                semantic_action,
                json.dumps(action_case.get("main_step") or {}, ensure_ascii=False, default=str),
            )
        ).lower()

        english_words = lambda *words: re.search(r"\b(?:" + "|".join(re.escape(word) for word in words) + r")\b", evidence) is not None

        if "delete_action" in evidence or "deletefile" in evidence or english_words("delete", "remove") or any(token in evidence for token in ("削除", "消去")):
            return "delete"
        if "search_normal" in evidence or english_words("search", "query", "find") or any(token in evidence for token in ("検索", "照会", "抽出")):
            return "read"
        if english_words("update", "modify", "edit", "save") or any(token in evidence for token in ("更新", "変更", "編集", "保存")):
            return "update"
        if "upload_submit" in evidence or english_words("create", "insert", "entry", "register", "add") or any(token in evidence for token in ("登録", "新規", "追加", "作成", "アップロード")):
            return "create"
        return None

    def _compare_database_operation(
        self,
        page_id: str,
        risk: str,
        action: str,
        action_type: Any,
        operation_kind: str,
        legacy_before: Dict[str, Any],
        new_before: Dict[str, Any],
        legacy_action: Dict[str, Any],
        new_action: Dict[str, Any],
        page_dir: Path,
        action_file_id: str,
    ) -> Dict[str, Any]:
        legacy_after = legacy_action.get("state") or {}
        new_after = new_action.get("state") or {}
        legacy_delta = compare_visual_screenshot(
            legacy_before.get("screenshot", ""),
            legacy_after.get("screenshot", ""),
            str(page_dir / f"{action_file_id}_legacy_before_after_diff.png"),
            threshold_percent=self.visual_threshold_percent,
        )
        new_delta = compare_visual_screenshot(
            new_before.get("screenshot", ""),
            new_after.get("screenshot", ""),
            str(page_dir / f"{action_file_id}_new_before_after_diff.png"),
            threshold_percent=self.visual_threshold_percent,
        )

        legacy_changed = legacy_delta.get("status") == "DIFF"
        new_changed = new_delta.get("status") == "DIFF"
        legacy_url_changed = self._normalized_url(legacy_before.get("url", "")) != self._normalized_url(legacy_after.get("url", ""))
        new_url_changed = self._normalized_url(new_before.get("url", "")) != self._normalized_url(new_after.get("url", ""))
        transition_match = legacy_changed == new_changed and legacy_url_changed == new_url_changed

        status = "PASS"
        reason = "Database operation compared within each system before/after."
        if "BLOCKED" in {legacy_action.get("status"), new_action.get("status"), legacy_delta.get("status"), new_delta.get("status")}:
            status = "BLOCKED"
            reason = "Database operation action or before/after capture was blocked."
        elif not transition_match:
            status = "DIFF"
            reason = "Legacy/New database operation transition shape differs."
        elif operation_kind in {"create", "update", "delete"} and not legacy_changed and not new_changed:
            status = "WARN"
            reason = "Database mutation completed but no visible before/after change was detected in either system."

        max_diff = max(
            [
                float(value)
                for value in (legacy_delta.get("diff_percent"), new_delta.get("diff_percent"))
                if isinstance(value, (int, float))
            ]
            or [0.0]
        )
        aggregate_visual = {
            "status": "PASS" if status in {"PASS", "WARN"} else status,
            "diff_percent": max_diff,
            "comparison_mode": "database_operation_before_after",
            "legacy_delta": legacy_delta,
            "new_delta": new_delta,
        }

        return {
            "page_id": page_id,
            "risk": risk,
            "action": action,
            "status": status,
            "reason": reason,
            "url_match": transition_match,
            "dom_match": None,
            "visual": aggregate_visual,
            "comparison_mode": "database_operation_before_after",
            "database_operation": operation_kind,
            "legacy_url": legacy_after.get("url"),
            "new_url": new_after.get("url"),
            "legacy_screenshot": legacy_after.get("screenshot"),
            "new_screenshot": new_after.get("screenshot"),
            "diff_screenshot": None,
            "legacy_before_screenshot": legacy_before.get("screenshot"),
            "legacy_after_screenshot": legacy_after.get("screenshot"),
            "legacy_diff_screenshot": legacy_delta.get("diff_screenshot"),
            "new_before_screenshot": new_before.get("screenshot"),
            "new_after_screenshot": new_after.get("screenshot"),
            "new_diff_screenshot": new_delta.get("diff_screenshot"),
            "legacy_before_url": legacy_before.get("url"),
            "legacy_after_url": legacy_after.get("url"),
            "new_before_url": new_before.get("url"),
            "new_after_url": new_after.get("url"),
            "legacy_delta": legacy_delta,
            "new_delta": new_delta,
            "legacy_delta_changed": legacy_changed,
            "new_delta_changed": new_changed,
            "legacy_url_changed": legacy_url_changed,
            "new_url_changed": new_url_changed,
            "legacy_frame": legacy_after.get("target_frame"),
            "new_frame": new_after.get("target_frame"),
            "frame_candidates": {
                "legacy": legacy_after.get("frame_candidates", []),
                "new": new_after.get("frame_candidates", []),
            },
            "legacy_action": legacy_action,
            "new_action": new_action,
        }

    @staticmethod
    def _requires_target_reopen_after_action(
        action_case: Dict[str, Any],
        action_type: Any,
        semantic_action: Any,
        legacy_action: Dict[str, Any],
        new_action: Dict[str, Any],
    ) -> bool:
        if legacy_action.get("page_closed_after_action") or new_action.get("page_closed_after_action"):
            return True
        if RegressionEngine._is_negative_case(action_case.get("case_type"), action_case.get("action_type") or action_type):
            return True

        evidence = " ".join(
            str(value or "")
            for value in (
                action_case.get("case_type"),
                action_case.get("action_type"),
                action_case.get("label"),
                action_case.get("semantic_key"),
                action_case.get("locator"),
                action_case.get("legacy_locator"),
                action_case.get("new_locator"),
                action_case.get("expected_type"),
                action_type,
                semantic_action,
                json.dumps(action_case.get("main_step") or {}, ensure_ascii=False, default=str),
            )
        ).lower()
        if any(marker in evidence for marker in ("close_window", "window.close", "parent.close")):
            return True
        if any(marker in evidence for marker in ("back_action", "キャンセル", "取消", "戻る", "戻り", "戻 ")):
            return True
        if re.search(r"\b(?:cancel|back|bak)\b", evidence):
            return True
        return False

    def _reopen_target_pair(
        self,
        legacy_page: Page,
        new_page: Page,
        mapping: Dict[str, Any],
        page_dir: Path,
        browser_name: str,
        *,
        reason: str,
    ) -> Tuple[Page, Page, Dict[str, Any]]:
        page_id = mapping.get("page_id") or "unknown"
        print(
            f"[{page_id}] REOPEN target page after leaving/closing action: "
            + json.dumps({"reason": reason}, ensure_ascii=False)
        )

        legacy_page, legacy_nav = self._reopen_target_side(
            legacy_page,
            mapping,
            page_dir,
            browser_name,
            side="legacy",
        )
        new_page, new_nav = self._reopen_target_side(
            new_page,
            mapping,
            page_dir,
            browser_name,
            side="new",
        )
        status = "PASS" if legacy_nav.get("status") == "PASS" and new_nav.get("status") == "PASS" else "BLOCKED"
        result = {
            "status": status,
            "reason": reason if status == "PASS" else "Failed to reopen target page after leaving/closing action.",
            "legacy": legacy_nav,
            "new": new_nav,
        }
        print(
            f"[{page_id}] REOPEN result: "
            + json.dumps(
                {
                    "status": status,
                    "legacy_status": legacy_nav.get("status"),
                    "legacy_reason": legacy_nav.get("reason"),
                    "new_status": new_nav.get("status"),
                    "new_reason": new_nav.get("reason"),
                },
                ensure_ascii=False,
                default=str,
            )
        )
        return legacy_page, new_page, result

    def _reopen_target_side(
        self,
        page: Page,
        mapping: Dict[str, Any],
        page_dir: Path,
        browser_name: str,
        *,
        side: str,
    ) -> Tuple[Page, Dict[str, Any]]:
        """Re-enter the current target page after an action leaves/closes it.

        Minimal lifecycle recovery policy:
        - If the current page is still one of the route-map step pages, reuse
          the current session and run the route map back to the target page.
        - Otherwise, reopen the login/base entry URL, perform the normal login
          from env/Config credentials if a login form is present, then run the
          route map back to the target page.
        - If no route map is available, keep the previous direct-url fallback.
        """
        page_id = mapping.get("page_id") or "unknown"
        entry_url = self.legacy_base_url if side == "legacy" else self.new_base_url
        route = self.route_map_catalog.find_for_target(page_id, side=side) or self.route_map_catalog.find_for_target(page_id)

        if route:
            route_step = self._page_is_route_map_step(page, route)
            pre_nav: Dict[str, Any]

            if route_step:
                pre_nav = {
                    "status": "PASS",
                    "strategy": "reuse_current_route_step",
                    "current_url": self._safe_page_url(page),
                }
            else:
                page, pre_nav = self._login_recovery_page(
                    page,
                    entry_url,
                    page_dir,
                    f"{side}_reopen_login_entry",
                )
                if pre_nav.get("status") != "PASS":
                    pre_nav["reopen_strategy"] = "login_recovery_before_route_map"
                    return page, pre_nav

            navigator = RouteNavigator(
                self.route_map_catalog,
                timeout=self.timeout,
                browser_name=browser_name,
                upload_file=self.upload_file,
            )
            page, nav = navigator.navigate(
                page,
                entry_url=entry_url,
                target_page=page_id,
                capture_dir=page_dir,
                side=side,
            )
            nav["pre_recovery"] = pre_nav
            nav["route_step_detected"] = route_step
            nav["reopen_strategy"] = "route_map_from_current_step" if route_step else "login_recovery_then_route_map"
            return page, nav

        target_url = self._page_url(
            self.legacy_base_url if side == "legacy" else self.new_base_url,
            mapping.get("entry_url") or mapping.get("resolved_entry_url") or page_id,
        )
        page, nav = self._open_or_reset_page(page, target_url, page_dir, f"{side}_reopen_direct")
        nav["reopen_strategy"] = "direct_url"
        if nav.get("status") == "PASS" and not self._page_matches_mapping(page, mapping):
            nav.update(
                {
                    "status": "BLOCKED",
                    "target_reached": False,
                    "reason": f"Reopen direct URL did not reach target page: {page_id}",
                }
            )
        elif nav.get("status") == "PASS":
            nav["target_reached"] = True
        return page, nav

    def _login_recovery_page(
        self,
        page: Page,
        entry_url: str,
        page_dir: Path,
        capture_name: str,
    ) -> Tuple[Page, Dict[str, Any]]:
        page, nav = self._open_or_reset_page(page, entry_url, page_dir, capture_name)
        if nav.get("status") != "PASS":
            return page, nav

        login_result = self._try_login_if_login_form(page)
        nav["login_recovery"] = login_result
        if login_result.get("status") == "BLOCKED":
            nav.update(
                {
                    "status": "BLOCKED",
                    "reason": login_result.get("reason") or "Login recovery failed.",
                }
            )
        return page, nav

    def _try_login_if_login_form(self, page: Page) -> Dict[str, Any]:
        password_selectors = [
            "input[type='password']",
            "input[name='password']",
            "input[name='passwd']",
            "input[name='pass']",
        ]
        password_selector = self._first_visible_selector(page, password_selectors, timeout=1500)
        if not password_selector:
            return {"status": "PASS", "reason": "login_form_not_detected"}

        username = self._env_or_config(
            "LOGIN_USERNAME",
            "LOGIN_USER",
            "LOGIN_USER_ID",
            "USERNAME",
            "USER_ID",
        )
        password = self._env_or_config(
            "LOGIN_PASSWORD",
            "LOGIN_PASS",
            "PASSWORD",
        )
        if not username or not password:
            return {
                "status": "BLOCKED",
                "reason": "Login form detected but username/password were not found in environment or Config.",
            }

        user_selector = self._first_visible_selector(
            page,
            [
                "input[name='user']",
                "input[name='username']",
                "input[name='userId']",
                "input[name='userid']",
                "input[name='loginId']",
                "input[name='login_id']",
                "input[type='text']",
            ],
            timeout=1500,
        )
        if not user_selector:
            return {"status": "BLOCKED", "reason": "Login username field was not found."}

        try:
            page.locator(user_selector).first().fill(username, timeout=3000)
            page.locator(password_selector).first().fill(password, timeout=3000)

            submit_selector = self._first_visible_selector(
                page,
                [
                    "input[type='button']",
                    "input[type='submit']",
                    "button[type='submit']",
                    "button",
                ],
                timeout=1000,
            )
            if submit_selector:
                page.locator(submit_selector).first().click(timeout=5000)
            else:
                page.locator(password_selector).first().press("Enter", timeout=3000)

            try:
                page.wait_for_load_state("domcontentloaded", timeout=max(self.timeout, 10000))
            except Exception:
                pass
            return {"status": "PASS", "strategy": "login_form_submit", "url": self._safe_page_url(page)}
        except Exception as exc:
            return {"status": "BLOCKED", "reason": f"Login recovery failed: {exc}"}

    @staticmethod
    def _env_or_config(*names: str) -> str:
        for name in names:
            value = os.environ.get(name)
            if value:
                return value
            value = getattr(Config, name, None)
            if value:
                return str(value)
        return ""

    @staticmethod
    def _first_visible_selector(page: Page, selectors: List[str], *, timeout: int = 1000) -> Optional[str]:
        for selector in selectors:
            try:
                locator = page.locator(selector).first()
                if locator.count() > 0 and locator.is_visible(timeout=timeout):
                    return selector
            except Exception:
                continue
        return None

    def _page_is_route_map_step(self, page: Page, route: Dict[str, Any]) -> bool:
        if self._page_is_closed(page):
            return False

        tokens = self._route_map_step_tokens(route)
        if not tokens:
            return False

        try:
            urls = [page.url]
            urls.extend(frame.url for frame in page.frames)
        except Exception:
            return False

        haystack = "\n".join(str(url or "").replace("\\", "/").lower() for url in urls)
        return any(token in haystack for token in tokens)

    @staticmethod
    def _route_map_step_tokens(route: Dict[str, Any]) -> List[str]:
        tokens: List[str] = []

        def add_token(value: Any) -> None:
            raw = str(value or "").replace("\\", "/").strip().lower()
            if not raw:
                return
            leaf = raw.rsplit("/", 1)[-1]
            candidates = [raw, leaf]
            stem = re.sub(r"\.(jsp|do|action)$", "", leaf, flags=re.IGNORECASE)
            if stem and stem != leaf:
                candidates.extend([stem, f"{stem}.do", f"{stem}.jsp"])
            for candidate in candidates:
                candidate = candidate.strip("/")
                if len(candidate) >= 3 and candidate not in tokens:
                    tokens.append(candidate)

        def walk(value: Any, parent_key: str = "") -> None:
            if isinstance(value, dict):
                for key, child in value.items():
                    key_text = str(key or "").lower()
                    if isinstance(child, str) and any(
                        marker in key_text
                        for marker in ("url", "page", "path", "action", "target", "entry", "href")
                    ):
                        add_token(child)
                    walk(child, key_text)
            elif isinstance(value, list):
                for child in value:
                    walk(child, parent_key)

        walk(route)
        return tokens

    @staticmethod
    def _safe_page_url(page: Page) -> str:
        try:
            return page.url
        except Exception:
            return ""


    def _compare_state(
        self,
        page_id: str,
        risk: str,
        action: str,
        legacy_state: Dict[str, Any],
        new_state: Dict[str, Any],
        diff_path: Path,
        **extra: Any,
    ) -> Dict[str, Any]:
        action_type = str(extra.get("action_type") or action or "page_snapshot")
        visual = {"status": "SKIPPED", "reason": "Visual comparison is disabled for this action type"}
        if action_type.lower() not in {"download", "close_window"}:
            visual = compare_visual_screenshot(
                legacy_state.get("screenshot", ""),
                new_state.get("screenshot", ""),
                str(diff_path),
                threshold_percent=self.visual_threshold_percent,
            )
        url_match = self._normalized_url(legacy_state.get("url", "")) == self._normalized_url(new_state.get("url", ""))
        dom_match = str(legacy_state.get("dom", "")) == str(new_state.get("dom", ""))
        status = _judge_compare_status(
            action_type=action_type,
            visual_status=visual.get("status"),
            url_match=url_match,
            legacy_action_status=(extra.get("legacy_action") or {}).get("status"),
            new_action_status=(extra.get("new_action") or {}).get("status"),
            legacy_nav_status=(extra.get("legacy_nav") or {}).get("status"),
            new_nav_status=(extra.get("new_nav") or {}).get("status"),
        )
        if status == "DIFF" and self._looks_like_data_variance(action_type, visual, legacy_state, new_state):
            status = "WARN"
            visual["data_variance_tolerated"] = True
            visual["reason"] = "Visual diff appears to be table/list data variance between environments."

        return {
            "page_id": page_id,
            "risk": risk,
            "action": action,
            "status": status,
            "url_match": url_match,
            "dom_match": dom_match,
            "visual": visual,
            "legacy_url": legacy_state.get("url"),
            "new_url": new_state.get("url"),
            "legacy_screenshot": legacy_state.get("screenshot"),
            "new_screenshot": new_state.get("screenshot"),
            "diff_screenshot": visual.get("diff_screenshot"),
            "legacy_frame": legacy_state.get("target_frame"),
            "new_frame": new_state.get("target_frame"),
            "frame_candidates": {
                "legacy": legacy_state.get("frame_candidates", []),
                "new": new_state.get("frame_candidates", []),
            },
            **extra,
        }

    @staticmethod
    def _looks_like_data_variance(
        action_type: str,
        visual: Dict[str, Any],
        legacy_state: Dict[str, Any],
        new_state: Dict[str, Any],
    ) -> bool:
        if str(visual.get("status") or "").upper() != "DIFF":
            return False
        try:
            diff_percent = float(visual.get("diff_percent") or 0)
        except (TypeError, ValueError):
            diff_percent = 0
        if diff_percent <= 0 or diff_percent > 10:
            return False

        normalized_action = str(action_type or "").lower()
        if normalized_action not in {"page_snapshot", "snapshot", "initial_display", "result_table_verify"}:
            return False

        legacy_lines = RegressionEngine._stable_text_lines(legacy_state.get("text", ""))
        new_lines = RegressionEngine._stable_text_lines(new_state.get("text", ""))
        if not legacy_lines or not new_lines:
            return False
        shared = set(legacy_lines) & set(new_lines)
        combined = " ".join(shared).lower()
        table_markers = ("一覧", "検索結果", "結果", "削除", "アップロード日時", "ファイル", "table", "list")
        return len(shared) >= 2 and any(marker.lower() in combined for marker in table_markers)

    @staticmethod
    def _stable_text_lines(value: Any) -> List[str]:
        lines: List[str] = []
        for raw_line in str(value or "").splitlines():
            line = " ".join(raw_line.split())
            if not line:
                continue
            # Drop row values that are commonly environment data, while keeping
            # titles, headers, and button labels useful for structure checks.
            if re.search(r"\d{4}年\d{1,2}月\d{1,2}日|\d{4}[-/]\d{1,2}[-/]\d{1,2}", line):
                continue
            if re.search(r"\.(csv|tsv|xls|xlsx|pdf|txt)\b", line, re.IGNORECASE):
                continue
            if len(line) > 120:
                continue
            lines.append(line)
        return lines

    def render_report(self, results: List[Dict[str, Any]]) -> Path:
        report_path, report_dir, report_page = self._report_location(results)
        browser_name = self.current_browser_name or "-"
        counts = Counter(item["status"] for item in results)
        rows = "\n".join(self._render_result(item, report_dir) for item in results)
        report_dir.mkdir(parents=True, exist_ok=True)
        report_path.write_text(
            f"""<!doctype html>
<html lang="ja">
<head>
  <meta charset="utf-8">
  <title>Moonlight Regression Report - {html.escape(report_page)}</title>
  <style>
    * {{ box-sizing: border-box; }}
    body {{ margin: 0; font-family: Arial, sans-serif; color: #17202a; background: #f5f7fb; }}
    header {{ padding: 24px 32px; background: #17202a; color: white; }}
    h1 {{ margin: 0 0 12px; font-size: 26px; letter-spacing: 0; }}
    .subtitle {{ margin: 0 0 16px; color: #d6eaf8; font-size: 13px; overflow-wrap: anywhere; }}
    .summary {{ display: flex; gap: 12px; flex-wrap: wrap; }}
    .summary-details {{ margin-top: 16px; display: grid; grid-template-columns: repeat(4, minmax(0, 1fr)); gap: 12px; }}
    .summary-card {{ background: #273746; border-radius: 8px; padding: 12px; }}
    .summary-card h2 {{ margin: 0 0 8px; font-size: 13px; color: #d6eaf8; }}
    .summary-card p {{ margin: 4px 0; font-size: 12px; color: #f8f9f9; overflow-wrap: anywhere; }}
    .summary-card .num {{ font-size: 20px; font-weight: 800; }}
    .pill {{ padding: 8px 12px; border-radius: 6px; background: #273746; font-weight: 700; }}
    main {{ padding: 24px 32px; }}
    .case {{ margin-bottom: 20px; border: 1px solid #d9e0ea; border-radius: 8px; background: white; overflow: hidden; }}
    .case-head {{ display: flex; align-items: center; gap: 12px; padding: 12px 16px; border-bottom: 1px solid #e7ecf3; flex-wrap: wrap; }}
    .case-title {{ display: flex; align-items: center; gap: 10px; min-width: min(100%, 420px); }}
    .case-title strong {{ overflow-wrap: anywhere; }}
    .status {{ padding: 4px 8px; border-radius: 4px; color: white; font-weight: 700; font-size: 12px; }}
    .PASS {{ background: #1e8449; }} .WARN {{ background: #b7950b; }} .DIFF {{ background: #b7950b; }} .BLOCKED {{ background: #922b21; }} .ERROR {{ background: #7b241c; }}
    .meta {{ color: #52616f; font-size: 13px; overflow-wrap: anywhere; }}
    .grid {{ display: grid; grid-template-columns: repeat(3, minmax(0, 1fr)); gap: 12px; padding: 16px; }}
    .db-grid {{ grid-template-columns: repeat(3, minmax(0, 1fr)); }}
    figure {{ margin: 0; }}
    figcaption {{ margin-bottom: 6px; font-size: 12px; color: #52616f; font-weight: 700; }}
    img {{ width: 100%; max-height: 520px; object-fit: contain; border: 1px solid #d9e0ea; background: #fff; }}
    .details {{ padding: 0 16px 16px; font-size: 13px; color: #34495e; }}
    .detail-grid {{ display: grid; grid-template-columns: 160px minmax(0, 1fr); gap: 6px 12px; margin-bottom: 12px; }}
    .detail-grid b {{ color: #17202a; }}
    table {{ width: 100%; border-collapse: collapse; }}
    td {{ border-top: 1px solid #e7ecf3; padding: 8px 10px; }}
    details {{ margin-top: 10px; }}
    summary {{ cursor: pointer; font-weight: 700; color: #17202a; }}
    code {{ font-family: Consolas, monospace; font-size: 12px; white-space: pre-wrap; }}
    @media (max-width: 900px) {{ .grid, .summary-details {{ grid-template-columns: 1fr; }} }}
  </style>
</head>
<body>
  <header>
    <h1>Moonlight Legacy/New Regression Report - {html.escape(report_page)}</h1>
    <p class="subtitle">Page: {html.escape(report_page)} / Browser: {html.escape(browser_name)} / Report: {html.escape(str(report_path))}</p>
    <div class="summary">
      <span class="pill">Total: {len(results)}</span>
      <span class="pill">PASS: {counts.get('PASS', 0)}</span>
      <span class="pill">DIFF: {counts.get('DIFF', 0)}</span>
      <span class="pill">BLOCKED: {counts.get('BLOCKED', 0)}</span>
      <span class="pill">ERROR: {counts.get('ERROR', 0)}</span>
    </div>
    {self._render_summary_details(results, counts)}
  </header>
  <main>{self._render_coverage_matrix(results)}{rows or '<p>No results.</p>'}</main>
</body>
</html>
""",
            encoding="utf-8",
        )
        return report_path

    def _report_location(self, results: List[Dict[str, Any]]) -> Tuple[Path, Path, str]:
        pages = sorted({str(item.get("page_id") or "-") for item in results})
        report_page = pages[0] if len(pages) == 1 else f"{len(pages)} pages"
        report_dir = self.output_dir

        if len(pages) == 1:
            asset_dir = self._first_asset_dir(results)
            if asset_dir is not None:
                report_dir = asset_dir
            else:
                report_dir = self.output_dir / f"0001_{self._safe_name(report_page)}"

        return report_dir / "regression_report.html", report_dir, report_page

    @staticmethod
    def _first_asset_dir(results: List[Dict[str, Any]]) -> Optional[Path]:
        for item in results:
            for key in (
                "legacy_screenshot",
                "new_screenshot",
                "diff_screenshot",
                "legacy_before_screenshot",
                "legacy_after_screenshot",
                "legacy_diff_screenshot",
                "new_before_screenshot",
                "new_after_screenshot",
                "new_diff_screenshot",
            ):
                value = item.get(key)
                if value:
                    return Path(value).resolve().parent
        return None

    def _render_result(self, item: Dict[str, Any], report_dir: Path) -> str:
        if item.get("comparison_mode") == "database_operation_before_after":
            return self._render_database_operation_result(item, report_dir)

        visual = item.get("visual") or {}
        diff_percent = visual.get("diff_percent")
        diff_text = "-" if diff_percent is None else f"{diff_percent:.4f}%"
        action_type = item.get("action_type") or self._action_type_from_payload(item)
        reason = item.get("reason") or self._blocked_reason(item) or "-"
        return f"""
<section class="case">
  <div class="case-head">
    <div class="case-title">
      <span class="status {html.escape(item.get('status', 'DIFF'))}">{html.escape(item.get('status', 'DIFF'))}</span>
      <strong>{html.escape(str(item.get('page_id')))}</strong>
    </div>
    <span class="meta">risk={html.escape(str(item.get('risk')))} / action={html.escape(str(item.get('action')))} / diff={diff_text}</span>
  </div>
  <div class="grid">
    {self._figure('Legacy', item.get('legacy_screenshot'), report_dir)}
    {self._figure('New', item.get('new_screenshot'), report_dir)}
    {self._figure('Diff', item.get('diff_screenshot'), report_dir)}
  </div>
  <div class="details">
    <div class="detail-grid">
      <b>URL match</b><span>{item.get('url_match')}</span>
      <b>Action type</b><span>{html.escape(str(action_type or '-'))}</span>
      <b>Legacy</b><span>{html.escape(str(item.get('legacy_url') or '-'))}</span>
      <b>New</b><span>{html.escape(str(item.get('new_url') or '-'))}</span>
      <b>Upload file</b><span>{html.escape(str(item.get('upload_file') or '-'))}</span>
      <b>Submit locator</b><span>{html.escape(str(item.get('submit_locator') or '-'))}</span>
      <b>After URL</b><span>{html.escape(str(item.get('legacy_after_url') or '-'))} / {html.escape(str(item.get('new_after_url') or '-'))}</span>
      <b>Runtime</b><span>navigation={item.get('navigation_detected')} / popup={item.get('popup_detected')} / frame={item.get('frame_changed')} / validation_only={item.get('validation_only')}</span>
      <b>Reason</b><span>{html.escape(str(reason))}</span>
    </div>
    {self._render_diagnostics(item)}
  </div>
</section>"""

    def _render_database_operation_result(self, item: Dict[str, Any], report_dir: Path) -> str:
        legacy_delta = item.get("legacy_delta") or {}
        new_delta = item.get("new_delta") or {}
        legacy_diff = legacy_delta.get("diff_percent")
        new_diff = new_delta.get("diff_percent")
        legacy_diff_text = "-" if legacy_diff is None else f"{legacy_diff:.4f}%"
        new_diff_text = "-" if new_diff is None else f"{new_diff:.4f}%"
        action_type = item.get("action_type") or self._action_type_from_payload(item)
        reason = item.get("reason") or self._blocked_reason(item) or "-"
        operation = item.get("database_operation") or "-"
        return f"""
<section class="case">
  <div class="case-head">
    <div class="case-title">
      <span class="status {html.escape(item.get('status', 'DIFF'))}">{html.escape(item.get('status', 'DIFF'))}</span>
      <strong>{html.escape(str(item.get('page_id')))}</strong>
    </div>
    <span class="meta">risk={html.escape(str(item.get('risk')))} / action={html.escape(str(item.get('action')))} / DB={html.escape(str(operation))} / legacy Δ={legacy_diff_text} / new Δ={new_diff_text}</span>
  </div>
  <div class="grid db-grid">
    {self._figure('Legacy Before', item.get('legacy_before_screenshot'), report_dir)}
    {self._figure('Legacy After', item.get('legacy_after_screenshot'), report_dir)}
    {self._figure('Legacy Before/After Diff', item.get('legacy_diff_screenshot'), report_dir)}
    {self._figure('New Before', item.get('new_before_screenshot'), report_dir)}
    {self._figure('New After', item.get('new_after_screenshot'), report_dir)}
    {self._figure('New Before/After Diff', item.get('new_diff_screenshot'), report_dir)}
  </div>
  <div class="details">
    <div class="detail-grid">
      <b>Compare mode</b><span>database_operation_before_after</span>
      <b>DB operation</b><span>{html.escape(str(operation))}</span>
      <b>Transition match</b><span>{item.get('url_match')}</span>
      <b>Action type</b><span>{html.escape(str(action_type or '-'))}</span>
      <b>Legacy delta</b><span>{html.escape(str(legacy_delta.get('status') or '-'))} / {legacy_diff_text}</span>
      <b>New delta</b><span>{html.escape(str(new_delta.get('status') or '-'))} / {new_diff_text}</span>
      <b>Legacy before</b><span>{html.escape(str(item.get('legacy_before_url') or '-'))}</span>
      <b>Legacy after</b><span>{html.escape(str(item.get('legacy_after_url') or '-'))}</span>
      <b>New before</b><span>{html.escape(str(item.get('new_before_url') or '-'))}</span>
      <b>New after</b><span>{html.escape(str(item.get('new_after_url') or '-'))}</span>
      <b>Reason</b><span>{html.escape(str(reason))}</span>
    </div>
    {self._render_diagnostics(item)}
  </div>
</section>"""

    def _figure(self, label: str, path: Optional[str], report_dir: Path) -> str:
        if not path:
            return f"<figure><figcaption>{html.escape(label)}</figcaption><div class=\"meta\">No screenshot</div></figure>"
        try:
            rel = Path(path).resolve().relative_to(report_dir.resolve())
        except ValueError:
            rel = Path(path).resolve()
        return f'<figure><figcaption>{html.escape(label)}</figcaption><img src="{html.escape(rel.as_posix())}" alt="{html.escape(label)}"></figure>'

    @staticmethod
    def _blocked_reason(item: Dict[str, Any]) -> Optional[str]:
        for key in ("legacy_action", "new_action", "legacy_nav", "new_nav"):
            value = item.get(key) or {}
            if value.get("reason"):
                return value.get("reason")
        return None

    @staticmethod
    def _action_type_from_payload(item: Dict[str, Any]) -> Optional[str]:
        for key in ("legacy_action", "new_action"):
            value = item.get(key) or {}
            if value.get("semantic_action"):
                return value.get("semantic_action")
        return None

    def _render_summary_details(self, results: List[Dict[str, Any]], counts: Counter) -> str:
        """Render compact diagnostic information in the top summary area."""
        total = len(results)
        pages = sorted({str(item.get("page_id") or "-") for item in results})
        action_counts = Counter(str(item.get("action_type") or self._action_type_from_payload(item) or "-") for item in results)
        risk_counts = Counter(str(item.get("risk") or "-") for item in results)
        visual_counts = Counter(str((item.get("visual") or {}).get("status") or "-") for item in results)
        url_true = sum(1 for item in results if item.get("url_match") is True)
        url_false = sum(1 for item in results if item.get("url_match") is False)
        diffs = [
            float((item.get("visual") or {}).get("diff_percent"))
            for item in results
            if isinstance((item.get("visual") or {}).get("diff_percent"), (int, float))
        ]
        avg_diff = sum(diffs) / len(diffs) if diffs else 0.0
        max_diff = max(diffs) if diffs else 0.0
        hotspot_counts = Counter(
            str(item.get("page_id") or "-")
            for item in results
            if item.get("status") not in {"PASS"}
        )
        hotspots = ", ".join(f"{page}:{count}" for page, count in hotspot_counts.most_common(5)) or "-"

        def fmt_counter(counter: Counter) -> str:
            return ", ".join(f"{html.escape(str(k))}:{v}" for k, v in counter.most_common()) or "-"

        return f"""
    <div class="summary-details">
      <div class="summary-card">
        <h2>Execution</h2>
        <p><span class="num">{total}</span> results / {len(pages)} page(s)</p>
        <p>Risk: {fmt_counter(risk_counts)}</p>
      </div>
      <div class="summary-card">
        <h2>Status</h2>
        <p>PASS: {counts.get('PASS', 0)} / WARN: {counts.get('WARN', 0)}</p>
        <p>DIFF: {counts.get('DIFF', 0)} / BLOCKED: {counts.get('BLOCKED', 0)} / ERROR: {counts.get('ERROR', 0)}</p>
      </div>
      <div class="summary-card">
        <h2>Comparison</h2>
        <p>URL match: {url_true} true / {url_false} false</p>
        <p>Visual: {fmt_counter(visual_counts)}</p>
        <p>Visual diff avg/max: {avg_diff:.4f}% / {max_diff:.4f}%</p>
      </div>
      <div class="summary-card">
        <h2>Actions / Hotspots</h2>
        <p>Actions: {fmt_counter(action_counts)}</p>
        <p>Problem pages: {html.escape(hotspots)}</p>
      </div>
    </div>"""

    def _render_coverage_matrix(self, results: List[Dict[str, Any]]) -> str:
        pages = sorted({self._target_page_name(item.get("page_id")) for item in results if item.get("page_id")})
        checklist_rows: List[Dict[str, Any]] = []
        for page in pages:
            checklist_rows.extend(self._checklist_case_rows.get(page) or [])

        if checklist_rows:
            body = "\n".join(
                "<tr>"
                f"<td>{html.escape(str(row.get('case_id') or '-'))}</td>"
                f"<td>{html.escape(str(row.get('test_title') or '-'))}</td>"
                f"<td>{html.escape(str(row.get('automation_mode') or '-'))}</td>"
                f"<td>{html.escape(str(row.get('destructive') or 'false'))}</td>"
                f"<td>{html.escape(str(row.get('excluded_reason') or '-'))}</td>"
                "</tr>"
                for row in checklist_rows
            )
        else:
            body = '<tr><td colspan="5">No checklist cases were loaded for this report.</td></tr>'

        return f"""
<section class="case">
  <div class="case-head"><strong>Checklist Cases</strong></div>
  <div class="details">
    <table>
      <thead>
        <tr>
          <td><b>case_id</b></td>
          <td><b>test_title</b></td>
          <td><b>automation_mode</b></td>
          <td><b>destructive</b></td>
          <td><b>excluded_reason</b></td>
        </tr>
      </thead>
      <tbody>{body}</tbody>
    </table>
  </div>
</section>"""

    def _render_diagnostics(self, item: Dict[str, Any]) -> str:
        legacy_action = item.get("legacy_action") or {}
        new_action = item.get("new_action") or {}
        legacy_frame = item.get("legacy_frame") or legacy_action.get("target_frame") or {}
        new_frame = item.get("new_frame") or new_action.get("target_frame") or {}
        diagnostics = {
            "legacy_locator": item.get("legacy_locator"),
            "new_locator": item.get("new_locator"),
            "legacy_frame": legacy_frame,
            "new_frame": new_frame,
            "legacy_wait": legacy_action.get("wait_state"),
            "new_wait": new_action.get("wait_state"),
            "legacy_selector_found": legacy_action.get("selector_found"),
            "new_selector_found": new_action.get("selector_found"),
            "legacy_blocked_reason": legacy_action.get("reason"),
            "new_blocked_reason": new_action.get("reason"),
            "upload_file": item.get("upload_file"),
            "submit_locator": item.get("submit_locator"),
            "legacy_after_url": item.get("legacy_after_url"),
            "new_after_url": item.get("new_after_url"),
            "navigation_detected": item.get("navigation_detected"),
            "popup_detected": item.get("popup_detected"),
            "frame_changed": item.get("frame_changed"),
            "validation_only": item.get("validation_only"),
            "post_action_reopen": item.get("post_action_reopen"),
            "frame_candidates": item.get("frame_candidates"),
        }
        return (
            "<details class=\"diag\">"
            "<summary>Diagnostics</summary>"
            f"<code>{html.escape(json.dumps(diagnostics, ensure_ascii=False, default=str, indent=2))}</code>"
            "</details>"
        )

    def _dedupe_actions(self, actions: List[Dict[str, Any]]) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
        """
        Backward-compatible dedupe used by older tests and callers.

        Keep the first action for the same semantic key. page_mapping places
        locator_changes before full_action_steps, so this preserves explicit
        migration mappings over later fallback scanner actions.
        """
        seen = set()
        unique: List[Dict[str, Any]] = []
        skipped: List[Dict[str, Any]] = []
        for action in actions:
            key = (
                action.get("semantic_key")
                or action.get("label")
                or action.get("legacy_locator")
                or action.get("locator")
                or json.dumps(action, sort_keys=True, ensure_ascii=False, default=str)
            )
            if key in seen:
                skipped.append(action)
                continue
            seen.add(key)
            unique.append(action)
        return unique, skipped

    @staticmethod
    def _page_url(base_url: str, page_id: str) -> str:
        return urljoin(base_url, str(page_id or "").lstrip("/"))

    @staticmethod
    def _route_map_paths(route_map_path: Optional[str]) -> Optional[List[Path]]:
        if not route_map_path:
            return None
        paths: List[Path] = []
        for raw in str(route_map_path).split(","):
            item = raw.strip()
            if not item:
                continue
            path = Path(item)
            if any(char in item for char in "*?[]"):
                paths.extend(Path(match) for match in sorted(glob.glob(item)))
            elif path.is_dir():
                paths.extend(sorted(path.rglob("usable_route_map*.json")))
            else:
                paths.append(path)
        return paths

    @staticmethod
    def _normalized_url(url: str) -> str:
        if not url:
            return ""
        parts = urlsplit(url)
        ignored = {"userid", "sessionid", "jsessionid", "token", "csrf", "_", "timestamp"}
        query_pairs = []
        for key, value in parse_qsl(parts.query, keep_blank_values=True):
            if key.lower() in ignored:
                continue
            query_pairs.append((key, value))
        query = urlencode(query_pairs)
        return f"{parts.path.rstrip('/') or '/'}?{query}".rstrip("?")

    @staticmethod
    def _safe_name(value: Any) -> str:
        return re.sub(r"[^A-Za-z0-9_.-]+", "_", str(value or "unknown")).strip("_")[:120] or "unknown"

    @staticmethod
    def _target_page_name(value: Any) -> str:
        name = PureWindowsPath(str(value or "").strip().replace("/", "\\")).name
        if name.lower().endswith(".do"):
            name = name[:-3] + ".jsp"
        return name.lower()

    def _page_matches_mapping(self, page: Page, mapping: Dict[str, Any]) -> bool:
        """
        Best-effort check that direct navigation landed on the requested page.

        Legacy Struts pages can render the useful state inside frames or
        popups, so this checks all frame URLs against both page_id and resolved
        action URL. Body presence alone is not enough because login/menu pages
        also satisfy that.
        """
        target_values = [
            mapping.get("page_id"),
            mapping.get("entry_url"),
            mapping.get("resolved_entry_url"),
        ]
        needles = []
        for value in target_values:
            raw = str(value or "").replace("\\", "/").strip().lower()
            if not raw:
                continue
            leaf = raw.rsplit("/", 1)[-1]
            stem = re.sub(r"\.(jsp|do|action)$", "", leaf, flags=re.IGNORECASE)
            for candidate in (raw, leaf, stem, f"{stem}.do"):
                candidate = candidate.strip("/")
                if len(candidate) >= 3 and candidate not in needles:
                    needles.append(candidate)

        if not needles:
            return True

        urls = []
        try:
            urls.append(page.url)
            urls.extend(frame.url for frame in page.frames)
        except PlaywrightError:
            return False

        haystack = "\n".join(str(url or "").replace("\\", "/").lower() for url in urls)
        return any(needle in haystack for needle in needles)

    @staticmethod
    def _page_is_closed(page: Page) -> bool:
        try:
            return page.is_closed()
        except Exception:
            return True

    def _open_or_reset_page(self, page: Page, url: str, capture_dir: Path, name: str) -> Tuple[Page, Dict[str, Any]]:
        try:
            if self._page_is_closed(page):
                page = page.context.new_page()
        except Exception as exc:
            return page, {"status": "BLOCKED", "url": url, "reason": f"Failed to create replacement page: {exc}"}

        nav = self._goto(page, url)
        if nav.get("status") != "PASS":
            try:
                nav["state"] = _capture_state(page, capture_dir, name)
            except Exception as exc:
                nav["capture_error"] = str(exc)
        return page, nav

    def _goto(self, page: Page, url: str) -> Dict[str, Any]:
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=max(self.timeout, 60000))

            try:
                page.wait_for_selector("body, form, table, input", timeout=10000)
            except Exception:
                pass

            return {"status": "PASS", "url": page.url}
        except (PlaywrightTimeoutError, PlaywrightError) as exc:
            return {"status": "BLOCKED", "url": url, "reason": str(exc)}
