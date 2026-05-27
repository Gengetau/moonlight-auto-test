import json
import re
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional


DEFAULT_CHECKLIST_PATH = "generated/valid/migration_checklist.xlsx"
DEFAULT_ROUTE_MAP_PATH = "generated/valid/route"
DEFAULT_NEGATIVE_PROFILES = [
    {
        "profile": "negative_js_error",
        "label": "JS error",
        "description": "console.error + JavaScript runtime error",
    },
    {
        "profile": "negative_http_500",
        "label": "HTTP 500",
        "description": "Mock target request as HTTP 500",
    },
    {
        "profile": "negative_network_abort",
        "label": "Network abort",
        "description": "Abort target request",
    },
    {
        "profile": "negative_file_upload",
        "label": "Invalid file upload",
        "description": "Invalid/negative file upload cases",
    },
]


def quote(value: Any) -> str:
    text = str(value)
    return '"' + text.replace('"', '\\"') + '"'


def browser_key(browser_name: str) -> str:
    return re.sub(r"[^a-z0-9_.-]+", "_", str(browser_name or "browser").lower()).strip("_") or "browser"


def safe_page_key(page_id: str) -> str:
    text = Path(str(page_id or "selected_pages").replace("\\", "/")).name
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", text).strip("_") or "selected_pages"


def split_case_types(value: Any) -> List[str]:
    return [item.strip() for item in re.split(r"[\r\n,;]+", str(value or "")) if item.strip()]


def target_page_name(value: Any) -> str:
    text = Path(str(value or "").replace("\\", "/")).name
    return text.lower()


def html_report_path(browser_name: str, page_id: str) -> Path:
    return Path("output/gui") / browser_key(browser_name) / safe_page_key(page_id) / "gui_report.html"


def regression_output_dir(browser_name: str) -> Path:
    return Path("output/regression") / browser_key(browser_name)


def upload_profile_config_path(browser_name: str, page_id: str) -> Path:
    return Path("generated/gui/upload_profiles") / browser_key(browser_name) / f"{safe_page_key(page_id)}.json"


def build_regression_command(config: Dict[str, Any], *, pytest_cmd: str) -> str:
    browser = str(config.get("browser") or "chrome_port")
    page_id = str(config.get("target_page") or "").strip()
    if not page_id:
        raise ValueError("target_page is required")

    cmd = f"{pytest_cmd} tests/test_migration.py --run-migration --test-browser={browser}"
    cmd += f" --target-page={quote(page_id)}"
    cmd += f" --regression-output-dir={quote(config.get('regression_output_dir') or regression_output_dir(browser))}"

    login_entry = str(config.get("login_entry") or "").strip()
    if login_entry:
        cmd += f" --login-entry={quote(login_entry)}"

    checklist_path = str(config.get("checklist_path") or "").strip()
    if checklist_path:
        cmd += f" --checklist-path={quote(checklist_path)}"

    route_map_path = str(config.get("route_map_path") or "").strip()
    if config.get("force_route_map"):
        cmd += " --force-route-map"
        if route_map_path:
            cmd += f" --route-map-path={quote(route_map_path)}"

    if config.get("manual"):
        cmd += " --manual"
    if config.get("risk_only"):
        cmd += " --risk-only"
    if config.get("include_semi_auto"):
        cmd += " --include-semi-auto"
    if config.get("include_destructive"):
        cmd += " --include-destructive"
    if config.get("include_negative"):
        cmd += " --include-negative"
        negative_profile = str(config.get("negative_profile") or "").strip()
        if negative_profile:
            cmd += f" --negative-profile={quote(negative_profile)}"

    upload_file = str(config.get("upload_file") or "").strip()
    if upload_file:
        cmd += f" --upload-file={quote(upload_file)}"

    upload_profile_config = str(config.get("upload_profile_config") or "").strip()
    if upload_profile_config:
        cmd += f" --upload-profile-config={quote(upload_profile_config)}"

    html_path = config.get("html_path") or html_report_path(browser, page_id)
    cmd += f" --html={quote(html_path)}"
    return cmd


def _page_name(value: Any) -> str:
    return Path(str(value or "").replace("\\", "/")).name


def _add_option(options: Dict[str, Dict[str, Any]], page_id: Any, **meta: Any) -> None:
    page = _page_name(page_id)
    if not page:
        return
    key = page.lower()
    existing = options.setdefault(key, {"page_id": page, "sources": []})
    for field in ("entry_url", "action", "risk", "route_map_path"):
        if meta.get(field) and not existing.get(field):
            existing[field] = meta[field]
    source = meta.get("source")
    if source and source not in existing["sources"]:
        existing["sources"].append(source)


def load_page_options(
    *,
    mapping_path: Path = Path("generated/valid/page_mapping.json"),
    route_dir: Path = Path("generated/valid/route"),
    report_dir: Path = Path("output/regression"),
) -> List[Dict[str, Any]]:
    options: Dict[str, Dict[str, Any]] = {}

    if mapping_path.exists():
        try:
            payload = json.loads(mapping_path.read_text(encoding="utf-8"))
            for item in payload.get("page_mappings") or []:
                _add_option(
                    options,
                    item.get("page_id") or item.get("target_page"),
                    entry_url=item.get("entry_url") or item.get("resolved_entry_url"),
                    action=item.get("entry_action") or item.get("action") or item.get("action_path"),
                    risk=item.get("risk"),
                    source="mapping",
                )
        except Exception:
            pass

    if route_dir.exists():
        for path in route_dir.rglob("usable_route_map*.json"):
            try:
                payload = json.loads(path.read_text(encoding="utf-8"))
            except Exception:
                continue
            for route in list(payload.get("verified") or []) + list(payload.get("manual_verified") or []):
                source_route = route.get("source_route") or {}
                target = route.get("target_page") or route.get("target_page_name") or source_route.get("target_page") or source_route.get("target_page_name")
                _add_option(options, target, route_map_path=str(path), source="route")

    if report_dir.exists():
        for report in report_dir.rglob("regression_report.html"):
            parent = report.parent.name
            page = re.sub(r"^\d+_", "", parent)
            _add_option(options, page, source="recent")

    return sorted(options.values(), key=lambda item: item["page_id"].lower())


def page_option_labels(options: Iterable[Dict[str, Any]]) -> List[str]:
    labels = []
    for option in options:
        details = []
        if option.get("entry_url"):
            details.append(f"entry: {option['entry_url']}")
        if option.get("action"):
            details.append(f"action: {option['action']}")
        if option.get("risk"):
            details.append(f"risk: {option['risk']}")
        if option.get("route_map_path"):
            details.append("route: yes")
        suffix = f"    {' / '.join(details)}" if details else ""
        labels.append(f"{option['page_id']}{suffix}")
    return labels


def _column_index(headers: List[str], *names: str) -> Optional[int]:
    wanted = {name.strip().lower() for name in names if str(name or "").strip()}
    for index, header in enumerate(headers):
        if header.strip().lower() in wanted:
            return index
    return None


def _row_value(row: Iterable[Any], index: Optional[int]) -> str:
    values = list(row)
    if index is None or index >= len(values):
        return ""
    value = values[index]
    return "" if value is None else str(value).strip()


def _is_upload_case(
    *,
    case_type: Any,
    action_type: Any,
    title: Any,
    submit_locator: Any,
    main_step: Any,
) -> bool:
    case_lower = str(case_type or "").lower()
    action_lower = str(action_type or "").lower()
    title_text = str(title or "")
    main_step_lower = str(main_step or "").lower()
    if case_lower == "upload_submit" or action_lower == "upload_submit":
        return True
    if "アップロード確認" in title_text:
        return True
    if submit_locator:
        return True
    return "submit" in main_step_lower or "click" in main_step_lower


def load_upload_case_options(checklist_path: Any, page_id: Any) -> List[Dict[str, str]]:
    checklist = Path(str(checklist_path or ""))
    target = target_page_name(page_id)
    if not target or not checklist.exists():
        return []

    try:
        from openpyxl import load_workbook
    except Exception:
        return []

    try:
        workbook = load_workbook(checklist, data_only=True, read_only=True)
        sheet = workbook["Checklist"] if "Checklist" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        rows = list(sheet.iter_rows(values_only=True))
    except Exception:
        return []

    if not rows:
        return []

    headers = [str(value or "").strip().lower() for value in rows[0]]
    page_col = _column_index(headers, "page_id", "page", "jsp")
    case_id_col = _column_index(headers, "case_id", "id", "no")
    title_col = _column_index(headers, "test_title", "title", "test_viewpoint")
    mode_col = _column_index(headers, "automation_mode", "mode")
    case_type_col = _column_index(headers, "case_type")
    action_type_col = _column_index(headers, "action_type")
    locator_col = _column_index(headers, "locator")
    submit_locator_col = _column_index(headers, "submit_locator")
    main_step_col = _column_index(headers, "main_step")
    destructive_col = _column_index(headers, "destructive")
    enabled_col = _column_index(headers, "enabled", "enable")

    options: List[Dict[str, str]] = []
    seen = set()
    for row in rows[1:]:
        row_page = _row_value(row, page_col)
        if target_page_name(row_page) != target:
            continue

        case_id = _row_value(row, case_id_col)
        title = _row_value(row, title_col)
        case_type = _row_value(row, case_type_col)
        action_type = _row_value(row, action_type_col)
        submit_locator = _row_value(row, submit_locator_col)
        main_step = _row_value(row, main_step_col)
        locator = _row_value(row, locator_col)
        if not _is_upload_case(
            case_type=case_type,
            action_type=action_type,
            title=title,
            submit_locator=submit_locator,
            main_step=main_step,
        ):
            continue

        option_id = case_id or title or f"upload_case_{len(options) + 1}"
        key = option_id.lower()
        if key in seen:
            continue
        seen.add(key)
        options.append(
            {
                "case_id": option_id,
                "test_title": title,
                "automation_mode": _row_value(row, mode_col),
                "case_type": case_type or action_type or "upload_submit",
                "action_type": action_type or case_type or "upload_submit",
                "locator": locator,
                "submit_locator": submit_locator,
                "destructive": _row_value(row, destructive_col) or "false",
                "enabled": _row_value(row, enabled_col) or "true",
            }
        )
    return options


def upload_case_option_labels(cases: Iterable[Dict[str, Any]]) -> List[str]:
    labels = []
    for case in cases:
        details = []
        if case.get("test_title"):
            details.append(str(case["test_title"]))
        if case.get("automation_mode"):
            details.append(f"mode: {case['automation_mode']}")
        if case.get("locator"):
            details.append(f"locator: {case['locator']}")
        suffix = f"    {' / '.join(details)}" if details else ""
        labels.append(f"{case.get('case_id') or 'upload_case'}{suffix}")
    return labels


def _is_negative_case_type(case_type: Any, action_type: Any) -> bool:
    text = " ".join(str(value or "").lower() for value in (case_type, action_type))
    return "negative" in text or text.startswith("error_")


def load_negative_profile_options(checklist_path: Any, page_id: Any) -> List[Dict[str, str]]:
    checklist = Path(str(checklist_path or ""))
    target = target_page_name(page_id)
    if not target or not checklist.exists():
        return DEFAULT_NEGATIVE_PROFILES

    try:
        from openpyxl import load_workbook
    except Exception:
        return DEFAULT_NEGATIVE_PROFILES

    try:
        workbook = load_workbook(checklist, data_only=True, read_only=True)
        sheet = workbook["Checklist"] if "Checklist" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        rows = list(sheet.iter_rows(values_only=True))
    except Exception:
        return DEFAULT_NEGATIVE_PROFILES

    if not rows:
        return DEFAULT_NEGATIVE_PROFILES

    headers = [str(value or "").strip().lower() for value in rows[0]]
    page_col = _column_index(headers, "page_id", "page", "jsp")
    case_type_col = _column_index(headers, "case_type")
    action_type_col = _column_index(headers, "action_type")
    title_col = _column_index(headers, "test_title", "title", "test_viewpoint")
    mode_col = _column_index(headers, "automation_mode", "mode")

    defaults = {item["profile"]: item for item in DEFAULT_NEGATIVE_PROFILES}
    options: Dict[str, Dict[str, str]] = {}
    for row in rows[1:]:
        row_page = _row_value(row, page_col)
        if target_page_name(row_page) != target:
            continue
        case_type = _row_value(row, case_type_col)
        action_type = _row_value(row, action_type_col)
        if not _is_negative_case_type(case_type, action_type):
            continue
        profile = (case_type or action_type).strip()
        if not profile:
            continue
        default = defaults.get(profile, {})
        options[profile] = {
            "profile": profile,
            "label": default.get("label") or profile,
            "description": _row_value(row, title_col) or default.get("description") or _row_value(row, mode_col),
        }

    if not options:
        return DEFAULT_NEGATIVE_PROFILES
    for default in DEFAULT_NEGATIVE_PROFILES:
        options.setdefault(default["profile"], dict(default))
    return list(options.values())


def negative_profile_labels(options: Iterable[Dict[str, Any]]) -> List[str]:
    labels = []
    for option in options:
        label = option.get("label") or option.get("profile") or "negative"
        description = option.get("description")
        suffix = f"    {description}" if description else ""
        labels.append(f"{option.get('profile') or label}{suffix}")
    return labels
