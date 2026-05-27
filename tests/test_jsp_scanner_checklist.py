import json

from src.checklist_generator import generate_cases, page_entries, plan_page_specific_cases, universal_checklist_markdown_lines, write_excel
from src.jsp_scanner import scan_jsp_source


def test_spring_form_tags_are_classified_without_inflating_forms():
    source = """
    <form:form id="searchForm" action="/search" method="post">
      <form:input path="keyword" />
      <form:hidden path="token" />
      <form:select path="status">
        <form:option value="A" />
      </form:select>
      <html:file property="uploadFile" />
      <input type="button" name="search" value="Search" />
    </form:form>
    """

    result = scan_jsp_source(source, "Search.jsp")

    assert result["counts"] == {
        "form": 1,
        "field": 4,
        "file": 1,
        "button": 1,
    }

    by_tag = {element["tag"]: element["kind"] for element in result["elements"]}
    assert by_tag["form:form"] == "form"
    assert by_tag["form:input"] == "field"
    assert by_tag["form:hidden"] == "field"
    assert by_tag["form:select"] == "field"
    assert by_tag["form:option"] == "field"
    assert by_tag["html:file"] == "file"


def test_plain_anchor_links_are_scanned_and_runtime_alternatives_are_deduped():
    source = """
    <span class="t12"><a href="#" onclick="fnSubmit('/ProjectMemberUploadTemplateDownload.do');return false"><bean:message key="label.patlics.all456" bundle="PATLICS_MESSAGE" /></a></span>
    <% if(enl.equals("en")) { %>
      <a href="/help_en/project_106_01.html" target="winHelp"><span class="t12"><bean:message bundle='PATLICS_MESSAGE' key='label.mapCitationHead.comment3'/></span></a>
    <% } else { %>
      <a href="/help/project_106_01.html" target="winHelp"><span class="t12"><bean:message bundle='PATLICS_MESSAGE' key='label.mapCitationHead.comment3'/></span></a>
    <% } %>
    <input type="button" name="entry" value="upload" onClick="fnSubmit('/ProjectMemberUpload.do')" />
    <input type="button" name="cancell" value="cancel" onClick="javascript:window.close()" />
    """

    result = scan_jsp_source(source, "ProjectMemberUploadDisp.jsp")
    links = [element for element in result["elements"] if element["kind"] == "link"]
    buttons = [element for element in result["elements"] if element["kind"] == "button"]

    assert result["counts"]["link"] == 2
    assert result["counts"]["button"] == 2
    assert len(links) == 2
    assert len(buttons) == 2
    assert links[0]["locator"] == 'a[onclick="fnSubmit(\'/ProjectMemberUploadTemplateDownload.do\');return false"]'
    assert links[0]["label_key"] == "label.patlics.all456"
    assert links[1]["locator"] == "a[href='/help/project_106_01.html']"
    assert links[1]["label_key"] == "label.mapCitationHead.comment3"


def test_field_elements_generate_automation_ready_cases_and_form_evidence():
    scan_data = {
        "source": "Search.jsp",
        "counts": {"form": 1, "field": 2, "file": 1},
        "elements": [
            {
                "kind": "form",
                "tag": "form:form",
                "line": 1,
                "attributes": {"id": "searchForm", "action": "/search"},
                "locator": "#searchForm",
            },
            {
                "kind": "field",
                "tag": "form:input",
                "line": 2,
                "attributes": {"path": "keyword"},
                "locator": None,
            },
            {
                "kind": "field",
                "tag": "form:hidden",
                "line": 3,
                "attributes": {"path": "token"},
                "locator": None,
            },
            {
                "kind": "file",
                "tag": "html:file",
                "line": 4,
                "attributes": {"property": "uploadFile"},
                "locator": "[name='uploadFile']",
            },
        ],
    }

    cases = generate_cases(scan_data)

    assert len(cases) > 8
    assert {"page", "form", "field", "file"}.issubset({case.kind for case in cases})
    assert any(case.kind == "field" for case in cases)
    assert any(case.automation_mode == "auto" for case in cases)
    form_cases = [case for case in cases if case.kind == "form"]
    assert all("字段完整性校验" in case.evidence for case in form_cases)
    assert all("form:input `keyword`" in case.evidence for case in form_cases)


def test_page_mapping_input_uses_full_mappings_and_missing_elements():
    scan_data = {
        "high_risk_pages": [
            {
                "page_id": "HighOnly.jsp",
                "missing_legacy_elements": [
                    {
                        "kind": "button",
                        "line": 10,
                        "locator": "#high",
                        "label": "high",
                    }
                ],
            }
        ],
        "medium_risk_pages": [
            {
                "page_id": "MediumOnly.jsp",
                "missing_legacy_elements": [
                    {
                        "kind": "button",
                        "line": 20,
                        "locator": "#medium",
                        "label": "medium",
                    }
                ],
            }
        ],
        "page_mappings": [
            {
                "page_id": "MatchedA.jsp",
                "elements": [
                    {
                        "kind": "form",
                        "tag": "form",
                        "line": 1,
                        "attributes": {"id": "matchedForm", "action": "/save"},
                        "locator": "#matchedForm",
                    }
                ],
                "missing_legacy_elements": [
                    {
                        "kind": "button",
                        "line": 2,
                        "locator": "#missing",
                        "label": "missing",
                    }
                ],
            },
            {
                "page_id": "MatchedB.jsp",
                "missing_legacy_elements": [
                    {
                        "kind": "link",
                        "line": 3,
                        "locator": "a.details",
                        "label": "details",
                    }
                ],
            },
            {
                "page_id": "MatchedC.jsp",
                "missing_legacy_elements": [],
            },
        ],
    }

    pages = list(page_entries(scan_data))
    cases = generate_cases(scan_data)

    assert [page["page_id"] for page in pages] == ["MatchedA.jsp", "MatchedB.jsp", "MatchedC.jsp"]
    assert {case.page for case in cases} == {"MatchedA.jsp", "MatchedB.jsp", "MatchedC.jsp"}
    assert {"initial_display", "negative_js_error"}.issubset({case.case_type for case in cases})
    assert {"auto", "auto-negative"}.issuperset({case.automation_mode for case in cases})
    assert all(case.generated_by in {"PageCasePlanner", "CaseExpansionRules"} for case in cases)


def test_universal_checklist_lines_include_xls_categories():
    lines = universal_checklist_markdown_lines()
    text = "\n".join(lines)

    assert "1-1 画面レイアウト" in text
    assert "7 ファイルアップロード" in text
    assert "13 マルチブラウザ動作確認" in text


def test_page_specific_excel_contains_profile_and_skipped_sheets(tmp_path):
    from openpyxl import load_workbook

    scan_data = {
        "page_mappings": [
            {
                "page_id": "Upload.jsp",
                "elements": [
                    {"kind": "file", "tag": "html:file", "attributes": {"property": "uploadFile"}, "locator": "input[name='uploadFile']"},
                    {"kind": "button", "tag": "input", "attributes": {"onclick": "fnSubmit('/Upload.do')"}, "locator": "input[name='entry']"},
                ],
            }
        ]
    }
    output = tmp_path / "migration_checklist.xlsx"

    write_excel(output, scan_data, generate_cases(scan_data))

    workbook = load_workbook(output, read_only=True)
    assert "Checklist" in workbook.sheetnames
    assert "PageProfile" in workbook.sheetnames
    assert "SkippedTemplates" in workbook.sheetnames


def test_checklist_generation_appends_existing_runtime_profile_pages(tmp_path):
    runtime_dir = tmp_path / "runtime_profile"
    runtime_dir.mkdir()
    profile_path = runtime_dir / "legacy_runtimeupload_route1.json"
    profile_path.write_text(
        json.dumps(
            {
                "schema": "moonlight.runtime_page_profile.v1",
                "page_id": "RuntimeUpload.jsp",
                "target_page": "RuntimeUpload.jsp",
                "route_id": "route1",
                "side": "legacy",
                "controls": [
                    {
                        "tag": "input",
                        "type": "file",
                        "name": "uploadFile",
                        "selector": "input[name=\"uploadFile\"][type=\"file\"]",
                        "visible": True,
                    },
                    {
                        "tag": "input",
                        "type": "button",
                        "name": "entry",
                        "value": "Upload",
                        "onclick": "fnSubmit('/RuntimeUpload.do')",
                        "selector": "input[name=\"entry\"][type=\"button\"]",
                        "visible": True,
                    },
                ],
            }
        ),
        encoding="utf-8",
    )
    scan_data = {"page_mappings": [{"page_id": "StaticOnly.jsp", "elements": []}]}

    without_runtime = generate_cases(scan_data, runtime_profile_dir=runtime_dir, include_runtime_profile_dir=False)
    with_runtime = generate_cases(scan_data, runtime_profile_dir=runtime_dir, include_runtime_profile_dir=True)
    _, _, profiles = plan_page_specific_cases(scan_data, runtime_profile_dir=runtime_dir, include_runtime_profile_dir=True)

    assert {case.page for case in without_runtime} == {"StaticOnly.jsp"}
    assert {"StaticOnly.jsp", "RuntimeUpload.jsp"}.issubset({case.page for case in with_runtime})
    assert any(case.page == "RuntimeUpload.jsp" and case.case_type == "upload_submit" for case in with_runtime)
    runtime_profiles = [profile for profile in profiles if profile.get("page_id") == "RuntimeUpload.jsp"]
    assert runtime_profiles
    assert runtime_profiles[0]["runtime_profile_path"] == str(profile_path)
