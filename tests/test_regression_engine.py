import json
from pathlib import Path

import pytest
from PIL import Image

from src.action_executor import (
    _accept_dialog_safely,
    _capture_state,
    _download_save_path,
    _negative_visual_evidence_payload,
    _opens_popup_hint,
    _render_console_evidence_image,
    _resolve_upload_file_value,
    _safe_download_filename,
    build_steps_from_page_mapping,
    infer_semantic_action,
)
import src.regression_engine as regression_engine_module
from src.assert_engine import compare_visual_screenshot
from src.regression_engine import RegressionEngine
from src.route_navigator import RouteMapCatalog


def test_compare_visual_screenshot_writes_diff(tmp_path):
    legacy = tmp_path / "legacy.png"
    new = tmp_path / "new.png"
    diff = tmp_path / "diff.png"
    Image.new("RGB", (2, 2), "white").save(legacy)
    image = Image.new("RGB", (2, 2), "white")
    image.putpixel((0, 0), (0, 0, 0))
    image.save(new)

    result = compare_visual_screenshot(str(legacy), str(new), str(diff), threshold_percent=0)

    assert result["status"] == "DIFF"
    assert result["diff_percent"] == 25.0
    assert diff.exists()


def test_select_pages_orders_high_then_medium(tmp_path):
    mapping = {
        "page_mappings": [
            {"page_id": "low.jsp", "risk": "Low"},
            {"page_id": "medium.jsp", "risk": "Medium"},
            {"page_id": "high.jsp", "risk": "High"},
        ]
    }
    mapping_path = tmp_path / "page_mapping.json"
    mapping_path.write_text(json.dumps(mapping), encoding="utf-8")

    engine = RegressionEngine(mapping_path=str(mapping_path), output_dir=str(tmp_path / "out"))
    pages = engine.select_pages(risk_only=True)

    assert [page["page_id"] for page in pages] == ["high.jsp", "medium.jsp"]


def test_select_pages_target_page_ignores_risk(tmp_path):
    mapping = {
        "page_mappings": [
            {"page_id": "low.jsp", "risk": "Low"},
            {"page_id": "high.jsp", "risk": "High"},
        ]
    }
    mapping_path = tmp_path / "page_mapping.json"
    mapping_path.write_text(json.dumps(mapping), encoding="utf-8")

    engine = RegressionEngine(mapping_path=str(mapping_path), output_dir=str(tmp_path / "out"))
    pages = engine.select_pages(risk_only=True, target_page="low.jsp")

    assert [page["page_id"] for page in pages] == ["low.jsp"]


def test_select_pages_target_page_reports_missing_mapping(tmp_path):
    mapping = {"page_mappings": [{"page_id": "exists.jsp", "risk": "High"}]}
    mapping_path = tmp_path / "page_mapping.json"
    mapping_path.write_text(json.dumps(mapping), encoding="utf-8")

    engine = RegressionEngine(mapping_path=str(mapping_path), output_dir=str(tmp_path / "out"))

    try:
        engine.select_pages(target_page="missing.jsp")
    except ValueError as exc:
        message = str(exc)
    else:
        raise AssertionError("Expected missing target page to raise ValueError")

    assert "Target JSP page not found" in message
    assert "missing.jsp" in message
    assert "exists.jsp" in message


def test_target_page_name_normalizes_case_and_action_suffix():
    assert RegressionEngine._target_page_name("/docroot/ProjectListUploadDisp.do") == "projectlistuploaddisp.jsp"
    assert RegressionEngine._target_page_name("projectlistuploaddisp.jsp") == "projectlistuploaddisp.jsp"


def test_route_map_catalog_selects_verified_route_for_target(tmp_path):
    route_map = tmp_path / "usable_route_map.json"
    route_map.write_text(
        json.dumps(
            {
                "schema": "moonlight.usable_route_map.v1",
                "verified": [
                    {
                        "route_id": "r1",
                        "status": "verified",
                        "target_page": "ProjectMemberUploadDisp.jsp",
                        "target_page_name": "projectmemberuploaddisp.jsp",
                        "source_route": {"length": 3},
                    }
                ],
            }
        ),
        encoding="utf-8",
    )

    catalog = RouteMapCatalog([route_map])
    route = catalog.find_for_target("ProjectMemberUploadDisp.jsp")

    assert route["route_id"] == "r1"
    assert route["route_map_path"] == str(route_map)


def test_page_matches_mapping_checks_frame_urls(tmp_path):
    mapping_path = tmp_path / "page_mapping.json"
    mapping_path.write_text(json.dumps({"page_mappings": []}), encoding="utf-8")
    engine = RegressionEngine(mapping_path=str(mapping_path), output_dir=str(tmp_path / "out"))

    class Frame:
        def __init__(self, url):
            self.url = url

    class Page:
        url = "https://legacy.example/patlics/PatlicsTopMain.do"
        frames = [
            Frame("about:blank"),
            Frame("https://legacy.example/patlics/ProjectMemberUploadDisp.do"),
        ]

    assert engine._page_matches_mapping(
        Page(),
        {"page_id": "ProjectMemberUploadDisp.jsp", "entry_url": "ProjectMemberUploadDisp.do"},
    )


def test_render_report_contains_side_by_side_sections(tmp_path):
    mapping_path = tmp_path / "page_mapping.json"
    mapping_path.write_text(json.dumps({"page_mappings": []}), encoding="utf-8")
    output_dir = tmp_path / "out"
    output_dir.mkdir()
    screenshot = output_dir / "shot.png"
    Image.new("RGB", (1, 1), "white").save(screenshot)

    engine = RegressionEngine(mapping_path=str(mapping_path), output_dir=str(output_dir))
    report = engine.render_report(
        [
            {
                "page_id": "A.jsp",
                "risk": "High",
                "action": "page_snapshot",
                "status": "PASS",
                "url_match": True,
                "dom_match": True,
                "legacy_screenshot": str(screenshot),
                "new_screenshot": str(screenshot),
                "diff_screenshot": str(screenshot),
                "visual": {"diff_percent": 0.0},
            }
        ]
    )

    html = Path(report).read_text(encoding="utf-8")
    assert "Legacy" in html
    assert "New" in html
    assert "Diff" in html
    assert "Result JSON" in html
    assert (Path(report).parent / "regression_results.json").exists()
    assert "Checklist Cases" in html
    assert "No checklist cases were loaded for this report." in html


def test_render_report_for_single_page_goes_next_to_screenshots(tmp_path):
    mapping_path = tmp_path / "page_mapping.json"
    mapping_path.write_text(json.dumps({"page_mappings": []}), encoding="utf-8")
    output_dir = tmp_path / "out"
    page_dir = output_dir / "0001_ProjectListUploadDisp.jsp"
    page_dir.mkdir(parents=True)
    screenshot = page_dir / "00_legacy_initial.png"
    Image.new("RGB", (1, 1), "white").save(screenshot)

    engine = RegressionEngine(mapping_path=str(mapping_path), output_dir=str(output_dir))
    report = engine.render_report(
        [
            {
                "page_id": "ProjectListUploadDisp.jsp",
                "risk": "High",
                "action": "page_snapshot",
                "status": "PASS",
                "url_match": True,
                "legacy_screenshot": str(screenshot),
                "new_screenshot": str(screenshot),
                "diff_screenshot": str(screenshot),
                "visual": {"diff_percent": 0.0},
            }
        ]
    )

    assert Path(report).parent == page_dir
    html = Path(report).read_text(encoding="utf-8")
    assert "Page: ProjectListUploadDisp.jsp" in html
    assert 'src="00_legacy_initial.png"' in html


def test_build_steps_from_page_mapping_is_risk_first():
    steps = build_steps_from_page_mapping(
        {
            "page_mappings": [
                {"page_id": "m.jsp", "risk": "Medium", "locator_changes": []},
                {"page_id": "h.jsp", "risk": "High", "locator_changes": []},
                {"page_id": "l.jsp", "risk": "Low", "locator_changes": []},
            ]
        }
    )

    assert [step["page_id"] for step in steps] == ["h.jsp", "m.jsp"]


def test_infer_semantic_action_uses_scanner_hints():
    assert infer_semantic_action("click", {"kind": "form", "locator": "[name='SearchForm']"}) == "submit"
    assert infer_semantic_action(None, {"kind": "file", "action_hint": "upload"}) == "upload"
    assert infer_semantic_action("click", {"raw": '<form:select path="country">'}) == "select"
    assert infer_semantic_action(None, {"kind": "link", "raw": '<html:link href="/next">'}) == "navigate"
    assert infer_semantic_action("negative_http_500", {}) == "negative_http_500"
    assert infer_semantic_action("negative_js_error", {}) == "negative_js_error"
    assert infer_semantic_action("check", {"kind": "checkbox"}) == "check"
    assert infer_semantic_action("uncheck", {"kind": "checkbox"}) == "uncheck"
    assert infer_semantic_action("file_download", {"locator": 'input[name="btSave"][type="button"]'}) == "download"
    assert infer_semantic_action("download_template", {"kind": "link"}) == "download"


def test_action_dedupe_prefers_locator_change_over_full_action_fallback(tmp_path):
    mapping_path = tmp_path / "page_mapping.json"
    mapping_path.write_text(json.dumps({"page_mappings": []}), encoding="utf-8")
    engine = RegressionEngine(mapping_path=str(mapping_path), output_dir=str(tmp_path / "out"))
    actions, skipped = engine._dedupe_actions(
        [
            {
                "kind": "button",
                "label": "save",
                "semantic_key": "action:save",
                "legacy_locator": "[name='save']",
                "new_locator": "#save",
            },
            {
                "kind": "button",
                "label": "save",
                "semantic_key": "action:save",
                "legacy_locator": "[name='save']",
                "new_locator": "[name='save']",
            },
            {
                "kind": "link",
                "label": "help",
                "semantic_key": "locator:a[href='/help.html']",
                "legacy_locator": "a[href='/help.html']",
                "new_locator": "a[href='/help.html']",
            },
        ]
    )

    assert len(actions) == 2
    assert len(skipped) == 1
    assert actions[0]["new_locator"] == "#save"


def test_popup_hint_detects_targeted_links():
    assert _opens_popup_hint({"attributes": {"target": "winHelp"}}) is True
    assert _opens_popup_hint({"attributes": {"target": "_self"}}) is False


def test_download_filename_is_windows_safe():
    name = _safe_download_filename('a[onclick="x"]?.xls', "chrome:port")

    assert name == "a_onclick_x_chrome_port.xls"


def test_download_save_path_uses_env_and_preserves_suggested_filename(tmp_path, monkeypatch):
    monkeypatch.setenv("DOWNLOAD_DIR", str(tmp_path))

    path = _download_save_path("プロジェクトリストアップロード.tsv")

    assert path == tmp_path / "プロジェクトリストアップロード.tsv"


def test_download_save_path_does_not_overwrite_existing_file(tmp_path, monkeypatch):
    monkeypatch.setenv("DOWNLOAD_DIR", str(tmp_path))
    existing = tmp_path / "report.pdf"
    existing.write_text("legacy", encoding="utf-8")

    path = _download_save_path("report.pdf")

    assert path == tmp_path / "report (1).pdf"


def test_console_evidence_image_renders_events(tmp_path):
    image_path = _render_console_evidence_image(
        [
            {"time": "12:00:01", "type": "CONSOLE", "level": "error", "detail": "error: boom", "url": "http://example/js"},
            {"time": "12:00:02", "type": "HTTP_ERROR", "level": "http_error", "detail": "HTTP 500", "status": 500, "url": "http://example/api"},
        ],
        tmp_path,
        "case_001",
    )

    assert Path(image_path).exists()
    with Image.open(image_path) as image:
        assert image.width >= 1000
        assert image.height >= 200


def test_negative_visual_evidence_payload_names_visible_error_state():
    payload = _negative_visual_evidence_payload(
        "negative_http_500",
        detail="UploadConf failed",
        url="**/UploadConf.do*",
        phase="after trigger",
    )

    assert payload["title"] == "Simulated HTTP 500"
    assert payload["phase"] == "after trigger"
    assert payload["detail"] == "UploadConf failed"
    assert payload["url"] == "**/UploadConf.do*"


def test_accept_dialog_safely_ignores_already_handled_dialog():
    class Dialog:
        def accept(self):
            raise RuntimeError("Dialog.accept: Cannot accept dialog which is already handled!")

    assert _accept_dialog_safely(Dialog()) == "already_handled"


def test_accept_dialog_safely_reports_accept_success():
    class Dialog:
        accepted = False

        def accept(self):
            self.accepted = True

    dialog = Dialog()

    assert _accept_dialog_safely(dialog) == "accepted"
    assert dialog.accepted is True


def test_upload_file_resolver_handles_placeholders_and_multiple_files(tmp_path):
    explicit = tmp_path / "valid.tsv"
    explicit.write_text("id\tname\n1\tmoonlight\n", encoding="utf-8")
    other = tmp_path / "other.tsv"
    other.write_text("id\tname\n2\tluna\n", encoding="utf-8")

    assert _resolve_upload_file_value(str(explicit), tmp_path) == str(explicit)
    assert _resolve_upload_file_value([str(explicit), str(other)], tmp_path) == [str(explicit), str(other)]

    placeholder = _resolve_upload_file_value("${UPLOAD_FILE}", tmp_path)
    assert Path(placeholder).exists()

    fakepath = _resolve_upload_file_value(r"C:\fakepath\missing.tsv", tmp_path)
    assert Path(fakepath).exists()


def test_upload_profile_resolver_prefers_checklist_then_profile_then_global(tmp_path):
    mapping_path = tmp_path / "page_mapping.json"
    mapping_path.write_text(json.dumps({"page_mappings": []}), encoding="utf-8")
    checklist_file = tmp_path / "checklist.tsv"
    profile_file = tmp_path / "profile.tsv"
    case_profile_file = tmp_path / "case_profile.tsv"
    fallback_file = tmp_path / "fallback.tsv"
    for path in (checklist_file, profile_file, case_profile_file, fallback_file):
        path.write_text("id\tname\n1\tmoonlight\n", encoding="utf-8")
    profile_config = tmp_path / "upload_profiles.json"
    profile_config.write_text(
        json.dumps(
            {
                "upload_profiles": [
                    {
                        "name": "valid_project_list_tsv",
                        "file": str(profile_file),
                        "page_patterns": ["ProjectListUploadDisp.jsp"],
                        "case_types": ["upload_submit"],
                        "locator": "input[name='uploadFile']",
                        "negative": False,
                    },
                    {
                        "name": "case_specific_upload",
                        "case_id": "project-upload-valid",
                        "file": str(case_profile_file),
                        "page_patterns": ["ProjectListUploadDisp.jsp"],
                        "case_types": ["upload_submit"],
                        "locator": "input[name='uploadFile']",
                        "negative": False,
                    }
                ]
            }
        ),
        encoding="utf-8",
    )
    engine = RegressionEngine(
        mapping_path=str(mapping_path),
        output_dir=str(tmp_path / "out"),
        upload_file=str(fallback_file),
        upload_profile_config=str(profile_config),
    )
    action_case = {"page_id": "ProjectListUploadDisp.jsp", "case_id": "project-upload-valid", "case_type": "upload_submit", "action_type": "upload_submit"}
    step = {"action_type": "upload", "locator": "input[name='uploadFile']"}

    assert engine._resolve_upload_value(str(checklist_file), action_case, step, "input[name='uploadFile']") == str(checklist_file)
    assert engine._resolve_upload_value(r"C:\fakepath\bad.tsv", action_case, step, "input[name='uploadFile']") == str(case_profile_file)
    assert engine._resolve_upload_value(r"C:\fakepath\bad.tsv", {**action_case, "case_id": "other-upload"}, step, "input[name='uploadFile']") == str(profile_file)
    assert engine._resolve_upload_value("${UPLOAD_FILE}", {**action_case, "page_id": "Other.jsp"}, step, "input[name='uploadFile']") == str(fallback_file)


def test_upload_submit_button_step_clicks_onclick_instead_of_request_submit(tmp_path, monkeypatch):
    mapping_path = tmp_path / "page_mapping.json"
    mapping_path.write_text(json.dumps({"page_mappings": []}), encoding="utf-8")
    upload_file = tmp_path / "upload.tsv"
    upload_file.write_text("id\tname\n1\tmoonlight\n", encoding="utf-8")
    engine = RegressionEngine(mapping_path=str(mapping_path), output_dir=str(tmp_path / "out"), upload_file=str(upload_file))
    calls = []

    def fake_execute_action(page, action_type, locator, value=None, **kwargs):
        context = kwargs.get("action_context") or {}
        calls.append(
            {
                "action_type": action_type,
                "locator": locator,
                "semantic_action": infer_semantic_action(action_type, context),
            }
        )
        return {"status": "PASS", "state": {"url": "http://example.test/after", "screenshot": str(tmp_path / "after.png")}}

    monkeypatch.setattr(regression_engine_module, "execute_action", fake_execute_action)
    action_case = {
        "case_id": "upload-1",
        "case_type": "upload_submit",
        "action_type": "upload_submit",
        "page_id": "Upload.jsp",
        "pre_steps": [{"action_type": "upload", "locator": "input[name='uploadFile']", "value": "${UPLOAD_FILE}"}],
        "main_step": {
            "action_type": "submit",
            "locator": "input[onclick*=\"submitForm('UploadForm','./UploadConfirm.do','')\"]",
        },
    }

    result = engine._execute_action_case(
        object(),
        action_case,
        side="legacy",
        browser_name="chrome_port",
        capture_dir=tmp_path,
        test_id="upload_case",
    )

    assert result["status"] == "PASS"
    assert calls[0]["semantic_action"] == "upload"
    assert calls[1]["action_type"] == "click"
    assert calls[1]["semantic_action"] == "click"


def test_scenario_stops_after_step_closes_page(tmp_path, monkeypatch):
    mapping_path = tmp_path / "page_mapping.json"
    mapping_path.write_text(json.dumps({"page_mappings": []}), encoding="utf-8")
    engine = RegressionEngine(mapping_path=str(mapping_path), output_dir=str(tmp_path / "out"))
    calls = []

    def fake_execute_action(page, action_type, locator, value=None, **kwargs):
        calls.append((action_type, locator))
        return {
            "status": "PASS",
            "page_closed_after_action": True,
            "download_filename": "report.pdf",
            "state": {"url": "about:closed", "page_closed": True, "screenshot": str(tmp_path / "closed.png")},
        }

    monkeypatch.setattr(regression_engine_module, "execute_action", fake_execute_action)
    action_case = {
        "case_id": "download-misgenerated",
        "case_type": "upload_submit",
        "action_type": "upload_submit",
        "page_id": "Download.jsp",
        "pre_steps": [{"action_type": "download", "locator": "#download"}],
        "main_step": {"action_type": "submit", "locator": "form[name='DownloadForm']"},
    }

    result = engine._execute_action_case(
        object(),
        action_case,
        side="legacy",
        browser_name="edge",
        capture_dir=tmp_path,
        test_id="download_case",
    )

    assert result["status"] == "PASS"
    assert result["page_closed_after_action"] is True
    assert calls == [("download", "#download")]


def test_checklist_loader_filters_optional_modes(tmp_path):
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    mapping_path = tmp_path / "page_mapping.json"
    mapping_path.write_text(json.dumps({"page_mappings": []}), encoding="utf-8")
    checklist = tmp_path / "migration_checklist.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Checklist"
    sheet.append(
        [
            "case_id",
            "page_id",
            "automation_mode",
            "case_type",
            "action_type",
            "locator",
            "submit_locator",
            "test_data",
            "destructive",
            "enabled",
        ]
    )
    sheet.append(["auto-1", "Page.jsp", "auto", "snapshot", "snapshot", "", "", "", "false", "true"])
    sheet.append(["semi-1", "Page.jsp", "semi-auto", "click", "click", "#semi", "", "", "false", "true"])
    sheet.append(["destroy-1", "Page.jsp", "auto", "delete_action", "click", "#delete", "", "", "true", "true"])
    sheet.append(["neg-1", "Page.jsp", "auto", "negative_file_upload", "upload", "input[type='file']", "", "", "false", "true"])
    workbook.save(checklist)

    default_engine = RegressionEngine(mapping_path=str(mapping_path), output_dir=str(tmp_path / "out"), checklist_path=str(checklist))
    default_cases = default_engine._load_checklist_cases("Page.jsp")
    assert [case["label"] for case in default_cases] == ["auto-1"]
    assert default_cases[0]["case_id"] == "auto-1"
    coverage = {row["case_id"]: row for row in default_engine._checklist_case_rows["page.jsp"]}
    assert set(coverage) == {"auto-1", "semi-1", "destroy-1", "neg-1"}
    assert coverage["auto-1"]["excluded_reason"] == ""
    assert "--include-semi-auto" in coverage["semi-1"]["excluded_reason"]
    assert "--include-destructive" in coverage["destroy-1"]["excluded_reason"]
    assert "--include-negative" in coverage["neg-1"]["excluded_reason"]

    full_engine = RegressionEngine(
        mapping_path=str(mapping_path),
        output_dir=str(tmp_path / "out2"),
        checklist_path=str(checklist),
        include_semi_auto=True,
        include_destructive=True,
        include_negative=True,
        negative_profile="negative_file_upload",
    )
    full_cases = full_engine._load_checklist_cases("Page.jsp")
    assert {case["label"] for case in full_cases} == {"auto-1", "semi-1", "destroy-1", "neg-1"}


def test_checklist_loader_preserves_negative_expected_value_and_steps(tmp_path):
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    mapping_path = tmp_path / "page_mapping.json"
    mapping_path.write_text(json.dumps({"page_mappings": []}), encoding="utf-8")
    checklist = tmp_path / "migration_checklist.xlsx"
    pre_steps = json.dumps([{"action_type": "upload", "locator": "input[type='file']", "value": "${UPLOAD_FILE}"}])
    main_step = json.dumps({"action_type": "negative_http_500", "locator": "#submit", "expected_value": "**/UploadConf.do*"})

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Checklist"
    sheet.append(
        [
            "case_id",
            "page_id",
            "automation_mode",
            "case_type",
            "action_type",
            "locator",
            "expected_type",
            "expected_value",
            "pre_steps",
            "main_step",
            "destructive",
            "enabled",
        ]
    )
    sheet.append(["neg-http", "Page.jsp", "auto-negative", "negative_http_500", "negative_http_500", "#submit", "http_error", "**/UploadConf.do*", pre_steps, main_step, "false", "true"])
    workbook.save(checklist)

    engine = RegressionEngine(
        mapping_path=str(mapping_path),
        output_dir=str(tmp_path / "out"),
        checklist_path=str(checklist),
        include_negative=True,
        negative_profile="negative_http_500",
    )
    cases = engine._load_checklist_cases("Page.jsp")

    assert len(cases) == 1
    assert cases[0]["expected_value"] == "**/UploadConf.do*"
    assert cases[0]["pre_steps"][0]["action_type"] == "upload"
    assert cases[0]["main_step"]["action_type"] == "negative_http_500"


def test_checklist_loader_keeps_download_case_out_of_upload_submit(tmp_path):
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    mapping_path = tmp_path / "page_mapping.json"
    mapping_path.write_text(json.dumps({"page_mappings": []}), encoding="utf-8")
    checklist = tmp_path / "migration_checklist.xlsx"
    pre_steps = json.dumps([{"action_type": "upload", "locator": 'input[name="btSave"][type="button"]', "value": ""}])
    main_step = json.dumps({"action_type": "submit", "locator": "form[name='fmPDFDownload']"})

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Checklist"
    sheet.append(
        [
            "case_id",
            "page_id",
            "automation_mode",
            "case_type",
            "action_type",
            "locator",
            "expected_type",
            "pre_steps",
            "main_step",
            "destructive",
            "enabled",
            "test_title",
        ]
    )
    sheet.append(
        [
            "jpgazettepdfdownloaddisp-file_download-001",
            "JpGazettePDFDownloadDisp.jsp",
            "auto",
            "upload_submit",
            "upload_submit",
            'input[name="btSave"][type="button"]',
            "download",
            pre_steps,
            main_step,
            "false",
            "true",
            "ファイル出力ダウンロード確認",
        ]
    )
    workbook.save(checklist)

    engine = RegressionEngine(mapping_path=str(mapping_path), output_dir=str(tmp_path / "out"), checklist_path=str(checklist))
    cases = engine._load_checklist_cases("JpGazettePDFDownloadDisp.jsp")

    assert len(cases) == 1
    assert cases[0]["case_type"] == "file_download"
    assert cases[0]["action_type"] == "file_download"
    assert cases[0]["legacy_locator"] == 'input[name="btSave"][type="button"]'
    assert "pre_steps" not in cases[0]
    assert "main_step" not in cases[0]


def test_checklist_loader_preserves_download_option_steps(tmp_path):
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    mapping_path = tmp_path / "page_mapping.json"
    mapping_path.write_text(json.dumps({"page_mappings": []}), encoding="utf-8")
    checklist = tmp_path / "migration_checklist.xlsx"
    pre_steps = json.dumps(
        [
            {"action_type": "check", "locator": 'input[name="cbKind"][type="checkbox"][value="11"]'},
            {"action_type": "uncheck", "locator": 'input[name="cbKind"][type="checkbox"][value="10"]'},
        ]
    )
    main_step = json.dumps({"action_type": "download", "locator": 'input[name="btSave"][type="button"]'})

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Checklist"
    sheet.append(
        [
            "case_id",
            "page_id",
            "automation_mode",
            "case_type",
            "action_type",
            "locator",
            "expected_type",
            "pre_steps",
            "main_step",
            "destructive",
            "enabled",
            "test_title",
        ]
    )
    sheet.append(
        [
            "jpgazettepdfdownloaddisp-file_download-001-option",
            "JpGazettePDFDownloadDisp.jsp",
            "auto",
            "file_download",
            "download",
            'input[name="btSave"][type="button"]',
            "download",
            pre_steps,
            main_step,
            "false",
            "true",
            "出力条件代表パターン確認",
        ]
    )
    workbook.save(checklist)

    engine = RegressionEngine(mapping_path=str(mapping_path), output_dir=str(tmp_path / "out"), checklist_path=str(checklist))
    cases = engine._load_checklist_cases("JpGazettePDFDownloadDisp.jsp")

    assert len(cases) == 1
    assert cases[0]["pre_steps"][0]["action_type"] == "check"
    assert cases[0]["pre_steps"][1]["action_type"] == "uncheck"
    assert cases[0]["main_step"]["action_type"] == "download"
    assert cases[0]["main_step"]["locator"] == 'input[name="btSave"][type="button"]'


def test_report_coverage_matrix_renders_checklist_case_rows(tmp_path):
    mapping_path = tmp_path / "page_mapping.json"
    mapping_path.write_text(json.dumps({"page_mappings": []}), encoding="utf-8")
    output_dir = tmp_path / "out"
    output_dir.mkdir()
    screenshot = output_dir / "shot.png"
    Image.new("RGB", (1, 1), "white").save(screenshot)

    engine = RegressionEngine(mapping_path=str(mapping_path), output_dir=str(output_dir))
    engine._checklist_case_rows = {
        "upload.jsp": [
            {
                "case_id": "upload-001",
                "test_title": "Upload main path",
                "automation_mode": "auto",
                "destructive": "false",
                "excluded_reason": "",
            },
            {
                "case_id": "upload-002",
                "test_title": "Delete uploaded row",
                "automation_mode": "auto-db",
                "destructive": "true",
                "excluded_reason": "destructive=true; requires --include-destructive",
            },
        ],
        "download.jsp": [
            {
                "case_id": "download-001",
                "test_title": "Download template",
                "automation_mode": "auto",
                "destructive": "false",
                "excluded_reason": "",
            }
        ],
    }
    report = engine.render_report(
        [
            {
                "page_id": "Upload.jsp",
                "risk": "High",
                "action": "uploadFile",
                "action_type": "upload",
                "status": "PASS",
                "url_match": True,
                "dom_match": True,
                "legacy_screenshot": str(screenshot),
                "new_screenshot": str(screenshot),
                "diff_screenshot": str(screenshot),
                "visual": {"diff_percent": 0.0},
            },
            {
                "page_id": "Download.jsp",
                "risk": "High",
                "action": "TemplateDownload",
                "action_type": "download",
                "status": "PASS",
                "url_match": True,
                "dom_match": True,
                "legacy_screenshot": str(screenshot),
                "new_screenshot": str(screenshot),
                "diff_screenshot": str(screenshot),
                "visual": {"diff_percent": 0.0},
                "download_filename_match": True,
                "legacy_download_filename": "template.tsv",
                "new_download_filename": "template.tsv",
                "legacy_download_path": str(tmp_path / "template.tsv"),
                "new_download_path": str(tmp_path / "template.tsv"),
                "legacy_console_screenshot": str(screenshot),
                "new_console_screenshot": str(screenshot),
                "legacy_console_error_count": 1,
                "new_console_error_count": 0,
                "legacy_http_error_count": 0,
                "new_http_error_count": 1,
                "legacy_request_failed_count": 0,
                "new_request_failed_count": 0,
            },
        ]
    )

    html = Path(report).read_text(encoding="utf-8")
    assert "Checklist Cases" in html
    assert "upload-001" in html
    assert "Upload main path" in html
    assert "auto-db" in html
    assert "destructive=true; requires --include-destructive" in html
    assert "download-001" in html
    assert "Download filename match" in html
    assert "template.tsv" in html
    assert "Legacy Console" in html
    assert "Console errors" in html
    assert "HTTP errors" in html


def test_download_compare_marks_filename_mismatch_as_diff(tmp_path):
    mapping_path = tmp_path / "page_mapping.json"
    mapping_path.write_text(json.dumps({"page_mappings": []}), encoding="utf-8")
    screenshot = tmp_path / "shot.png"
    Image.new("RGB", (1, 1), "white").save(screenshot)

    engine = RegressionEngine(mapping_path=str(mapping_path), output_dir=str(tmp_path))
    compared = engine._compare_state(
        "Download.jsp",
        "High",
        "download",
        {"url": "https://legacy/app", "dom": "", "screenshot": str(screenshot)},
        {"url": "https://new/app", "dom": "", "screenshot": str(screenshot)},
        tmp_path / "diff.png",
        action_type="download",
        legacy_action={"status": "PASS", "saved_filename": "legacy.tsv", "download_path": str(tmp_path / "legacy.tsv")},
        new_action={"status": "PASS", "saved_filename": "new.tsv", "download_path": str(tmp_path / "new.tsv")},
    )

    assert compared["status"] == "DIFF"
    assert compared["download_filename_match"] is False
    assert compared["legacy_download_filename"] == "legacy.tsv"
    assert compared["new_download_filename"] == "new.tsv"


def test_download_compare_uses_suggested_filename_before_unique_saved_name(tmp_path):
    mapping_path = tmp_path / "page_mapping.json"
    mapping_path.write_text(json.dumps({"page_mappings": []}), encoding="utf-8")
    engine = RegressionEngine(mapping_path=str(mapping_path), output_dir=str(tmp_path))

    compared = engine._compare_state(
        "Download.jsp",
        "Low",
        "file_download",
        {"url": "about:closed", "dom": "", "screenshot": ""},
        {"url": "about:closed", "dom": "", "screenshot": ""},
        tmp_path / "diff.png",
        action_type="file_download",
        legacy_action={"status": "PASS", "download_filename": "report.pdf", "saved_filename": "report.pdf"},
        new_action={"status": "PASS", "download_filename": "report.pdf", "saved_filename": "report (1).pdf"},
    )

    assert compared["status"] == "PASS"
    assert compared["download_filename_match"] is True
    assert compared["legacy_download_filename"] == "report.pdf"
    assert compared["new_download_filename"] == "report.pdf"


def test_file_download_compare_uses_download_filename_fields(tmp_path):
    mapping_path = tmp_path / "page_mapping.json"
    mapping_path.write_text(json.dumps({"page_mappings": []}), encoding="utf-8")
    engine = RegressionEngine(mapping_path=str(mapping_path), output_dir=str(tmp_path))

    compared = engine._compare_state(
        "Download.jsp",
        "High",
        "file output",
        {"url": "about:closed", "dom": "", "screenshot": ""},
        {"url": "about:closed", "dom": "", "screenshot": ""},
        tmp_path / "diff.png",
        action_type="file_download",
        legacy_action={"status": "PASS", "saved_filename": "output.pdf", "download_path": str(tmp_path / "output.pdf")},
        new_action={"status": "PASS", "saved_filename": "output.pdf", "download_path": str(tmp_path / "output.pdf")},
    )

    assert compared["status"] == "PASS"
    assert compared["visual"]["status"] == "SKIPPED"
    assert compared["download_filename_match"] is True
    assert compared["legacy_download_filename"] == "output.pdf"


def test_render_report_contains_semantic_diagnostics(tmp_path):
    mapping_path = tmp_path / "page_mapping.json"
    mapping_path.write_text(json.dumps({"page_mappings": []}), encoding="utf-8")
    output_dir = tmp_path / "out"
    output_dir.mkdir()
    screenshot = output_dir / "shot.png"
    Image.new("RGB", (1, 1), "white").save(screenshot)

    engine = RegressionEngine(mapping_path=str(mapping_path), output_dir=str(output_dir))
    report = engine.render_report(
        [
            {
                "page_id": "FramePage.jsp",
                "risk": "High",
                "action": "submitForm",
                "action_type": "submit",
                "status": "BLOCKED",
                "url_match": False,
                "dom_match": False,
                "legacy_locator": "[name='LegacyForm']",
                "new_locator": "[name='NewForm']",
                "legacy_url": "https://legacy.example/app/FramePage.jsp",
                "new_url": "https://new.example/app/FramePage.jsp",
                "legacy_screenshot": str(screenshot),
                "new_screenshot": str(screenshot),
                "diff_screenshot": str(screenshot),
                "visual": {"diff_percent": 0.0},
                "legacy_frame": {"name": "frEditFrame", "url": "https://legacy.example/app/inner.do"},
                "new_action": {"reason": "Timeout waiting for locator", "selector_found": False},
            }
        ]
    )

    html = Path(report).read_text(encoding="utf-8")
    assert "<b>Action type</b><span>submit</span>" in html
    assert "frEditFrame" in html
    assert "new_locator" in html
    assert "Timeout waiting for locator" in html


def test_capture_state_handles_closed_page(tmp_path):
    class ClosedPage:
        def is_closed(self):
            return True

    state = _capture_state(ClosedPage(), tmp_path, "closed")

    assert state["page_closed"] is True
    assert state["url"] == "about:closed"
    assert Path(state["screenshot"]).exists()


def test_compare_state_treats_matching_closed_pages_as_pass(tmp_path):
    mapping_path = tmp_path / "page_mapping.json"
    mapping_path.write_text(json.dumps({"page_mappings": []}), encoding="utf-8")
    output_dir = tmp_path / "out"
    output_dir.mkdir()

    class ClosedPage:
        def is_closed(self):
            return True

    legacy_state = _capture_state(ClosedPage(), output_dir, "legacy_closed")
    new_state = _capture_state(ClosedPage(), output_dir, "new_closed")
    engine = RegressionEngine(mapping_path=str(mapping_path), output_dir=str(output_dir))

    result = engine._compare_state(
        "ClosePage.jsp",
        "High",
        "cancel",
        legacy_state,
        new_state,
        output_dir / "diff.png",
        legacy_action={"status": "PASS", "page_closed_after_action": True},
        new_action={"status": "PASS", "page_closed_after_action": True},
    )

    assert result["status"] == "PASS"
    assert result["url_match"] is True
    assert result["dom_match"] is True


def test_missing_dynamic_jsp_row_control_is_skipped_by_static_mapping():
    assert RegressionEngine._is_dynamic_jsp_row_control(
        {
            "kind": "button",
            "key": "field:deleteFileName",
            "label": "deleteFileName",
            "locator": "input[type=\"button\"][onclick=\"deleteFile('<bean:write name=\"]",
        }
    )


def test_compare_state_treats_table_data_variance_as_warn(tmp_path):
    mapping_path = tmp_path / "page_mapping.json"
    mapping_path.write_text(json.dumps({"page_mappings": []}), encoding="utf-8")
    output_dir = tmp_path / "out"
    output_dir.mkdir()

    legacy_image = output_dir / "legacy.png"
    new_image = output_dir / "new.png"
    Image.new("RGB", (10, 10), "white").save(legacy_image)
    image = Image.new("RGB", (10, 10), "white")
    image.putpixel((0, 0), (0, 0, 0))
    image.save(new_image)

    shared_text = "\n".join(
        [
            "社内分類更新処理待ちファイル一覧(国内)",
            "アップロード日時",
            "更新処理待ちファイル",
            "削除",
            "2026年04月24日 UopcSampleDataJp.csv",
        ]
    )
    legacy_state = {
        "screenshot": str(legacy_image),
        "url": "http://legacy.example/patlics/UopcUploadListDispJP.do",
        "dom": "<table><tr><td>legacy rows</td></tr></table>",
        "text": shared_text + "\nlegacy-only-row",
    }
    new_state = {
        "screenshot": str(new_image),
        "url": "http://new.example/patlics/UopcUploadListDispJP.do",
        "dom": "<table><tr><td>new rows</td></tr></table>",
        "text": shared_text + "\nnew-only-row",
    }
    engine = RegressionEngine(mapping_path=str(mapping_path), output_dir=str(output_dir), visual_threshold_percent=0.1)

    result = engine._compare_state(
        "UopcUploadListDispJP.jsp",
        "High",
        "initial display",
        legacy_state,
        new_state,
        output_dir / "diff.png",
        action_type="initial_display",
    )

    assert result["status"] == "WARN"
    assert result["visual"]["data_variance_tolerated"] is True


def test_database_operation_compares_each_side_before_after(tmp_path):
    mapping_path = tmp_path / "page_mapping.json"
    mapping_path.write_text(json.dumps({"page_mappings": []}), encoding="utf-8")
    output_dir = tmp_path / "out"
    output_dir.mkdir()

    legacy_before = output_dir / "legacy_before.png"
    legacy_after = output_dir / "legacy_after.png"
    new_before = output_dir / "new_before.png"
    new_after = output_dir / "new_after.png"
    Image.new("RGB", (10, 10), "white").save(legacy_before)
    Image.new("RGB", (10, 10), "white").save(new_before)
    legacy_changed = Image.new("RGB", (10, 10), "white")
    legacy_changed.putpixel((0, 0), (0, 0, 0))
    legacy_changed.save(legacy_after)
    new_changed = Image.new("RGB", (10, 10), "white")
    new_changed.putpixel((0, 0), (0, 0, 0))
    new_changed.save(new_after)

    engine = RegressionEngine(mapping_path=str(mapping_path), output_dir=str(output_dir), visual_threshold_percent=0.1)
    result = engine._compare_database_operation(
        "UopcUploadListDispJP.jsp",
        "High",
        "削除ボタン動作確認",
        "delete_action",
        "delete",
        {"screenshot": str(legacy_before), "url": "http://legacy/app/List.do"},
        {"screenshot": str(new_before), "url": "http://new/app/List.do"},
        {"status": "PASS", "state": {"screenshot": str(legacy_after), "url": "http://legacy/app/List.do"}},
        {"status": "PASS", "state": {"screenshot": str(new_after), "url": "http://new/app/List.do"}},
        output_dir,
        "delete_case",
    )

    assert result["status"] == "PASS"
    assert result["comparison_mode"] == "database_operation_before_after"
    assert result["database_operation"] == "delete"
    assert result["legacy_delta"]["status"] == "DIFF"
    assert result["new_delta"]["status"] == "DIFF"


def test_database_operation_kind_covers_crud_actions():
    assert RegressionEngine._database_operation_kind({"case_type": "create_action", "label": "登録"}, "click", "click") == "create"
    assert RegressionEngine._database_operation_kind({"case_type": "update_action", "label": "更新"}, "click", "click") == "update"
    assert RegressionEngine._database_operation_kind({"case_type": "delete_action", "label": "削除"}, "click", "click") == "delete"
    assert RegressionEngine._database_operation_kind({"case_type": "search_normal", "label": "検索"}, "search", "click") == "read"
    assert RegressionEngine._database_operation_kind({"case_type": "upload_submit", "label": "アップロード確認"}, "upload_submit", "upload") is None


def test_leaving_actions_require_target_reopen():
    assert RegressionEngine._requires_target_reopen_after_action(
        {"case_type": "back_action", "label": "戻る"},
        "back_action",
        "click",
        {"status": "PASS"},
        {"status": "PASS"},
    )
    assert RegressionEngine._requires_target_reopen_after_action(
        {"case_type": "close_window", "locator": "input[onclick*='window.close']"},
        "close_window",
        "click",
        {"status": "PASS"},
        {"status": "PASS"},
    )
    assert RegressionEngine._requires_target_reopen_after_action(
        {"case_type": "click", "label": "キャンセル"},
        "click",
        "click",
        {"status": "PASS"},
        {"status": "PASS"},
    )
    assert RegressionEngine._requires_target_reopen_after_action(
        {"case_type": "click", "label": "next"},
        "click",
        "click",
        {"status": "PASS", "page_closed_after_action": True},
        {"status": "PASS"},
    )
    assert RegressionEngine._requires_target_reopen_after_action(
        {"case_type": "negative_file_upload", "label": "invalid upload"},
        "negative_file_upload",
        "upload",
        {"status": "PASS"},
        {"status": "PASS"},
    )
    assert not RegressionEngine._requires_target_reopen_after_action(
        {"case_type": "delete_action", "label": "削除"},
        "delete_action",
        "click",
        {"status": "PASS"},
        {"status": "PASS"},
    )


def test_reopen_prefers_open_route_step_page_before_login_recovery(tmp_path):
    class Context:
        def __init__(self):
            self.pages = []

    class Page:
        def __init__(self, url, *, closed=False):
            self.url = url
            self._closed = closed
            self.frames = []
            self.context = None
            self.front = 0

        def is_closed(self):
            return self._closed

        def bring_to_front(self):
            self.front += 1

        def wait_for_load_state(self, state, timeout=None):
            return None

    context = Context()
    closed_popup = Page("about:closed", closed=True)
    parent = Page("http://example.test/patlics/JpBiblioListForEasySearch.do")
    login = Page("http://example.test/patlics/login.jsp")
    for page in (closed_popup, parent, login):
        page.context = context
    context.pages = [closed_popup, login, parent]

    mapping_path = tmp_path / "page_mapping.json"
    mapping_path.write_text(json.dumps({"page_mappings": []}), encoding="utf-8")
    engine = RegressionEngine(mapping_path=str(mapping_path), output_dir=str(tmp_path))

    selected, nav = engine._takeover_recovery_page(
        closed_popup,
        {"steps": [{"active_page_url": "http://example.test/patlics/JpBiblioListForEasySearch.do"}]},
    )

    assert selected is parent
    assert nav["status"] == "PASS"
    assert nav["strategy"] == "context_sibling_takeover"
    assert nav["route_step_detected"] is True
    assert parent.front == 1


def test_closing_action_report_uses_reopened_target_state(tmp_path, monkeypatch):
    class Page:
        def __init__(self, url):
            self.url = url

    mapping_path = tmp_path / "page_mapping.json"
    mapping_path.write_text(json.dumps({"page_mappings": []}), encoding="utf-8")
    engine = RegressionEngine(mapping_path=str(mapping_path), output_dir=str(tmp_path))

    def fake_capture_state(page, output_dir, name):
        return {
            "screenshot": str(Path(output_dir) / f"{name}.png"),
            "url": getattr(page, "url", ""),
            "dom": name,
            "text": name,
        }

    def fake_compare_state(page_id, risk, action, legacy_state, new_state, diff_path, **extra):
        return {
            "page_id": page_id,
            "risk": risk,
            "action": action,
            "status": "PASS",
            "legacy_screenshot": legacy_state.get("screenshot"),
            "new_screenshot": new_state.get("screenshot"),
            **extra,
        }

    monkeypatch.setattr(regression_engine_module, "_capture_state", fake_capture_state)
    engine._build_action_plan = lambda page_id, mapping: (
        [
            {
                "case_type": "file_download",
                "label": "download",
                "legacy_locator": "#download",
                "new_locator": "#download",
            }
        ],
        "test",
    )
    engine._execute_action_case = lambda page, action_case, **kwargs: {
        "status": "PASS",
        "page_closed_after_action": True,
        "state": {
            "screenshot": str(tmp_path / "closed.png"),
            "url": "about:closed",
            "dom": "<page-closed />",
            "page_closed": True,
        },
    }
    engine._reopen_target_pair = lambda legacy_page, new_page, mapping, page_dir, browser_name, *, reason: (
        Page("http://legacy/target.jsp"),
        Page("http://new/target.jsp"),
        {"status": "PASS", "reason": reason},
    )
    engine._compare_state = fake_compare_state

    results = engine._run_captured_page_pair(
        Page("http://legacy/start.jsp"),
        Page("http://new/start.jsp"),
        {"page_id": "Download.jsp", "risk": "Low"},
        tmp_path,
        "chrome",
        {"status": "PASS"},
        {"status": "PASS"},
        manual=False,
    )

    action_result = next(item for item in results if item["action"] == "download")
    assert action_result["legacy_screenshot"].endswith("01_download_legacy_after_reopen.png")
    assert action_result["new_screenshot"].endswith("01_download_new_after_reopen.png")
    assert action_result["legacy_action"]["state_before_reopen"]["page_closed"] is True
    assert action_result["post_action_reopen"]["status"] == "PASS"
    log_text = (tmp_path / "full_test_log.jsonl").read_text(encoding="utf-8")
    assert '"event": "action_executed"' in log_text
    assert '"event": "target_reopen_finished"' in log_text
    assert '"event": "compare_result"' in log_text
